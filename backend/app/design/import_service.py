"""设计预约 — 批量导入"""

import logging
from datetime import date, datetime
from io import BytesIO
from typing import Optional
from urllib.parse import quote

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from fastapi.responses import StreamingResponse
from sqlalchemy import and_, func, Integer, case
from sqlalchemy.orm import Session

from app.design.models import (
    DesignScheduleRequest,
    DesignScheduleTask,
    DesignDesigner,
    DesignUnavailableDate,
    DesignCapacityConfig,
    DesignAuditLog,
)
from app.design.state_machine import RequestStatus, OperatorRole, transition
from app.design.conflict_engine import check_conflict, get_scheduling_mode as _get_mode
from app.design.utils import generate_request_no, generate_task_no
from app.system.service import get_label_map as _get_dict_map
from app.design.schemas import (
    DesignRequestCreate,
    DesignRequestAudit,
    DesignRequestAction,
    TaskReschedule,
    UnavailableDateCreate,
    CapacityUpdate,
    ModeUpdate,
)
from app.design.audit_log import write_audit_log as _write_audit_log

logger = __import__("logging").getLogger("design")


def batch_import_requests(
    db: Session,
    file_bytes: bytes,
    operator_id: int,
    operator_name: str,
    operator_role: str,
) -> dict:
    """从 Excel 批量导入预约申请"""
    from openpyxl import load_workbook
    from io import BytesIO as _BytesIO

    wb = load_workbook(filename=_BytesIO(file_bytes), read_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(min_row=1, values_only=True))
    if not rows:
        return {"code": 400, "message": "Excel 文件为空", "data": None}

    # Expected header: 客户名称, 业务员姓名, 拍摄类型, 期望开始日期, 期望结束日期, 优先级, 备注
    header = [str(c).strip() if c else "" for c in rows[0]]

    shoot_type_map = _get_dict_map(db, "shoot_type")
    shoot_type_reverse = {v: k for k, v in shoot_type_map.items()}
    valid_codes = set(shoot_type_map.keys())
    PRIORITY_REVERSE = {"普通": "normal", "加急": "urgent"}

    total = 0
    success = 0
    failed = 0
    errors = []

    for row_idx, row in enumerate(rows[1:], start=2):
        total += 1
        try:
            if not row or not row[0]:
                raise ValueError("客户名称为空")

            customer_name = str(row[0]).strip()
            salesperson_name = str(row[1]).strip() if row[1] else ""
            shoot_type_raw = str(row[2]).strip() if row[2] else ""
            parts = [p.strip() for p in shoot_type_raw.replace("、", ",").split(",") if p.strip()]
            codes = [shoot_type_reverse.get(p, p) for p in parts]
            if any(c not in valid_codes for c in codes):
                invalid = [p for p in parts if shoot_type_reverse.get(p, p) not in valid_codes]
                raise ValueError(f"无效的拍摄类型: {'、'.join(invalid)}")
            shoot_type = ",".join(codes)

            # Parse dates (handle both string and date objects)
            expect_start = row[3]
            expect_end = row[4]
            if isinstance(expect_start, str):
                expect_start = date.fromisoformat(expect_start.strip())
            elif isinstance(expect_start, datetime):
                expect_start = expect_start.date()
            elif not isinstance(expect_start, date):
                raise ValueError("期望开始日期格式错误")

            if isinstance(expect_end, str):
                expect_end = date.fromisoformat(expect_end.strip())
            elif isinstance(expect_end, datetime):
                expect_end = expect_end.date()
            elif not isinstance(expect_end, date):
                raise ValueError("期望结束日期格式错误")

            priority_raw = str(row[5]).strip() if row[5] else "普通"
            priority = PRIORITY_REVERSE.get(priority_raw, "normal")

            remark = str(row[6]).strip() if len(row) > 6 and row[6] else None

            # Conflict check
            conflict = check_conflict(db, expect_start, expect_end, start_period="am", end_period="pm")
            batch_status = "pending_audit" if conflict["has_conflict"] else "pending_design"

            request_no = generate_request_no(db)
            req = DesignScheduleRequest(
                request_no=request_no,
                customer_name=customer_name,
                salesperson_id=operator_id,
                salesperson_name=salesperson_name or operator_name,
                shoot_type=shoot_type,
                expect_start_date=expect_start,
                expect_start_period="am",
                expect_end_date=expect_end,
                expect_end_period="pm",
                priority=priority,
                remark=remark,
                status=batch_status,
                conflict_detail=conflict if conflict["has_conflict"] else None,
            )
            db.add(req)
            db.flush()

            _write_audit_log(
                db,
                request_id=req.id,
                operator_id=operator_id,
                operator_name=operator_name,
                operator_role=operator_role,
                action="submit",
                to_status=batch_status,
                comment=f"批量导入 (行{row_idx})",
            )
            success += 1

        except Exception as e:
            failed += 1
            errors.append({"row": row_idx, "reason": str(e)})

    db.commit()
    wb.close()

    return {
        "code": 200,
        "message": f"导入完成: 成功 {success}, 失败 {failed}",
        "data": {
            "total": total,
            "success": success,
            "failed": failed,
            "errors": errors,
        },
    }
