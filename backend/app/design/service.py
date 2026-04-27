"""设计预约 — 业务逻辑层"""

import logging
from datetime import date, datetime
from io import BytesIO
from typing import Optional
from urllib.parse import quote

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from fastapi.responses import StreamingResponse
from sqlalchemy import and_, func, Integer, case
from sqlalchemy.orm import Session

from app.design.models import (
    DesignScheduleRequest,
    DesignScheduleTask,
    DesignDesigner,
    DesignUnavailableDate,
    DesignCapacityConfig,
    DesignAuditLog,
)
from app.design.state_machine import RequestStatus, OperatorRole, transition
from app.design.conflict_engine import check_conflict, get_scheduling_mode as _get_mode
from app.design.utils import generate_request_no, generate_task_no
from app.design.schemas import (
    DesignRequestCreate,
    DesignRequestAudit,
    DesignRequestAction,
    TaskReschedule,
    UnavailableDateCreate,
    CapacityUpdate,
    ModeUpdate,
)

logger = logging.getLogger("design")


# ── 辅助 ──────────────────────────────────────────────────


def _write_audit_log(
    db: Session,
    *,
    request_id: int,
    task_id: Optional[int] = None,
    operator_id: int,
    operator_name: str,
    operator_role: str,
    action: str,
    from_status: Optional[str] = None,
    to_status: Optional[str] = None,
    comment: Optional[str] = None,
    snapshot: Optional[dict] = None,
):
    log = DesignAuditLog(
        request_id=request_id,
        task_id=task_id,
        operator_id=operator_id,
        operator_name=operator_name,
        operator_role=operator_role,
        action=action,
        from_status=from_status,
        to_status=to_status,
        comment=comment,
        snapshot=snapshot,
    )
    db.add(log)


# ── 申请单 ────────────────────────────────────────────────


def create_request(
    db: Session,
    data: DesignRequestCreate,
    operator_id: int,
    operator_name: str,
    operator_role: str,
) -> dict:
    """创建设计预约申请"""
    # 冲突检测
    conflict = check_conflict(db, data.expect_start_date, data.expect_end_date)

    request_no = generate_request_no(db)

    req = DesignScheduleRequest(
        request_no=request_no,
        customer_name=data.customer_name,
        salesperson_id=data.salesperson_id,
        salesperson_name=data.salesperson_name,
        shoot_type=data.shoot_type,
        shoot_type_remark=data.shoot_type_remark,
        expect_start_date=data.expect_start_date,
        expect_end_date=data.expect_end_date,
        priority=data.priority,
        remark=data.remark,
        status="pending_audit",
        conflict_detail=conflict if conflict["has_conflict"] else None,
    )
    db.add(req)
    db.flush()

    _write_audit_log(
        db,
        request_id=req.id,
        operator_id=operator_id,
        operator_name=operator_name,
        operator_role=operator_role,
        action="submit",
        to_status="pending_audit",
        comment=data.remark,
    )
    db.commit()
    db.refresh(req)

    return {
        "code": 200,
        "message": "申请已提交",
        "data": {
            "id": req.id,
            "request_no": req.request_no,
            "status": req.status,
            "conflict": conflict,
        },
    }


def audit_request(
    db: Session,
    request_id: int,
    data: DesignRequestAudit,
    operator_id: int,
    operator_name: str,
) -> dict:
    """主管审批（approve / reject）"""
    req = db.query(DesignScheduleRequest).filter(
        DesignScheduleRequest.id == request_id,
        DesignScheduleRequest.deleted_at.is_(None),
    ).first()
    if not req:
        raise ValueError("申请单不存在")

    from_status = req.status
    new_status = transition(
        RequestStatus(from_status),
        data.action,
        OperatorRole.supervisor,
    )
    req.status = new_status.value
    req.updated_at = datetime.now()

    _write_audit_log(
        db,
        request_id=req.id,
        operator_id=operator_id,
        operator_name=operator_name,
        operator_role="supervisor",
        action=data.action,
        from_status=from_status,
        to_status=new_status.value,
        comment=data.comment,
    )
    db.commit()
    db.refresh(req)

    return {
        "code": 200,
        "message": f"审批操作 {data.action} 完成",
        "data": {"id": req.id, "status": req.status},
    }


def action_request(
    db: Session,
    request_id: int,
    data: DesignRequestAction,
    operator_id: int,
    operator_name: str,
    operator_role: str,
) -> dict:
    """执行申请单动作：confirm / start / complete / cancel"""
    req = db.query(DesignScheduleRequest).filter(
        DesignScheduleRequest.id == request_id,
        DesignScheduleRequest.deleted_at.is_(None),
    ).first()
    if not req:
        raise ValueError("申请单不存在")

    from_status = req.status
    new_status = transition(
        RequestStatus(from_status),
        data.action,
        OperatorRole(operator_role),
    )
    req.status = new_status.value
    req.updated_at = datetime.now()

    task_id = None

    if data.action == "confirm":
        # 确认排期 → 创建任务
        if not data.designer_id:
            raise ValueError("确认排期需要指定设计师")
        task_no = generate_task_no(db)
        task = DesignScheduleTask(
            request_id=req.id,
            task_no=task_no,
            designer_id=data.designer_id,
            task_name=f"{req.customer_name} - {req.shoot_type}",
            customer_name=req.customer_name,
            salesperson_name=req.salesperson_name,
            shoot_type=req.shoot_type,
            priority=req.priority,
            plan_start_date=data.plan_start_date or req.expect_start_date,
            plan_end_date=data.plan_end_date or req.expect_end_date,
            status="scheduled",
        )
        db.add(task)
        db.flush()
        task_id = task.id
        req.assigned_designer_id = data.designer_id

    elif data.action == "start":
        req.actual_start_date = date.today()
        # 同步更新相关任务
        for t in req.tasks:
            if t.status == "scheduled":
                t.status = "in_progress"
                t.actual_start_date = date.today()
                t.updated_at = datetime.now()

    elif data.action == "complete":
        req.actual_end_date = date.today()
        for t in req.tasks:
            if t.status == "in_progress":
                t.status = "completed"
                t.actual_end_date = date.today()
                t.updated_at = datetime.now()

    elif data.action == "cancel":
        for t in req.tasks:
            if t.status not in ("completed", "cancelled"):
                t.status = "cancelled"
                t.updated_at = datetime.now()

    _write_audit_log(
        db,
        request_id=req.id,
        task_id=task_id,
        operator_id=operator_id,
        operator_name=operator_name,
        operator_role=operator_role,
        action=data.action,
        from_status=from_status,
        to_status=new_status.value,
        comment=data.comment,
    )
    db.commit()
    db.refresh(req)

    return {
        "code": 200,
        "message": f"操作 {data.action} 完成",
        "data": {"id": req.id, "status": req.status},
    }


# ── 甘特图 ────────────────────────────────────────────────


def get_gantt_data(
    db: Session,
    start_date: date,
    end_date: date,
    designer_id: Optional[int] = None,
) -> dict:
    """获取甘特图数据"""
    designer_query = db.query(DesignDesigner).filter(DesignDesigner.is_active == True)
    if designer_id:
        designer_query = designer_query.filter(DesignDesigner.id == designer_id)
    designers = designer_query.all()

    result = []
    for d in designers:
        tasks = db.query(DesignScheduleTask).filter(
            DesignScheduleTask.designer_id == d.id,
            DesignScheduleTask.status.in_(["scheduled", "in_progress", "completed"]),
            DesignScheduleTask.plan_start_date <= end_date,
            DesignScheduleTask.plan_end_date >= start_date,
        ).all()
        result.append({
            "id": d.id,
            "name": d.name,
            "tasks": [
                {
                    "id": t.id,
                    "request_id": t.request_id,
                    "task_no": t.task_no,
                    "designer_id": t.designer_id,
                    "task_name": t.task_name,
                    "customer_name": t.customer_name,
                    "salesperson_name": t.salesperson_name,
                    "shoot_type": t.shoot_type,
                    "priority": t.priority,
                    "plan_start_date": t.plan_start_date.isoformat() if t.plan_start_date else None,
                    "plan_end_date": t.plan_end_date.isoformat() if t.plan_end_date else None,
                    "actual_start_date": t.actual_start_date.isoformat() if t.actual_start_date else None,
                    "actual_end_date": t.actual_end_date.isoformat() if t.actual_end_date else None,
                    "status": t.status,
                    "remark": t.remark,
                    "created_at": t.created_at.isoformat() if t.created_at else None,
                    "updated_at": t.updated_at.isoformat() if t.updated_at else None,
                }
                for t in tasks
            ],
        })

    # 获取不可用日期
    unavailable = db.query(DesignUnavailableDate).filter(
        DesignUnavailableDate.date >= start_date,
        DesignUnavailableDate.date <= end_date,
    ).all()
    unavailable_dates = [row.date.isoformat() for row in unavailable]

    return {
        "code": 200,
        "message": "ok",
        "data": {
            "designers": result,
            "unavailable_dates": unavailable_dates,
        },
    }


# ── 任务改期 ──────────────────────────────────────────────


def reschedule_task(
    db: Session,
    task_id: int,
    data: TaskReschedule,
    operator_id: int,
    operator_name: str,
) -> dict:
    """改期任务"""
    task = db.query(DesignScheduleTask).filter(
        DesignScheduleTask.id == task_id,
    ).first()
    if not task:
        raise ValueError("任务不存在")

    if task.status in ("completed", "cancelled"):
        raise ValueError(f"任务状态 {task.status} 不允许改期")

    old_start = task.plan_start_date
    old_end = task.plan_end_date
    old_designer = task.designer_id

    task.plan_start_date = data.plan_start_date
    task.plan_end_date = data.plan_end_date
    if data.designer_id is not None:
        task.designer_id = data.designer_id
    task.updated_at = datetime.now()

    _write_audit_log(
        db,
        request_id=task.request_id,
        task_id=task.id,
        operator_id=operator_id,
        operator_name=operator_name,
        operator_role="design_staff",
        action="reschedule",
        from_status=task.status,
        to_status=task.status,
        comment=data.comment,
        snapshot={
            "old_start": old_start.isoformat() if old_start else None,
            "old_end": old_end.isoformat() if old_end else None,
            "old_designer_id": old_designer,
            "new_start": data.plan_start_date.isoformat(),
            "new_end": data.plan_end_date.isoformat(),
            "new_designer_id": data.designer_id,
        },
    )
    db.commit()
    db.refresh(task)

    return {
        "code": 200,
        "message": "改期成功",
        "data": {"task_id": task.id, "task_no": task.task_no},
    }


# ── 不可用日期 ────────────────────────────────────────────


def create_unavailable_dates(
    db: Session,
    data: UnavailableDateCreate,
    operator_id: int,
    operator_name: str,
) -> dict:
    """批量设置不可用日期"""
    created = []
    for d in data.dates:
        existing = db.query(DesignUnavailableDate).filter(
            DesignUnavailableDate.date == d,
        ).first()
        if existing:
            continue
        row = DesignUnavailableDate(
            date=d,
            reason=data.reason,
            created_by=operator_id,
        )
        db.add(row)
        created.append(d.isoformat())

    if created:
        _write_audit_log(
            db,
            request_id=0,
            operator_id=operator_id,
            operator_name=operator_name,
            operator_role="design_staff",
            action="set_unavailable",
            comment=f"设置不可用日期: {', '.join(created)}; 原因: {data.reason}",
        )
    db.commit()

    return {
        "code": 200,
        "message": f"已设置 {len(created)} 个不可用日期",
        "data": {"dates": created},
    }


def delete_unavailable_date(
    db: Session,
    date_str: str,
    operator_id: int,
    operator_name: str,
) -> dict:
    """删除不可用日期"""
    from datetime import date as date_type
    try:
        target = date_type.fromisoformat(date_str)
    except ValueError:
        raise ValueError("日期格式错误，应为 YYYY-MM-DD")

    row = db.query(DesignUnavailableDate).filter(
        DesignUnavailableDate.date == target,
    ).first()
    if not row:
        raise ValueError("该日期未设置为不可用")

    db.delete(row)
    _write_audit_log(
        db,
        request_id=0,
        operator_id=operator_id,
        operator_name=operator_name,
        operator_role="design_staff",
        action="remove_unavailable",
        comment=f"移除不可用日期: {date_str}",
    )
    db.commit()

    return {"code": 200, "message": "已移除不可用日期", "data": {"date": date_str}}


# ── 容量配置 ──────────────────────────────────────────────


def get_capacity(db: Session) -> list:
    """获取所有容量配置"""
    rows = db.query(DesignCapacityConfig).all()
    return [
        {
            "id": r.id,
            "config_date": r.config_date.isoformat() if r.config_date else None,
            "designer_id": r.designer_id,
            "max_parallel_tasks": r.max_parallel_tasks,
            "scheduling_mode": r.scheduling_mode,
            "updated_by": r.updated_by,
            "updated_at": r.updated_at.isoformat() if r.updated_at else None,
        }
        for r in rows
    ]


def update_capacity(
    db: Session,
    data: CapacityUpdate,
    operator_id: int,
    operator_name: str,
) -> dict:
    """更新容量配置"""
    for entry in data.entries:
        existing = db.query(DesignCapacityConfig).filter(
            DesignCapacityConfig.config_date == entry.config_date if entry.config_date else DesignCapacityConfig.config_date.is_(None),
            DesignCapacityConfig.designer_id == entry.designer_id if entry.designer_id else DesignCapacityConfig.designer_id.is_(None),
        ).first()
        if existing:
            existing.max_parallel_tasks = entry.max_parallel_tasks
            existing.updated_by = operator_id
            existing.updated_at = datetime.now()
        else:
            row = DesignCapacityConfig(
                config_date=entry.config_date,
                designer_id=entry.designer_id,
                max_parallel_tasks=entry.max_parallel_tasks,
                updated_by=operator_id,
            )
            db.add(row)

    _write_audit_log(
        db,
        request_id=0,
        operator_id=operator_id,
        operator_name=operator_name,
        operator_role="design_staff",
        action="update_capacity",
        comment=f"更新容量配置 {len(data.entries)} 条",
    )
    db.commit()

    return {"code": 200, "message": "容量配置已更新", "data": None}


# ── 排期模式 ──────────────────────────────────────────────


def get_scheduling_mode_info(db: Session) -> dict:
    """获取当前排期模式"""
    mode = _get_mode(db)
    return {"code": 200, "message": "ok", "data": {"scheduling_mode": mode}}


def update_scheduling_mode(
    db: Session,
    data: ModeUpdate,
    operator_id: int,
    operator_name: str,
) -> dict:
    """更新排期模式"""
    row = db.query(DesignCapacityConfig).filter(
        DesignCapacityConfig.config_date.is_(None),
        DesignCapacityConfig.designer_id.is_(None),
    ).first()
    if row:
        row.scheduling_mode = data.scheduling_mode
        row.updated_by = operator_id
        row.updated_at = datetime.now()
    else:
        row = DesignCapacityConfig(
            config_date=None,
            designer_id=None,
            max_parallel_tasks=3,
            scheduling_mode=data.scheduling_mode,
            updated_by=operator_id,
        )
        db.add(row)

    _write_audit_log(
        db,
        request_id=0,
        operator_id=operator_id,
        operator_name=operator_name,
        operator_role="design_staff",
        action="update_capacity",
        comment=f"切换排期模式为 {data.scheduling_mode}",
    )
    db.commit()

    return {"code": 200, "message": f"排期模式已切换为 {data.scheduling_mode}", "data": None}


# ── 导出 Excel ────────────────────────────────────────────

HEADER_FONT = Font(bold=True)
HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center")

SHOOT_TYPE_LABELS = {
    "product_photo": "产品图",
    "model_photo": "模特图",
    "video": "视频",
    "product_video": "产品视频",
    "other": "其他",
}

PRIORITY_LABELS = {"normal": "普通", "urgent": "加急"}

TASK_STATUS_LABELS = {
    "pending_design": "待排期",
    "scheduled": "已排期",
    "in_progress": "进行中",
    "completed": "已完成",
    "cancelled": "已取消",
}


def _to_streaming(wb: Workbook, filename: str) -> StreamingResponse:
    """将 Workbook 转为 StreamingResponse"""
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    encoded = quote(filename, safe="")
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{encoded}",
        },
    )


def export_tasks_excel(
    db: Session,
    start_date: date,
    end_date: date,
) -> StreamingResponse:
    """导出任务列表为 Excel"""
    tasks = db.query(DesignScheduleTask).filter(
        DesignScheduleTask.plan_start_date <= end_date,
        DesignScheduleTask.plan_end_date >= start_date,
    ).order_by(DesignScheduleTask.plan_start_date).all()

    # Fetch designer names
    designers = db.query(DesignDesigner).all()
    designer_map = {d.id: d.name for d in designers}

    columns = [
        "任务编号", "客户名称", "业务员", "设计师", "拍摄类型",
        "优先级", "计划开始", "计划结束", "实际开始", "实际结束", "状态",
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "设计任务"

    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN

    for row_idx, t in enumerate(tasks, 2):
        ws.cell(row=row_idx, column=1, value=t.task_no)
        ws.cell(row=row_idx, column=2, value=t.customer_name or "")
        ws.cell(row=row_idx, column=3, value=t.salesperson_name or "")
        ws.cell(row=row_idx, column=4, value=designer_map.get(t.designer_id, ""))
        ws.cell(row=row_idx, column=5, value=SHOOT_TYPE_LABELS.get(t.shoot_type, t.shoot_type or ""))
        ws.cell(row=row_idx, column=6, value=PRIORITY_LABELS.get(t.priority, t.priority or ""))
        ws.cell(row=row_idx, column=7, value=t.plan_start_date.isoformat() if t.plan_start_date else "")
        ws.cell(row=row_idx, column=8, value=t.plan_end_date.isoformat() if t.plan_end_date else "")
        ws.cell(row=row_idx, column=9, value=t.actual_start_date.isoformat() if t.actual_start_date else "")
        ws.cell(row=row_idx, column=10, value=t.actual_end_date.isoformat() if t.actual_end_date else "")
        ws.cell(row=row_idx, column=11, value=TASK_STATUS_LABELS.get(t.status, t.status or ""))

    # Auto-width columns
    for col in ws.columns:
        max_len = 0
        for cell in col:
            val = str(cell.value) if cell.value else ""
            max_len = max(max_len, len(val))
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 30)

    filename = f"design_tasks_{start_date.isoformat()}_{end_date.isoformat()}.xlsx"
    return _to_streaming(wb, filename)


# ── 统计报表 ──────────────────────────────────────────────


def get_design_stats(
    db: Session,
    start_date: date,
    end_date: date,
) -> dict:
    """获取设计模块统计数据"""
    base_query = db.query(DesignScheduleTask).filter(
        DesignScheduleTask.plan_start_date <= end_date,
        DesignScheduleTask.plan_end_date >= start_date,
    )

    total = base_query.count()
    completed = base_query.filter(DesignScheduleTask.status == "completed").count()
    in_progress = base_query.filter(DesignScheduleTask.status == "in_progress").count()
    scheduled = base_query.filter(DesignScheduleTask.status == "scheduled").count()

    # Designer-level stats
    designer_rows = db.query(
        DesignScheduleTask.designer_id,
        func.count(DesignScheduleTask.id).label("total"),
        func.sum(
            case((DesignScheduleTask.status == "completed", 1), else_=0)
        ).label("completed"),
        func.sum(
            case((DesignScheduleTask.status == "in_progress", 1), else_=0)
        ).label("in_progress"),
    ).filter(
        DesignScheduleTask.plan_start_date <= end_date,
        DesignScheduleTask.plan_end_date >= start_date,
    ).group_by(DesignScheduleTask.designer_id).all()

    # Fetch designer names
    designers = db.query(DesignDesigner).all()
    designer_map = {d.id: d.name for d in designers}

    # Calculate avg_duration for completed tasks per designer
    completed_tasks = db.query(DesignScheduleTask).filter(
        DesignScheduleTask.plan_start_date <= end_date,
        DesignScheduleTask.plan_end_date >= start_date,
        DesignScheduleTask.status == "completed",
        DesignScheduleTask.actual_start_date.isnot(None),
        DesignScheduleTask.actual_end_date.isnot(None),
    ).all()

    duration_sums = {}
    duration_counts = {}
    for t in completed_tasks:
        did = t.designer_id
        days = (t.actual_end_date - t.actual_start_date).days + 1
        duration_sums[did] = duration_sums.get(did, 0) + days
        duration_counts[did] = duration_counts.get(did, 0) + 1

    designer_stats = []
    for row in designer_rows:
        did = row.designer_id
        avg_days = round(duration_sums.get(did, 0) / duration_counts[did], 1) if duration_counts.get(did) else 0
        designer_stats.append({
            "designer_id": did,
            "designer_name": designer_map.get(did, str(did)),
            "total": row.total,
            "completed": row.completed or 0,
            "in_progress": row.in_progress or 0,
            "avg_duration_days": avg_days,
        })

    return {
        "code": 200,
        "message": "ok",
        "data": {
            "summary": {
                "total": total,
                "completed": completed,
                "in_progress": in_progress,
                "scheduled": scheduled,
            },
            "designer_stats": designer_stats,
        },
    }


# ── 批量导入 ──────────────────────────────────────────────


def batch_import_requests(
    db: Session,
    file_bytes: bytes,
    operator_id: int,
    operator_name: str,
    operator_role: str,
) -> dict:
    """从 Excel 批量导入预约申请"""
    from openpyxl import load_workbook
    from io import BytesIO as _BytesIO

    wb = load_workbook(filename=_BytesIO(file_bytes), read_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(min_row=1, values_only=True))
    if not rows:
        return {"code": 400, "message": "Excel 文件为空", "data": None}

    # Expected header: 客户名称, 业务员姓名, 拍摄类型, 期望开始日期, 期望结束日期, 优先级, 备注
    header = [str(c).strip() if c else "" for c in rows[0]]

    SHOOT_TYPE_REVERSE = {v: k for k, v in SHOOT_TYPE_LABELS.items()}
    PRIORITY_REVERSE = {"普通": "normal", "加急": "urgent"}

    total = 0
    success = 0
    failed = 0
    errors = []

    for row_idx, row in enumerate(rows[1:], start=2):
        total += 1
        try:
            if not row or not row[0]:
                raise ValueError("客户名称为空")

            customer_name = str(row[0]).strip()
            salesperson_name = str(row[1]).strip() if row[1] else ""
            shoot_type_raw = str(row[2]).strip() if row[2] else ""
            shoot_type = SHOOT_TYPE_REVERSE.get(shoot_type_raw, shoot_type_raw)
            if shoot_type not in ("product_photo", "model_photo", "video", "product_video", "other"):
                raise ValueError(f"无效的拍摄类型: {shoot_type_raw}")

            # Parse dates (handle both string and date objects)
            expect_start = row[3]
            expect_end = row[4]
            if isinstance(expect_start, str):
                expect_start = date.fromisoformat(expect_start.strip())
            elif isinstance(expect_start, datetime):
                expect_start = expect_start.date()
            elif not isinstance(expect_start, date):
                raise ValueError("期望开始日期格式错误")

            if isinstance(expect_end, str):
                expect_end = date.fromisoformat(expect_end.strip())
            elif isinstance(expect_end, datetime):
                expect_end = expect_end.date()
            elif not isinstance(expect_end, date):
                raise ValueError("期望结束日期格式错误")

            priority_raw = str(row[5]).strip() if row[5] else "普通"
            priority = PRIORITY_REVERSE.get(priority_raw, "normal")

            remark = str(row[6]).strip() if len(row) > 6 and row[6] else None

            # Conflict check
            conflict = check_conflict(db, expect_start, expect_end)

            request_no = generate_request_no(db)
            req = DesignScheduleRequest(
                request_no=request_no,
                customer_name=customer_name,
                salesperson_id=operator_id,
                salesperson_name=salesperson_name or operator_name,
                shoot_type=shoot_type,
                expect_start_date=expect_start,
                expect_end_date=expect_end,
                priority=priority,
                remark=remark,
                status="pending_audit",
                conflict_detail=conflict if conflict["has_conflict"] else None,
            )
            db.add(req)
            db.flush()

            _write_audit_log(
                db,
                request_id=req.id,
                operator_id=operator_id,
                operator_name=operator_name,
                operator_role=operator_role,
                action="submit",
                to_status="pending_audit",
                comment=f"批量导入 (行{row_idx})",
            )
            success += 1

        except Exception as e:
            failed += 1
            errors.append({"row": row_idx, "reason": str(e)})

    db.commit()
    wb.close()

    return {
        "code": 200,
        "message": f"导入完成: 成功 {success}, 失败 {failed}",
        "data": {
            "total": total,
            "success": success,
            "failed": failed,
            "errors": errors,
        },
    }
