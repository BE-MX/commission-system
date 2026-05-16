"""设计预约 — 统计 / 导出 Excel"""

import logging
from datetime import date, datetime
from io import BytesIO
from typing import Optional
from urllib.parse import quote

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from fastapi.responses import StreamingResponse
from sqlalchemy import and_, func, Integer, case
from sqlalchemy.orm import Session

from app.design.models import (
    DesignScheduleRequest,
    DesignScheduleTask,
    DesignDesigner,
    DesignUnavailableDate,
    DesignCapacityConfig,
    DesignAuditLog,
)
from app.design.state_machine import RequestStatus, OperatorRole, transition
from app.design.conflict_engine import check_conflict, get_scheduling_mode as _get_mode
from app.design.utils import generate_request_no, generate_task_no
from app.system.service import get_label_map as _get_dict_map
from app.design.schemas import (
    DesignRequestCreate,
    DesignRequestAudit,
    DesignRequestAction,
    TaskReschedule,
    UnavailableDateCreate,
    CapacityUpdate,
    ModeUpdate,
)

logger = __import__("logging").getLogger("design")


def _to_streaming(wb: Workbook, filename: str) -> StreamingResponse:
    """将 Workbook 转为 StreamingResponse"""
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    encoded = quote(filename, safe="")
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{encoded}",
        },
    )


def export_tasks_excel(
    db: Session,
    start_date: date,
    end_date: date,
) -> StreamingResponse:
    """导出任务列表为 Excel"""
    tasks = db.query(DesignScheduleTask).filter(
        DesignScheduleTask.plan_start_date <= end_date,
        DesignScheduleTask.plan_end_date >= start_date,
    ).order_by(DesignScheduleTask.plan_start_date).all()

    # Fetch designer names
    designers = db.query(DesignDesigner).all()
    designer_map = {d.id: d.name for d in designers}

    PERIOD_LABELS = {"am": "上午", "pm": "下午"}

    shoot_type_map = _get_dict_map(db, "shoot_type")
    def _fmt_shoot_type(raw: str | None) -> str:
        if not raw:
            return ""
        codes = [c.strip() for c in raw.split(",") if c.strip()]
        return "、".join(shoot_type_map.get(c, c) for c in codes)

    columns = [
        "任务编号", "客户名称", "业务员", "设计师", "拍摄类型",
        "优先级", "计划开始", "开始时段", "计划结束", "结束时段",
        "实际开始", "实际结束", "状态",
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "设计任务"

    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN

    for row_idx, t in enumerate(tasks, 2):
        ws.cell(row=row_idx, column=1, value=t.task_no)
        ws.cell(row=row_idx, column=2, value=t.customer_name or "")
        ws.cell(row=row_idx, column=3, value=t.salesperson_name or "")
        ws.cell(row=row_idx, column=4, value=designer_map.get(t.designer_id, ""))
        ws.cell(row=row_idx, column=5, value=_fmt_shoot_type(t.shoot_type))
        ws.cell(row=row_idx, column=6, value=PRIORITY_LABELS.get(t.priority, t.priority or ""))
        ws.cell(row=row_idx, column=7, value=t.plan_start_date.isoformat() if t.plan_start_date else "")
        ws.cell(row=row_idx, column=8, value=PERIOD_LABELS.get(t.plan_start_period, "全天"))
        ws.cell(row=row_idx, column=9, value=t.plan_end_date.isoformat() if t.plan_end_date else "")
        ws.cell(row=row_idx, column=10, value=PERIOD_LABELS.get(t.plan_end_period, "全天"))
        ws.cell(row=row_idx, column=11, value=t.actual_start_date.isoformat() if t.actual_start_date else "")
        ws.cell(row=row_idx, column=12, value=t.actual_end_date.isoformat() if t.actual_end_date else "")
        ws.cell(row=row_idx, column=13, value=TASK_STATUS_LABELS.get(t.status, t.status or ""))

    # Auto-width columns
    for col in ws.columns:
        max_len = 0
        for cell in col:
            val = str(cell.value) if cell.value else ""
            max_len = max(max_len, len(val))
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 30)

    filename = f"design_tasks_{start_date.isoformat()}_{end_date.isoformat()}.xlsx"
    return _to_streaming(wb, filename)


def get_design_stats(
    db: Session,
    start_date: date,
    end_date: date,
) -> dict:
    """获取设计模块统计数据"""
    base_query = db.query(DesignScheduleTask).filter(
        DesignScheduleTask.plan_start_date <= end_date,
        DesignScheduleTask.plan_end_date >= start_date,
    )

    total = base_query.count()
    completed = base_query.filter(DesignScheduleTask.status == "completed").count()
    in_progress = base_query.filter(DesignScheduleTask.status == "in_progress").count()
    scheduled = base_query.filter(DesignScheduleTask.status == "scheduled").count()

    # Designer-level stats
    designer_rows = db.query(
        DesignScheduleTask.designer_id,
        func.count(DesignScheduleTask.id).label("total"),
        func.sum(
            case((DesignScheduleTask.status == "completed", 1), else_=0)
        ).label("completed"),
        func.sum(
            case((DesignScheduleTask.status == "in_progress", 1), else_=0)
        ).label("in_progress"),
    ).filter(
        DesignScheduleTask.plan_start_date <= end_date,
        DesignScheduleTask.plan_end_date >= start_date,
    ).group_by(DesignScheduleTask.designer_id).all()

    # Fetch designer names
    designers = db.query(DesignDesigner).all()
    designer_map = {d.id: d.name for d in designers}

    # Calculate avg_duration for completed tasks per designer
    completed_tasks = db.query(DesignScheduleTask).filter(
        DesignScheduleTask.plan_start_date <= end_date,
        DesignScheduleTask.plan_end_date >= start_date,
        DesignScheduleTask.status == "completed",
        DesignScheduleTask.actual_start_date.isnot(None),
        DesignScheduleTask.actual_end_date.isnot(None),
    ).all()

    duration_sums = {}
    duration_counts = {}
    for t in completed_tasks:
        did = t.designer_id
        days = (t.actual_end_date - t.actual_start_date).days + 1
        duration_sums[did] = duration_sums.get(did, 0) + days
        duration_counts[did] = duration_counts.get(did, 0) + 1

    designer_stats = []
    for row in designer_rows:
        did = row.designer_id
        avg_days = round(duration_sums.get(did, 0) / duration_counts[did], 1) if duration_counts.get(did) else 0
        designer_stats.append({
            "designer_id": did,
            "designer_name": designer_map.get(did, str(did)),
            "total": row.total,
            "completed": row.completed or 0,
            "in_progress": row.in_progress or 0,
            "avg_duration_days": avg_days,
        })

    return {
        "code": 200,
        "message": "ok",
        "data": {
            "summary": {
                "total": total,
                "completed": completed,
                "in_progress": in_progress,
                "scheduled": scheduled,
            },
            "designer_stats": designer_stats,
        },
    }
