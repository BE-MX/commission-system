"""员工属性管理路由"""

import tempfile
from datetime import date

from fastapi import APIRouter, Depends, UploadFile, File, Query
from openpyxl import load_workbook
from sqlalchemy import or_
from sqlalchemy.orm import Session

from app.api.deps import get_db
from app.models.employee import EmployeeAttributeHistory
from app.models.business import UserBasic
from app.schemas.common import ResponseModel, PageResponse
from app.schemas.employee import (
    EmployeeListItem, EmployeeAttributeRequest,
    EmployeeAttributeHistoryItem, EmployeeAttributeResult, EmployeeImportResult,
    EmployeeAttributeBatchHistoryRequest,
)

router = APIRouter()


def _set_employee_attribute(
    db: Session, employee_id: str, attribute_type: str,
    effective_date: date = None,
) -> str:
    """设置/变更员工属性，返回操作类型"""
    current = db.query(EmployeeAttributeHistory).filter(
        EmployeeAttributeHistory.employee_id == employee_id,
        EmployeeAttributeHistory.is_current == True,
    ).first()

    if current and current.attribute_type == attribute_type:
        return "skipped"

    eff_date = effective_date or date.today()

    if current:
        current.is_current = False
        current.effective_end = eff_date

    new_record = EmployeeAttributeHistory(
        employee_id=employee_id,
        attribute_type=attribute_type,
        effective_start=eff_date,
        is_current=True,
    )
    db.add(new_record)
    return "updated" if current else "created"


@router.get("/list", summary="查询员工列表（含当前属性）")
def list_employees(
    keyword: str = Query("", description="搜索姓名/ID"),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    db: Session = Depends(get_db),
) -> ResponseModel[PageResponse[EmployeeListItem]]:
    """查询员工列表，LEFT JOIN 当前属性"""
    query = db.query(
        UserBasic.user_id,
        UserBasic.full_name,
        UserBasic.nickname,
        EmployeeAttributeHistory.attribute_type.label("current_attribute"),
    ).outerjoin(
        EmployeeAttributeHistory,
        (UserBasic.user_id == EmployeeAttributeHistory.employee_id)
        & (EmployeeAttributeHistory.is_current == True),
    )

    if keyword:
        like_pattern = f"%{keyword}%"
        query = query.filter(or_(
            UserBasic.user_id.like(like_pattern),
            UserBasic.full_name.like(like_pattern),
            UserBasic.nickname.like(like_pattern),
        ))

    total = query.count()
    rows = query.offset((page - 1) * page_size).limit(page_size).all()

    items = [
        EmployeeListItem(
            user_id=r.user_id,
            full_name=r.full_name,
            nickname=r.nickname,
            current_attribute=r.current_attribute,
        )
        for r in rows
    ]
    page_data = PageResponse(items=items, total=total, page=page, page_size=page_size)
    return ResponseModel(data=page_data)


@router.post("/attribute", summary="设置/变更员工属性")
def set_employee_attribute(
    req: EmployeeAttributeRequest,
    db: Session = Depends(get_db),
) -> ResponseModel[EmployeeAttributeResult]:
    """设置或变更员工属性（开发/分配）"""
    user = db.query(UserBasic).filter(UserBasic.user_id == req.employee_id).first()
    if not user:
        return ResponseModel(code=404, message=f"员工 {req.employee_id} 不存在")

    action = _set_employee_attribute(db, req.employee_id, req.attribute_type, req.effective_date)
    db.commit()

    return ResponseModel(data=EmployeeAttributeResult(
        employee_id=req.employee_id,
        attribute_type=req.attribute_type,
        action=action,
    ))


@router.get("/attribute/history", summary="查询员工属性变更历史")
def get_attribute_history(
    employee_id: str = Query(..., description="员工ID"),
    db: Session = Depends(get_db),
) -> ResponseModel[list[EmployeeAttributeHistoryItem]]:
    """查询指定员工的属性变更历史"""
    records = (
        db.query(EmployeeAttributeHistory)
        .filter(EmployeeAttributeHistory.employee_id == employee_id)
        .order_by(EmployeeAttributeHistory.effective_start.desc())
        .all()
    )
    items = [EmployeeAttributeHistoryItem.model_validate(r) for r in records]
    return ResponseModel(data=items)


@router.post("/attribute/import", summary="批量导入员工属性（Excel）")
def import_employee_attributes(
    file: UploadFile = File(..., description="Excel文件"),
    db: Session = Depends(get_db),
) -> ResponseModel[EmployeeImportResult]:
    """从 Excel 批量导入员工属性。模板列：员工ID(user_id) | 属性(开发/分配)"""
    attr_map = {"开发": "develop", "分配": "distribute"}
    result = EmployeeImportResult(total_rows=0, success=0, failed=0)

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp.write(file.file.read())
        tmp_path = tmp.name

    wb = load_workbook(tmp_path, data_only=True)
    ws = wb.active

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or not row[0]:
            break
        result.total_rows += 1

        try:
            emp_id = str(row[0]).strip()
            attr_raw = str(row[1]).strip()
            attr_type = attr_map.get(attr_raw, attr_raw)

            if attr_type not in ("develop", "distribute"):
                raise ValueError(f"属性无效: '{attr_raw}'，应为 开发/分配")

            user = db.query(UserBasic).filter(UserBasic.user_id == emp_id).first()
            if not user:
                raise ValueError(f"员工 {emp_id} 在业务库中不存在")

            _set_employee_attribute(db, emp_id, attr_type)
            result.success += 1
        except Exception as e:
            result.failures.append(f"第 {row_idx} 行: {e}")
            result.failed += 1

    wb.close()
    db.commit()
    return ResponseModel(data=result)


@router.post("/attribute/batch-history", summary="批量设置员工属性历史记录")
def batch_set_attribute_history(
    items: list[EmployeeAttributeBatchHistoryRequest],
    db: Session = Depends(get_db),
) -> ResponseModel[dict]:
    """
    批量写入员工属性历史记录。

    会先清除指定员工的所有历史记录，再按时间顺序重建。
    最后一条记录自动标记为 is_current=True。
    """
    from app.services.customer_reset_service import refresh_snapshots_by_employees

    affected_employee_ids = []

    for item in items:
        emp_id = item.employee_id
        user = db.query(UserBasic).filter(UserBasic.user_id == emp_id).first()
        if not user:
            return ResponseModel(code=404, message=f"员工 {emp_id} 不存在")

        if not item.records:
            continue

        sorted_records = sorted(item.records, key=lambda r: r.effective_start)

        db.query(EmployeeAttributeHistory).filter(
            EmployeeAttributeHistory.employee_id == emp_id,
        ).delete(synchronize_session="fetch")

        for idx, rec in enumerate(sorted_records):
            is_last = idx == len(sorted_records) - 1
            effective_end = sorted_records[idx + 1].effective_start if not is_last else None

            new_record = EmployeeAttributeHistory(
                employee_id=emp_id,
                attribute_type=rec.attribute_type,
                effective_start=rec.effective_start,
                effective_end=effective_end,
                is_current=is_last,
            )
            db.add(new_record)

        affected_employee_ids.append(emp_id)

    db.flush()

    refresh_result = refresh_snapshots_by_employees(db, affected_employee_ids, operator="api")

    db.commit()

    return ResponseModel(
        message=f"已更新 {len(affected_employee_ids)} 位员工的属性历史，"
                f"刷新快照 {refresh_result.updated} 条",
        data={
            "employees_updated": len(affected_employee_ids),
            "snapshots_refreshed": refresh_result.updated,
            "snapshots_skipped": refresh_result.skipped,
        },
    )
