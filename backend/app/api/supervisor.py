"""主管关系管理路由"""

import tempfile
from datetime import date

from fastapi import APIRouter, Depends, UploadFile, File, Query
from openpyxl import load_workbook
from sqlalchemy import or_
from sqlalchemy.orm import Session, aliased

from app.api.deps import get_db
from app.models.employee import SupervisorRelationHistory
from app.models.business import UserBasic
from app.schemas.common import ResponseModel, PageResponse
from app.schemas.supervisor import (
    SupervisorRelationListItem, SupervisorRelationRequest,
    SupervisorHistoryItem, SupervisorRelationResult, SupervisorImportResult,
)

router = APIRouter()


def _set_supervisor_relation(
    db: Session, salesperson_id: str, supervisor_id: str,
) -> str:
    """设置/变更主管关系，返回操作类型"""
    current = db.query(SupervisorRelationHistory).filter(
        SupervisorRelationHistory.salesperson_id == salesperson_id,
        SupervisorRelationHistory.is_current == True,
    ).first()

    if current and current.supervisor_id == supervisor_id:
        return "skipped"

    today = date.today()

    if current:
        current.is_current = False
        current.effective_end = today

    new_record = SupervisorRelationHistory(
        salesperson_id=salesperson_id,
        supervisor_id=supervisor_id,
        effective_start=today,
        is_current=True,
    )
    db.add(new_record)
    return "updated" if current else "created"


@router.get("/list", summary="查询主管关系列表")
def list_supervisor_relations(
    keyword: str = Query("", description="搜索业务员姓名/ID"),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    db: Session = Depends(get_db),
) -> ResponseModel[PageResponse[SupervisorRelationListItem]]:
    """查询当前有效的主管关系列表"""
    SpUser = aliased(UserBasic)
    SvUser = aliased(UserBasic)

    query = db.query(
        SupervisorRelationHistory.salesperson_id,
        SpUser.full_name.label("salesperson_name"),
        SupervisorRelationHistory.supervisor_id,
        SvUser.full_name.label("supervisor_name"),
        SupervisorRelationHistory.effective_start,
    ).join(
        SpUser, SupervisorRelationHistory.salesperson_id == SpUser.user_id,
    ).outerjoin(
        SvUser, SupervisorRelationHistory.supervisor_id == SvUser.user_id,
    ).filter(
        SupervisorRelationHistory.is_current == True,
    )

    if keyword:
        like_pattern = f"%{keyword}%"
        query = query.filter(or_(
            SupervisorRelationHistory.salesperson_id.like(like_pattern),
            SpUser.full_name.like(like_pattern),
            SpUser.nickname.like(like_pattern),
        ))

    total = query.count()
    rows = query.offset((page - 1) * page_size).limit(page_size).all()

    items = [
        SupervisorRelationListItem(
            salesperson_id=r.salesperson_id,
            salesperson_name=r.salesperson_name,
            supervisor_id=r.supervisor_id,
            supervisor_name=r.supervisor_name,
            effective_start=r.effective_start,
        )
        for r in rows
    ]
    page_data = PageResponse(items=items, total=total, page=page, page_size=page_size)
    return ResponseModel(data=page_data)


@router.post("/relation", summary="设置/变更主管关系")
def set_supervisor_relation(
    req: SupervisorRelationRequest,
    db: Session = Depends(get_db),
) -> ResponseModel[SupervisorRelationResult]:
    """设置或变更业务员的主管关系"""
    sp_user = db.query(UserBasic).filter(UserBasic.user_id == req.salesperson_id).first()
    if not sp_user:
        return ResponseModel(code=404, message=f"业务员 {req.salesperson_id} 不存在")

    sv_user = db.query(UserBasic).filter(UserBasic.user_id == req.supervisor_id).first()
    if not sv_user:
        return ResponseModel(code=404, message=f"主管 {req.supervisor_id} 不存在")

    action = _set_supervisor_relation(db, req.salesperson_id, req.supervisor_id)
    db.commit()

    return ResponseModel(data=SupervisorRelationResult(
        salesperson_id=req.salesperson_id,
        supervisor_id=req.supervisor_id,
        action=action,
    ))


@router.get("/history", summary="查询主管关系变更历史")
def get_supervisor_history(
    salesperson_id: str = Query(..., description="业务员ID"),
    db: Session = Depends(get_db),
) -> ResponseModel[list[SupervisorHistoryItem]]:
    """查询指定业务员的主管关系变更历史"""
    records = (
        db.query(SupervisorRelationHistory)
        .filter(SupervisorRelationHistory.salesperson_id == salesperson_id)
        .order_by(SupervisorRelationHistory.effective_start.desc())
        .all()
    )
    items = [SupervisorHistoryItem.model_validate(r) for r in records]
    return ResponseModel(data=items)


@router.post("/import", summary="批量导入主管关系（Excel）")
def import_supervisor_relations(
    file: UploadFile = File(..., description="Excel文件"),
    db: Session = Depends(get_db),
) -> ResponseModel[SupervisorImportResult]:
    """从 Excel 批量导入主管关系。模板列：业务员ID(user_id) | 主管ID(user_id)"""
    result = SupervisorImportResult(total_rows=0, success=0, failed=0)

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp.write(file.file.read())
        tmp_path = tmp.name

    wb = load_workbook(tmp_path, data_only=True)
    ws = wb.active

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or not row[0]:
            break
        result.total_rows += 1

        try:
            sp_id = str(row[0]).strip()
            sv_id = str(row[1]).strip()

            sp_user = db.query(UserBasic).filter(UserBasic.user_id == sp_id).first()
            if not sp_user:
                raise ValueError(f"业务员 {sp_id} 在业务库中不存在")

            sv_user = db.query(UserBasic).filter(UserBasic.user_id == sv_id).first()
            if not sv_user:
                raise ValueError(f"主管 {sv_id} 在业务库中不存在")

            _set_supervisor_relation(db, sp_id, sv_id)
            result.success += 1
        except Exception as e:
            result.failures.append(f"第 {row_idx} 行: {e}")
            result.failed += 1

    wb.close()
    db.commit()
    return ResponseModel(data=result)
