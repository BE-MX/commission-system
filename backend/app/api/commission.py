"""提成计算路由"""

import json
from calendar import monthrange
from datetime import date, datetime
from decimal import Decimal
from io import BytesIO
from urllib.parse import quote

from fastapi import APIRouter, Depends, Query, Path, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy import func, or_, desc
from sqlalchemy.orm import Session, aliased

from app.api.deps import get_db
from app.auth.dependencies import require_any_permission, require_permission
from app.auth.models import ArkUser, ArkUserExternalBinding
from app.models.commission import (
    CommissionBatch, CommissionDetail, SyncedPayment, PaymentCommissionStatus,
    CommissionBatchFeedback, CommissionBatchConfirmation,
)
from app.models.customer import CustomerCommissionSnapshot
from app.models.business import UserBasic, CustomerInfo, OkkiOrder, OkkiReceipt
from app.schemas.common import ResponseModel, PageResponse
from app.schemas.commission import (
    CommissionBatchCreateRequest, CommissionBatchListItem,
    CommissionDetailListItem, CommissionCalcResponse,
    CommissionConfirmRequest, CommissionSendConfirmRequest, CommissionBatchSummary,
    SalesCommissionBatchListItem, SalesCommissionBatchDetail,
    SalesCommissionSummary, SalesCommissionMonthlySummary,
    SalesCommissionDetailItem, SalesCommissionFeedbackRequest,
    SalesCommissionConfirmRequest,
    CommissionConfirmationProgress,
)
from app.services.commission_calculator import (
    calculate_commission, void_batch, confirm_batch, send_confirm, revoke_confirm,
)

router = APIRouter()

SELF_VISIBLE_BATCH_STATUSES = ("confirming", "confirmed")
CONFIRMATION_TEXT = "我已确认"
ORDER_SOURCE_FIELD_ID = "45285192666116"


def _to_float(value) -> float:
    return float(value or 0)


def _display_exchange_rate(value) -> float:
    """业务员侧按需求展示 exchange_rate/100。"""
    if value is None:
        return 0
    return float(Decimal(str(value)) / Decimal("100"))


def _commission_rmb(commission_usd, exchange_rate) -> float:
    return round(_to_float(commission_usd) * _display_exchange_rate(exchange_rate), 2)


def _order_source_from_custom_fields(custom_fields: str | None) -> str | None:
    if not custom_fields:
        return None
    try:
        data = json.loads(custom_fields)
    except (TypeError, ValueError):
        return None
    if not isinstance(data, dict):
        return None
    value = data.get(ORDER_SOURCE_FIELD_ID)
    return str(value).strip() if value else None


def _month_bounds(month: str) -> tuple[date, date]:
    try:
        year, month_num = [int(part) for part in month.split("-", 1)]
        last_day = monthrange(year, month_num)[1]
        return date(year, month_num, 1), date(year, month_num, last_day)
    except Exception as exc:
        raise HTTPException(status_code=400, detail="月份格式应为 YYYY-MM") from exc


def _current_business_context(db: Session, current_user: dict) -> tuple[ArkUser | None, list[str]]:
    ark_user = None
    sub = current_user.get("sub")
    if sub:
        try:
            ark_user = db.query(ArkUser).filter(ArkUser.id == int(sub)).first()
        except (TypeError, ValueError):
            ark_user = None

    candidate_ids = set()
    if ark_user:
        binding_ids = [
            row.external_account_id
            for row in db.query(ArkUserExternalBinding.external_account_id).filter(
                ArkUserExternalBinding.ark_user_id == ark_user.id,
                ArkUserExternalBinding.binding_status == "active",
                ArkUserExternalBinding.deleted_at.is_(None),
                ArkUserExternalBinding.external_account_id.isnot(None),
            ).all()
        ]
        candidate_ids.update(binding_ids)

        filters = []
        if candidate_ids:
            filters.append(UserBasic.user_id.in_(candidate_ids))
        if ark_user.username:
            filters.append(UserBasic.user_id == ark_user.username)
        if ark_user.real_name:
            filters.append(UserBasic.full_name == ark_user.real_name)
            filters.append(UserBasic.nickname == ark_user.real_name)
        if ark_user.phone:
            filters.append(UserBasic.user_mobile == ark_user.phone)

        if filters:
            matched = db.query(UserBasic.user_id).filter(or_(*filters)).all()
            candidate_ids.update(row.user_id for row in matched)

    return ark_user, sorted(uid for uid in candidate_ids if uid)


def _self_role_condition(user_ids: list[str], role: str = ""):
    if not user_ids:
        return CommissionDetail.id == None
    if role == "salesperson":
        return CommissionDetail.salesperson_id.in_(user_ids)
    if role == "supervisor":
        return CommissionDetail.supervisor_id.in_(user_ids)
    if role == "second_supervisor":
        return CommissionDetail.second_supervisor_id.in_(user_ids)
    return or_(
        CommissionDetail.salesperson_id.in_(user_ids),
        CommissionDetail.supervisor_id.in_(user_ids),
        CommissionDetail.second_supervisor_id.in_(user_ids),
    )


def _related_roles(detail: CommissionDetail, user_ids: list[str]) -> list[str]:
    roles = []
    user_id_set = set(user_ids)
    if detail.salesperson_id in user_id_set:
        roles.append("salesperson")
    if detail.supervisor_id in user_id_set:
        roles.append("supervisor")
    if detail.second_supervisor_id in user_id_set:
        roles.append("second_supervisor")
    return roles


def _role_commissions(detail: CommissionDetail, user_ids: list[str]) -> tuple[float, float, float]:
    user_id_set = set(user_ids)
    sp = _to_float(detail.salesperson_commission) if detail.salesperson_id in user_id_set else 0
    sv = _to_float(detail.supervisor_commission) if detail.supervisor_id in user_id_set else 0
    sv2 = _to_float(detail.second_supervisor_commission) if detail.second_supervisor_id in user_id_set else 0
    return sp, sv, sv2


def _parse_business_user_ids(value: str | None) -> set[str]:
    if not value:
        return set()
    return {item.strip() for item in value.split(",") if item.strip()}


def _batch_expected_business_user_ids(db: Session, batch_id: int) -> set[str]:
    rows = db.query(
        CommissionDetail.salesperson_id,
        CommissionDetail.supervisor_id,
        CommissionDetail.second_supervisor_id,
    ).filter(
        CommissionDetail.batch_id == batch_id,
        CommissionDetail.status != "voided",
    ).all()

    user_ids = set()
    for salesperson_id, supervisor_id, second_supervisor_id in rows:
        for user_id in (salesperson_id, supervisor_id, second_supervisor_id):
            if user_id:
                user_ids.add(user_id)
    return user_ids


def _confirmed_business_user_ids(db: Session, batch_id: int) -> set[str]:
    rows = db.query(CommissionBatchConfirmation.business_user_ids).filter(
        CommissionBatchConfirmation.batch_id == batch_id,
        CommissionBatchConfirmation.status == "confirmed",
    ).all()
    confirmed_ids = set()
    for row in rows:
        confirmed_ids.update(_parse_business_user_ids(row.business_user_ids))
    return confirmed_ids


def _confirmation_progress(db: Session, batch_id: int) -> CommissionConfirmationProgress:
    expected_ids = _batch_expected_business_user_ids(db, batch_id)
    confirmed_ids = _confirmed_business_user_ids(db, batch_id) & expected_ids
    feedback_count = db.query(func.count(CommissionBatchFeedback.id)).filter(
        CommissionBatchFeedback.batch_id == batch_id,
    ).scalar() or 0

    expected_count = len(expected_ids)
    confirmed_count = len(confirmed_ids)
    pending_count = max(expected_count - confirmed_count, 0)
    if expected_count == 0:
        status = "not_required"
    elif confirmed_count == 0:
        status = "not_started"
    elif pending_count == 0:
        status = "all_confirmed"
    else:
        status = "partial_confirmed"

    return CommissionConfirmationProgress(
        expected_confirm_count=expected_count,
        confirmed_count=confirmed_count,
        pending_confirm_count=pending_count,
        feedback_count=feedback_count,
        confirmation_status=status,
    )


def _batch_list_item(db: Session, batch: CommissionBatch) -> CommissionBatchListItem:
    progress = _confirmation_progress(db, batch.id)
    return CommissionBatchListItem(
        id=batch.id,
        batch_name=batch.batch_name,
        period_type=batch.period_type,
        period_start=batch.period_start,
        period_end=batch.period_end,
        status=batch.status,
        created_at=batch.created_at,
        created_by=batch.created_by,
        confirmed_at=batch.confirmed_at,
        confirmed_by=batch.confirmed_by,
        expected_confirm_count=progress.expected_confirm_count,
        confirmed_count=progress.confirmed_count,
        pending_confirm_count=progress.pending_confirm_count,
        feedback_count=progress.feedback_count,
        confirmation_status=progress.confirmation_status,
    )


def _my_confirmation(db: Session, batch_id: int, ark_user_id: str, user_ids: list[str]) -> CommissionBatchConfirmation | None:
    rows = db.query(CommissionBatchConfirmation).filter(
        CommissionBatchConfirmation.batch_id == batch_id,
        CommissionBatchConfirmation.status == "confirmed",
    ).all()
    user_id_set = set(user_ids)
    for row in rows:
        if row.ark_user_id == ark_user_id:
            return row
        if _parse_business_user_ids(row.business_user_ids) & user_id_set:
            return row
    return None


def _query_self_detail_rows(db: Session, batch_id: int, user_ids: list[str], role: str = ""):
    SpUser = aliased(UserBasic)
    SvUser = aliased(UserBasic)
    Sv2User = aliased(UserBasic)

    return db.query(
        CommissionDetail,
        CustomerInfo.company_name.label("customer_name"),
        CustomerInfo.country_name.label("customer_country"),
        SpUser.full_name.label("salesperson_name"),
        SvUser.full_name.label("supervisor_name"),
        Sv2User.full_name.label("second_supervisor_name"),
        SyncedPayment.payment_date.label("payment_date"),
        SyncedPayment.exchange_rate.label("payment_exchange_rate"),
        OkkiReceipt.collection_date.label("collection_date"),
        OkkiReceipt.type.label("receipt_type"),
        OkkiReceipt.exchange_rate.label("receipt_exchange_rate"),
        OkkiReceipt.order_no.label("order_no"),
        OkkiOrder.name.label("order_name"),
        SyncedPayment.service_fee.label("payment_service_fee"),
        OkkiReceipt.service_fee_amount_usd.label("receipt_service_fee"),
        OkkiOrder.custom_fields.label("order_custom_fields"),
    ).outerjoin(
        CustomerInfo, CommissionDetail.customer_id == CustomerInfo.company_id,
    ).outerjoin(
        SpUser, CommissionDetail.salesperson_id == SpUser.user_id,
    ).outerjoin(
        SvUser, CommissionDetail.supervisor_id == SvUser.user_id,
    ).outerjoin(
        Sv2User, CommissionDetail.second_supervisor_id == Sv2User.user_id,
    ).outerjoin(
        SyncedPayment, CommissionDetail.payment_id == SyncedPayment.payment_id,
    ).outerjoin(
        OkkiReceipt, CommissionDetail.payment_id == OkkiReceipt.cash_collection_id,
    ).outerjoin(
        OkkiOrder, CommissionDetail.order_id == OkkiOrder.order_id,
    ).filter(
        CommissionDetail.batch_id == batch_id,
        CommissionDetail.status != "voided",
        _self_role_condition(user_ids, role),
    ).order_by(
        func.coalesce(OkkiReceipt.collection_date, SyncedPayment.payment_date),
        CommissionDetail.id,
    ).all()


def _detail_schema_from_row(row) -> SalesCommissionDetailItem:
    (
        detail, customer_name, customer_country, salesperson_name, supervisor_name,
        second_supervisor_name, payment_date, payment_exchange_rate, collection_date,
        receipt_type, receipt_exchange_rate, order_no, order_name, payment_service_fee,
        receipt_service_fee, order_custom_fields,
    ) = row
    exchange_source = payment_exchange_rate if payment_exchange_rate is not None else receipt_exchange_rate
    return SalesCommissionDetailItem(
        id=detail.id,
        payment_id=detail.payment_id,
        order_id=detail.order_id,
        order_no=order_no,
        order_name=order_name,
        customer_id=detail.customer_id,
        customer_name=customer_name,
        customer_country=customer_country,
        collection_date=collection_date or payment_date,
        type=receipt_type,
        payment_amount=_to_float(detail.payment_amount),
        service_fee=_to_float(payment_service_fee if payment_service_fee is not None else receipt_service_fee),
        order_source=_order_source_from_custom_fields(order_custom_fields),
        exchange_rate=_display_exchange_rate(exchange_source),
        salesperson_id=detail.salesperson_id,
        salesperson_name=salesperson_name,
        salesperson_rate=_to_float(detail.salesperson_rate),
        salesperson_commission=_to_float(detail.salesperson_commission),
        salesperson_commission_rmb=_commission_rmb(detail.salesperson_commission, exchange_source),
        supervisor_id=detail.supervisor_id,
        supervisor_name=supervisor_name,
        supervisor_rate=_to_float(detail.supervisor_rate),
        supervisor_commission=_to_float(detail.supervisor_commission),
        supervisor_commission_rmb=_commission_rmb(detail.supervisor_commission, exchange_source),
        second_supervisor_id=detail.second_supervisor_id,
        second_supervisor_name=second_supervisor_name,
        second_supervisor_rate=_to_float(detail.second_supervisor_rate),
        second_supervisor_commission=_to_float(detail.second_supervisor_commission),
        second_supervisor_commission_rmb=_commission_rmb(detail.second_supervisor_commission, exchange_source),
        calc_rule_note=detail.calc_rule_note,
        status=detail.status,
    )


def _build_sales_batch_item(
    batch: CommissionBatch,
    rows,
    user_ids: list[str],
    my_confirmation: CommissionBatchConfirmation | None = None,
) -> SalesCommissionBatchListItem:
    role_set = set()
    payment_total = 0
    sp_total = 0
    sv_total = 0
    sv2_total = 0
    rmb_total = 0

    for row in rows:
        detail = row[0]
        exchange_source = row[7] if row[7] is not None else row[10]
        role_set.update(_related_roles(detail, user_ids))
        payment_total += _to_float(detail.payment_amount)
        sp, sv, sv2 = _role_commissions(detail, user_ids)
        sp_total += sp
        sv_total += sv
        sv2_total += sv2
        rmb_total += _commission_rmb(sp + sv + sv2, exchange_source)

    return SalesCommissionBatchListItem(
        id=batch.id,
        batch_name=batch.batch_name,
        period_start=batch.period_start,
        period_end=batch.period_end,
        status=batch.status,
        related_roles=sorted(role_set),
        total_payment_amount=round(payment_total, 2),
        total_salesperson_commission=round(sp_total, 2),
        total_supervisor_commission=round(sv_total, 2),
        total_second_supervisor_commission=round(sv2_total, 2),
        total_commission=round(sp_total + sv_total + sv2_total, 2),
        total_commission_rmb=round(rmb_total, 2),
        detail_count=len(rows),
        created_at=batch.created_at,
        confirmed_at=batch.confirmed_at,
        is_confirmed_by_me=my_confirmation is not None,
        my_confirmed_at=my_confirmation.confirmed_at if my_confirmation else None,
    )


def _build_sales_summary(rows, user_ids: list[str]) -> SalesCommissionSummary:
    item = _build_sales_batch_item(
        CommissionBatch(
            id=0,
            batch_name="",
            period_start=date.today(),
            period_end=date.today(),
            status="confirming",
        ),
        rows,
        user_ids,
    )
    return SalesCommissionSummary(
        total_payment_amount=item.total_payment_amount,
        total_salesperson_commission=item.total_salesperson_commission,
        total_supervisor_commission=item.total_supervisor_commission,
        total_second_supervisor_commission=item.total_second_supervisor_commission,
        total_commission=item.total_commission,
        total_commission_rmb=item.total_commission_rmb,
    )


def _build_monthly_summary(rows, user_ids: list[str]) -> list[SalesCommissionMonthlySummary]:
    buckets = {}
    for row in rows:
        detail = row[0]
        payment_date = row[8] or row[6]
        if not payment_date:
            continue
        exchange_source = row[7] if row[7] is not None else row[10]
        month = payment_date.strftime("%Y-%m")
        sp, sv, sv2 = _role_commissions(detail, user_ids)
        commission_usd = sp + sv + sv2
        bucket = buckets.setdefault(month, {"usd": 0, "rmb": 0, "rates": []})
        bucket["usd"] += commission_usd
        bucket["rmb"] += _commission_rmb(commission_usd, exchange_source)
        if exchange_source is not None:
            bucket["rates"].append(_display_exchange_rate(exchange_source))

    result = []
    for month in sorted(buckets):
        bucket = buckets[month]
        rates = bucket["rates"]
        avg_rate = sum(rates) / len(rates) if rates else 0
        result.append(SalesCommissionMonthlySummary(
            month=month,
            total_commission_usd=round(bucket["usd"], 2),
            average_exchange_rate=round(avg_rate, 6),
            total_commission_rmb=round(bucket["rmb"], 2),
        ))
    return result


@router.get("/self/batch/list", summary="业务员查询自己的提成批次")
def list_my_commission_batches(
    status: str = Query("", description="状态筛选：confirming/confirmed"),
    keyword: str = Query("", description="批次名称关键字"),
    month: str = Query("", description="月份筛选，格式 YYYY-MM"),
    role: str = Query("", description="关联角色：salesperson/supervisor/second_supervisor"),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_any_permission("commission:self_read", "commission:read", "commission:write")),
) -> ResponseModel[PageResponse[SalesCommissionBatchListItem]]:
    """业务员只能看到确认中/已确认，且与自己相关的批次。"""
    _, user_ids = _current_business_context(db, current_user)
    if not user_ids:
        return ResponseModel(data=PageResponse(items=[], total=0, page=page, page_size=page_size))

    visible_statuses = SELF_VISIBLE_BATCH_STATUSES
    if status:
        if status not in SELF_VISIBLE_BATCH_STATUSES:
            return ResponseModel(data=PageResponse(items=[], total=0, page=page, page_size=page_size))
        visible_statuses = (status,)

    related_batch_ids = db.query(CommissionDetail.batch_id).filter(
        _self_role_condition(user_ids, role),
        CommissionDetail.status != "voided",
    ).distinct()

    query = db.query(CommissionBatch).filter(
        CommissionBatch.status.in_(visible_statuses),
        CommissionBatch.id.in_(related_batch_ids),
    )

    if keyword:
        query = query.filter(CommissionBatch.batch_name.like(f"%{keyword}%"))
    if month:
        start, end = _month_bounds(month)
        query = query.filter(CommissionBatch.period_start <= end, CommissionBatch.period_end >= start)

    query = query.order_by(desc(CommissionBatch.created_at), desc(CommissionBatch.id))
    total = query.count()
    batches = query.offset((page - 1) * page_size).limit(page_size).all()

    items = []
    ark_user_id = str(current_user.get("sub") or "")
    for batch in batches:
        rows = _query_self_detail_rows(db, batch.id, user_ids, role)
        my_confirm = _my_confirmation(db, batch.id, ark_user_id, user_ids)
        items.append(_build_sales_batch_item(batch, rows, user_ids, my_confirm))

    return ResponseModel(data=PageResponse(items=items, total=total, page=page, page_size=page_size))


@router.get("/self/batch/{batch_id}", summary="业务员查询自己的提成批次详情")
def get_my_commission_batch_detail(
    batch_id: int = Path(..., description="批次ID"),
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_any_permission("commission:self_read", "commission:read", "commission:write")),
) -> ResponseModel[SalesCommissionBatchDetail]:
    ark_user, user_ids = _current_business_context(db, current_user)
    batch = db.query(CommissionBatch).filter(CommissionBatch.id == batch_id).first()
    if not batch or batch.status not in SELF_VISIBLE_BATCH_STATUSES:
        raise HTTPException(status_code=404, detail="提成批次不存在或不可查看")
    if not user_ids:
        raise HTTPException(status_code=403, detail="未匹配到业务库员工身份")

    all_rows = _query_self_detail_rows(db, batch_id, user_ids)
    if not all_rows:
        raise HTTPException(status_code=404, detail="提成批次不存在或不可查看")

    salesperson_rows = _query_self_detail_rows(db, batch_id, user_ids, "salesperson")
    supervisor_rows = _query_self_detail_rows(db, batch_id, user_ids, "supervisor")
    second_supervisor_rows = _query_self_detail_rows(db, batch_id, user_ids, "second_supervisor")

    my_confirm = _my_confirmation(db, batch_id, str(current_user.get("sub") or ""), user_ids)
    data = SalesCommissionBatchDetail(
        batch=_build_sales_batch_item(batch, all_rows, user_ids, my_confirm),
        summary=_build_sales_summary(all_rows, user_ids),
        monthly_summary=_build_monthly_summary(all_rows, user_ids),
        salesperson_details=[_detail_schema_from_row(row) for row in salesperson_rows],
        supervisor_details=[_detail_schema_from_row(row) for row in supervisor_rows],
        second_supervisor_details=[_detail_schema_from_row(row) for row in second_supervisor_rows],
    )
    return ResponseModel(data=data)


@router.post("/self/batch/{batch_id}/feedback", summary="业务员提交提成批次问题反馈")
def submit_my_commission_feedback(
    batch_id: int,
    req: SalesCommissionFeedbackRequest,
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_any_permission("commission:self_read", "commission:read", "commission:write")),
) -> ResponseModel[dict]:
    ark_user, user_ids = _current_business_context(db, current_user)
    batch = db.query(CommissionBatch).filter(CommissionBatch.id == batch_id).first()
    if not batch or batch.status != "confirming":
        raise HTTPException(status_code=404, detail="提成批次不存在或当前不可反馈")
    if not user_ids or not _query_self_detail_rows(db, batch_id, user_ids):
        raise HTTPException(status_code=403, detail="无权反馈该提成批次")
    if _my_confirmation(db, batch_id, str(current_user.get("sub") or ""), user_ids):
        raise HTTPException(status_code=400, detail="已提交确认，不能继续反馈")

    content = (req.content or "").strip()
    if not content:
        raise HTTPException(status_code=400, detail="反馈内容不能为空")
    if len(content) > 2000:
        raise HTTPException(status_code=400, detail="反馈内容不能超过 2000 字")

    feedback = CommissionBatchFeedback(
        batch_id=batch_id,
        ark_user_id=str(current_user.get("sub") or ""),
        user_name=(ark_user.real_name if ark_user else current_user.get("username")),
        business_user_ids=",".join(user_ids),
        content=content,
    )
    db.add(feedback)
    db.commit()
    db.refresh(feedback)
    return ResponseModel(message="反馈已提交", data={"id": feedback.id, "batch_id": batch_id})


@router.post("/self/batch/{batch_id}/confirm", summary="业务员确认自己的提成批次")
def confirm_my_commission_batch(
    batch_id: int,
    req: SalesCommissionConfirmRequest,
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_any_permission("commission:self_read", "commission:read", "commission:write")),
) -> ResponseModel[dict]:
    ark_user, user_ids = _current_business_context(db, current_user)
    if req.confirmation_text != CONFIRMATION_TEXT:
        raise HTTPException(status_code=400, detail="请输入“我已确认”后再提交")

    batch = db.query(CommissionBatch).filter(CommissionBatch.id == batch_id).first()
    if not batch or batch.status != "confirming":
        raise HTTPException(status_code=404, detail="提成批次不存在或当前不可确认")
    if not user_ids or not _query_self_detail_rows(db, batch_id, user_ids):
        raise HTTPException(status_code=403, detail="无权确认该提成批次")

    ark_user_id = str(current_user.get("sub") or "")
    existing = _my_confirmation(db, batch_id, ark_user_id, user_ids)
    if existing:
        return ResponseModel(message="已确认", data={"batch_id": batch_id, "confirmed_at": existing.confirmed_at})

    confirmation = CommissionBatchConfirmation(
        batch_id=batch_id,
        ark_user_id=ark_user_id,
        user_name=(ark_user.real_name if ark_user else current_user.get("username")),
        business_user_ids=",".join(user_ids),
        confirmation_text=req.confirmation_text,
        status="confirmed",
        confirmed_at=datetime.utcnow(),
    )
    db.add(confirmation)
    db.commit()
    db.refresh(confirmation)
    return ResponseModel(message="确认已提交", data={"batch_id": batch_id, "confirmed_at": confirmation.confirmed_at})


@router.get("/self/batch/{batch_id}/export", summary="业务员导出自己的提成明细")
def export_my_commission_batch(
    batch_id: int,
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_any_permission("commission:self_read", "commission:read", "commission:write")),
) -> StreamingResponse:
    _, user_ids = _current_business_context(db, current_user)
    batch = db.query(CommissionBatch).filter(CommissionBatch.id == batch_id).first()
    if not batch or batch.status not in SELF_VISIBLE_BATCH_STATUSES:
        raise HTTPException(status_code=404, detail="提成批次不存在或不可导出")
    if not user_ids:
        raise HTTPException(status_code=403, detail="未匹配到业务库员工身份")

    all_rows = _query_self_detail_rows(db, batch_id, user_ids)
    if not all_rows:
        raise HTTPException(status_code=403, detail="无权导出该提成批次")

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    summary_sheet = wb.active
    summary_sheet.title = "月度汇总"
    headers = ["月份", "总提成(美元)", "月平均汇率", "总提成(人民币)"]
    summary_sheet.append(headers)
    for item in _build_monthly_summary(all_rows, user_ids):
        summary_sheet.append([
            item.month,
            item.total_commission_usd,
            item.average_exchange_rate,
            item.total_commission_rmb,
        ])

    def write_detail_sheet(title: str, rows, commission_key: str, rmb_key: str):
        sheet = wb.create_sheet(title)
        sheet.append([
            "回款日期", "支付方式", "回款单ID", "订单名称", "客户名称", "客户国家",
            "回款金额(美元)", "汇率", "提成(美元)", "提成(人民币)", "计算说明",
        ])
        for row in rows:
            item = _detail_schema_from_row(row)
            sheet.append([
                item.collection_date,
                item.type,
                item.payment_id,
                item.order_name or item.order_no or item.order_id,
                item.customer_name,
                item.customer_country,
                item.payment_amount,
                item.exchange_rate,
                getattr(item, commission_key),
                getattr(item, rmb_key),
                item.calc_rule_note,
            ])

    write_detail_sheet(
        "业务提成",
        _query_self_detail_rows(db, batch_id, user_ids, "salesperson"),
        "salesperson_commission",
        "salesperson_commission_rmb",
    )
    write_detail_sheet(
        "一级主管提成",
        _query_self_detail_rows(db, batch_id, user_ids, "supervisor"),
        "supervisor_commission",
        "supervisor_commission_rmb",
    )
    write_detail_sheet(
        "二级主管提成",
        _query_self_detail_rows(db, batch_id, user_ids, "second_supervisor"),
        "second_supervisor_commission",
        "second_supervisor_commission_rmb",
    )

    for sheet in wb.worksheets:
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="2F5597")
        for column_cells in sheet.columns:
            sheet.column_dimensions[column_cells[0].column_letter].width = min(max(len(str(c.value or "")) for c in column_cells) + 2, 28)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    filename = quote(f"{batch.batch_name}-我的提成明细.xlsx")
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"},
    )


@router.post("/batch", summary="创建提成批次")
def create_batch(
    req: CommissionBatchCreateRequest,
    db: Session = Depends(get_db),
    _user: dict = Depends(require_permission("commission:write")),
) -> ResponseModel[CommissionBatchListItem]:
    """创建提成批次（status=draft）"""
    batch = CommissionBatch(
        batch_name=req.batch_name,
        period_type=req.period_type,
        period_start=req.period_start,
        period_end=req.period_end,
        status="draft",
    )
    db.add(batch)
    db.commit()
    db.refresh(batch)

    return ResponseModel(data=CommissionBatchListItem.model_validate(batch))


@router.get("/batch/list", summary="查询批次列表")
def list_batches(
    status: str = Query("", description="状态筛选"),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    sort_field: str = Query("created_at"),
    sort_order: str = Query("desc"),
    db: Session = Depends(get_db),
    _user: dict = Depends(require_any_permission("commission:read", "commission:write")),
) -> ResponseModel[PageResponse[CommissionBatchListItem]]:
    """查询提成批次列表"""
    from sqlalchemy import desc as _desc
    query = db.query(CommissionBatch)

    if status:
        query = query.filter(CommissionBatch.status == status)

    SORT_MAP = {
        "batch_name": CommissionBatch.batch_name,
        "period_start": CommissionBatch.period_start,
        "period_end": CommissionBatch.period_end,
        "status": CommissionBatch.status,
        "created_at": CommissionBatch.created_at,
    }
    sort_col = SORT_MAP.get(sort_field, CommissionBatch.created_at)
    order_fn = _desc if sort_order == "desc" else lambda c: c

    query = query.order_by(order_fn(sort_col))
    total = query.count()
    rows = query.offset((page - 1) * page_size).limit(page_size).all()

    items = [_batch_list_item(db, b) for b in rows]
    page_data = PageResponse(items=items, total=total, page=page, page_size=page_size)
    return ResponseModel(data=page_data)


@router.post("/batch/{batch_id}/calculate", summary="执行提成计算")
def execute_calculation(
    batch_id: int = Path(..., description="批次ID"),
    db: Session = Depends(get_db),
    _user: dict = Depends(require_permission("commission:write")),
) -> ResponseModel[CommissionCalcResponse]:
    """对指定批次执行提成计算"""
    result = calculate_commission(db, batch_id)
    db.commit()

    return ResponseModel(data=CommissionCalcResponse(
        total_payments=result.total_payments,
        total_salesperson_commission=float(result.total_salesperson_commission),
        total_supervisor_commission=float(result.total_supervisor_commission),
        total_second_supervisor_commission=float(result.total_second_supervisor_commission),
        skipped_incomplete=result.skipped_incomplete,
        skipped_no_snapshot=result.skipped_no_snapshot,
        errors=result.errors,
    ))


@router.get("/batch/{batch_id}/details", summary="查询提成明细")
def list_commission_details(
    batch_id: int = Path(..., description="批次ID"),
    salesperson_id: str = Query("", description="业务员ID"),
    supervisor_id: str = Query("", description="主管ID"),
    customer_id: str = Query("", description="客户ID"),
    keyword: str = Query("", description="搜索关键字"),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    sort_field: str = Query("payment_amount"),
    sort_order: str = Query("desc"),
    db: Session = Depends(get_db),
    _user: dict = Depends(require_any_permission("commission:read", "commission:write")),
) -> ResponseModel[PageResponse[CommissionDetailListItem]]:
    """查询指定批次的提成明细"""
    SpUser = aliased(UserBasic)
    SvUser = aliased(UserBasic)
    Sv2User = aliased(UserBasic)

    query = db.query(
        CommissionDetail,
        CustomerInfo.company_name.label("customer_name"),
        SpUser.full_name.label("salesperson_name"),
        SvUser.full_name.label("supervisor_name"),
        Sv2User.full_name.label("second_supervisor_name"),
    ).outerjoin(
        CustomerInfo,
        CommissionDetail.customer_id == CustomerInfo.company_id,
    ).outerjoin(
        SpUser, CommissionDetail.salesperson_id == SpUser.user_id,
    ).outerjoin(
        SvUser, CommissionDetail.supervisor_id == SvUser.user_id,
    ).outerjoin(
        Sv2User, CommissionDetail.second_supervisor_id == Sv2User.user_id,
    ).filter(
        CommissionDetail.batch_id == batch_id,
    )

    if salesperson_id:
        query = query.filter(CommissionDetail.salesperson_id == salesperson_id)
    if supervisor_id:
        query = query.filter(CommissionDetail.supervisor_id == supervisor_id)
    if customer_id:
        query = query.filter(CommissionDetail.customer_id == customer_id)
    if keyword:
        like_pattern = f"%{keyword}%"
        query = query.filter(
            CustomerInfo.company_name.like(like_pattern)
            | SpUser.full_name.like(like_pattern)
            | SvUser.full_name.like(like_pattern)
            | CommissionDetail.payment_id.like(like_pattern)
        )

    from sqlalchemy import desc as _desc
    SORT_MAP = {
        "customer_name": CustomerInfo.company_name,
        "salesperson_name": SpUser.full_name,
        "payment_amount": CommissionDetail.payment_amount,
        "commission_amount": CommissionDetail.salesperson_commission,
    }
    sort_col = SORT_MAP.get(sort_field, CommissionDetail.payment_amount)
    order_fn = _desc if sort_order == "desc" else lambda c: c
    query = query.order_by(order_fn(sort_col))

    total = query.count()
    rows = query.offset((page - 1) * page_size).limit(page_size).all()

    items = []
    for detail, customer_name, sp_name, sv_name, sv2_name in rows:
        items.append(CommissionDetailListItem(
            id=detail.id,
            payment_id=detail.payment_id,
            order_id=detail.order_id,
            customer_id=detail.customer_id,
            customer_name=customer_name,
            payment_amount=float(detail.payment_amount),
            salesperson_id=detail.salesperson_id,
            salesperson_name=sp_name,
            salesperson_rate=float(detail.salesperson_rate),
            salesperson_commission=float(detail.salesperson_commission),
            supervisor_id=detail.supervisor_id,
            supervisor_name=sv_name,
            supervisor_rate=float(detail.supervisor_rate) if detail.supervisor_rate else None,
            supervisor_commission=float(detail.supervisor_commission),
            second_supervisor_id=detail.second_supervisor_id,
            second_supervisor_name=sv2_name,
            second_supervisor_rate=float(detail.second_supervisor_rate) if detail.second_supervisor_rate else None,
            second_supervisor_commission=float(detail.second_supervisor_commission),
            calc_rule_note=detail.calc_rule_note,
            status=detail.status,
        ))

    page_data = PageResponse(items=items, total=total, page=page, page_size=page_size)
    return ResponseModel(data=page_data)


@router.post("/batch/{batch_id}/confirm", summary="确认批次")
def confirm_commission_batch(
    batch_id: int = Path(..., description="批次ID"),
    req: CommissionConfirmRequest = ...,
    db: Session = Depends(get_db),
    _user: dict = Depends(require_permission("commission:write")),
) -> ResponseModel[dict]:
    """确认提成批次"""
    confirm_batch(db, batch_id, req.confirmed_by)
    db.commit()
    return ResponseModel(message="批次已确认", data={"batch_id": batch_id})


@router.post("/batch/{batch_id}/send-confirm", summary="发送确认")
async def send_confirm_batch(
    batch_id: int = Path(..., description="批次ID"),
    req: CommissionSendConfirmRequest | None = None,
    db: Session = Depends(get_db),
    _user: dict = Depends(require_permission("commission:write")),
) -> ResponseModel[dict]:
    """发送确认：批次状态变为 confirming，可选是否推送钉钉通知。"""
    from app.dingtalk.events import notify_commission_confirm

    notify_dingtalk = True if req is None else req.notify_dingtalk
    result = send_confirm(db, batch_id)
    db.commit()

    if notify_dingtalk and result["dingtalk_ids"]:
        await notify_commission_confirm(result["dingtalk_ids"], result["batch_name"])

    return ResponseModel(
        message="已发送确认",
        data={
            "batch_id": batch_id,
            "notify_dingtalk": notify_dingtalk,
            "notified_count": len(result["dingtalk_ids"]) if notify_dingtalk else 0,
        },
    )


@router.post("/batch/{batch_id}/revoke-confirm", summary="撤销确认")
def revoke_confirm_batch(
    batch_id: int = Path(..., description="批次ID"),
    db: Session = Depends(get_db),
    _user: dict = Depends(require_permission("commission:write")),
) -> ResponseModel[dict]:
    """撤销确认，批次状态回到 calculated"""
    revoke_confirm(db, batch_id)
    db.commit()
    return ResponseModel(message="已撤销确认", data={"batch_id": batch_id})


@router.post("/batch/{batch_id}/void", summary="作废批次")
def void_commission_batch(
    batch_id: int = Path(..., description="批次ID"),
    db: Session = Depends(get_db),
    _user: dict = Depends(require_permission("commission:write")),
) -> ResponseModel[dict]:
    """作废提成批次"""
    void_batch(db, batch_id)
    db.commit()
    return ResponseModel(message="批次已作废", data={"batch_id": batch_id})


@router.get("/batch/{batch_id}/summary", summary="批次汇总统计")
def get_batch_summary(
    batch_id: int = Path(..., description="批次ID"),
    db: Session = Depends(get_db),
    _user: dict = Depends(require_any_permission("commission:read", "commission:write")),
) -> ResponseModel[CommissionBatchSummary]:
    """查询批次汇总统计"""
    batch = db.query(CommissionBatch).filter(CommissionBatch.id == batch_id).first()
    if not batch:
        return ResponseModel(code=404, message=f"批次 {batch_id} 不存在")

    # 聚合提成明细（排除 voided）
    agg = db.query(
        func.count(CommissionDetail.id).label("total_payments"),
        func.coalesce(func.sum(CommissionDetail.payment_amount), 0).label("total_payment_amount"),
        func.coalesce(func.sum(CommissionDetail.salesperson_commission), 0).label("total_sp_commission"),
        func.coalesce(func.sum(CommissionDetail.supervisor_commission), 0).label("total_sv_commission"),
        func.coalesce(func.sum(CommissionDetail.second_supervisor_commission), 0).label("total_sv2_commission"),
        func.count(func.distinct(CommissionDetail.salesperson_id)).label("salesperson_count"),
        func.count(func.distinct(CommissionDetail.supervisor_id)).label("supervisor_count"),
        func.count(func.distinct(CommissionDetail.second_supervisor_id)).label("second_supervisor_count"),
    ).filter(
        CommissionDetail.batch_id == batch_id,
        CommissionDetail.status != "voided",
    ).first()

    # 计算跳过数：期间内已同步但未参与计算的回款
    period_payment_count = db.query(func.count(SyncedPayment.id)).filter(
        SyncedPayment.payment_date >= batch.period_start,
        SyncedPayment.payment_date <= batch.period_end,
    ).scalar() or 0

    calculated_count = agg.total_payments if agg else 0
    total_skipped = period_payment_count - calculated_count

    # 分析跳过原因
    skipped_incomplete = 0
    skipped_no_snapshot = 0
    if total_skipped > 0:
        calculated_pids = db.query(PaymentCommissionStatus.payment_id).filter(
            PaymentCommissionStatus.batch_id == batch_id,
        ).subquery()

        skipped_payments = db.query(SyncedPayment.customer_id).filter(
            SyncedPayment.payment_date >= batch.period_start,
            SyncedPayment.payment_date <= batch.period_end,
            ~SyncedPayment.payment_id.in_(db.query(calculated_pids.c.payment_id)),
        ).all()

        skipped_cids = {r.customer_id for r in skipped_payments}
        for cid in skipped_cids:
            snap = db.query(CustomerCommissionSnapshot).filter(
                CustomerCommissionSnapshot.customer_id == cid,
                CustomerCommissionSnapshot.is_current == True,
            ).first()
            if not snap:
                skipped_no_snapshot += 1
            elif not snap.is_complete:
                skipped_incomplete += 1

    total_sp = float(agg.total_sp_commission) if agg else 0
    total_sv = float(agg.total_sv_commission) if agg else 0
    total_sv2 = float(agg.total_sv2_commission) if agg else 0
    progress = _confirmation_progress(db, batch_id)

    return ResponseModel(data=CommissionBatchSummary(
        batch_name=batch.batch_name,
        status=batch.status,
        total_payments=calculated_count,
        total_payment_amount=float(agg.total_payment_amount) if agg else 0,
        total_salesperson_commission=total_sp,
        total_supervisor_commission=total_sv,
        total_second_supervisor_commission=total_sv2,
        total_commission=total_sp + total_sv + total_sv2,
        salesperson_count=agg.salesperson_count if agg else 0,
        supervisor_count=agg.supervisor_count if agg else 0,
        second_supervisor_count=agg.second_supervisor_count if agg else 0,
        skipped_incomplete=skipped_incomplete,
        skipped_no_snapshot=skipped_no_snapshot,
        expected_confirm_count=progress.expected_confirm_count,
        confirmed_count=progress.confirmed_count,
        pending_confirm_count=progress.pending_confirm_count,
        feedback_count=progress.feedback_count,
        confirmation_status=progress.confirmation_status,
    ))
