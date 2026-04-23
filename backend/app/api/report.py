"""报表导出路由"""

from collections import defaultdict
from decimal import Decimal
from io import BytesIO

from fastapi import APIRouter, Depends, Query, Path
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from sqlalchemy import func
from sqlalchemy.orm import Session, aliased

from app.api.deps import get_db
from app.models.commission import CommissionBatch, CommissionDetail, SyncedPayment
from app.models.business import UserBasic, CustomerInfo

router = APIRouter()

HEADER_FONT = Font(bold=True)
HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center")

DETAIL_COLUMNS = [
    "回款ID", "订单ID", "客户ID", "客户名称", "回款日期", "回款金额", "交易手续费", "提成金额",
    "业务员ID", "业务员", "业务员比例", "业务员提成",
    "一级主管ID", "一级主管", "一级主管比例", "一级主管提成",
    "二级主管ID", "二级主管", "二级主管比例", "二级主管提成",
    "计算规则", "状态",
]


def _style_header(ws, columns):
    """为表头行应用样式"""
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN


def _write_detail_row(ws, row_num, d, customer_name, sp_name, sv_name, sv2_name="", payment_date=None, service_fee=0):
    """写入一行明细数据"""
    amount = float(d.payment_amount)
    fee = float(service_fee or 0)
    ws.cell(row=row_num, column=1, value=d.payment_id)
    ws.cell(row=row_num, column=2, value=d.order_id)
    ws.cell(row=row_num, column=3, value=d.customer_id)
    ws.cell(row=row_num, column=4, value=customer_name or "")
    ws.cell(row=row_num, column=5, value=payment_date)
    ws.cell(row=row_num, column=6, value=amount)
    ws.cell(row=row_num, column=7, value=fee)
    ws.cell(row=row_num, column=8, value=amount - fee)
    ws.cell(row=row_num, column=9, value=d.salesperson_id)
    ws.cell(row=row_num, column=10, value=sp_name or "")
    ws.cell(row=row_num, column=11, value=float(d.salesperson_rate))
    ws.cell(row=row_num, column=12, value=float(d.salesperson_commission))
    ws.cell(row=row_num, column=13, value=d.supervisor_id or "")
    ws.cell(row=row_num, column=14, value=sv_name or "")
    ws.cell(row=row_num, column=15, value=float(d.supervisor_rate) if d.supervisor_rate else 0)
    ws.cell(row=row_num, column=16, value=float(d.supervisor_commission))
    ws.cell(row=row_num, column=17, value=d.second_supervisor_id or "")
    ws.cell(row=row_num, column=18, value=sv2_name or "")
    ws.cell(row=row_num, column=19, value=float(d.second_supervisor_rate) if d.second_supervisor_rate else 0)
    ws.cell(row=row_num, column=20, value=float(d.second_supervisor_commission))
    ws.cell(row=row_num, column=21, value=d.calc_rule_note or "")
    ws.cell(row=row_num, column=22, value=d.status)


def _add_subtotal_row(ws, row_num, count):
    """写入小计行"""
    bold = Font(bold=True)
    ws.cell(row=row_num, column=4, value="小计").font = bold
    ws.cell(row=row_num, column=6, value=f"=SUM(F2:F{row_num - 1})").font = bold
    ws.cell(row=row_num, column=7, value=f"=SUM(G2:G{row_num - 1})").font = bold
    ws.cell(row=row_num, column=8, value=f"=SUM(H2:H{row_num - 1})").font = bold
    ws.cell(row=row_num, column=12, value=f"=SUM(L2:L{row_num - 1})").font = bold
    ws.cell(row=row_num, column=16, value=f"=SUM(P2:P{row_num - 1})").font = bold
    ws.cell(row=row_num, column=20, value=f"=SUM(T2:T{row_num - 1})").font = bold


SECTION_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
SECTION_FONT = Font(bold=True, size=12)


def _add_section_subtotal(ws, row_num, start_row):
    """写入分类小计行（指定起止行）"""
    bold = Font(bold=True)
    ws.cell(row=row_num, column=4, value="小计").font = bold
    if start_row <= row_num - 1:
        ws.cell(row=row_num, column=6, value=f"=SUM(F{start_row}:F{row_num - 1})").font = bold
        ws.cell(row=row_num, column=7, value=f"=SUM(G{start_row}:G{row_num - 1})").font = bold
        ws.cell(row=row_num, column=8, value=f"=SUM(H{start_row}:H{row_num - 1})").font = bold
        ws.cell(row=row_num, column=12, value=f"=SUM(L{start_row}:L{row_num - 1})").font = bold
        ws.cell(row=row_num, column=16, value=f"=SUM(P{start_row}:P{row_num - 1})").font = bold
        ws.cell(row=row_num, column=20, value=f"=SUM(T{start_row}:T{row_num - 1})").font = bold


def _write_salesperson_grouped_sheets(wb, rows):
    """按业务员分sheet，每人含作为业务员/一级主管/二级主管三个分类"""
    person_roles = defaultdict(lambda: {"as_sp": [], "as_sv": [], "as_sv2": []})
    person_names = {}

    for d, cname, spname, svname, sv2name, pdate, sfee in rows:
        sp_id = d.salesperson_id
        sv_id = d.supervisor_id
        sv2_id = d.second_supervisor_id
        row_data = (d, cname, spname, svname, sv2name, pdate, sfee)

        if sp_id:
            person_roles[sp_id]["as_sp"].append(row_data)
            if spname:
                person_names[sp_id] = spname
        if sv_id and sv_id != sp_id:
            person_roles[sv_id]["as_sv"].append(row_data)
            if svname:
                person_names[sv_id] = svname
        if sv2_id and sv2_id != sp_id and sv2_id != sv_id:
            person_roles[sv2_id]["as_sv2"].append(row_data)
            if sv2name:
                person_names[sv2_id] = sv2name

    sections = [
        ("as_sp", "作为业务员"),
        ("as_sv", "作为一级主管"),
        ("as_sv2", "作为二级主管"),
    ]

    for person_id, role_data in person_roles.items():
        name = person_names.get(person_id, person_id)
        sheet_name = str(name)[:31]
        ws = wb.create_sheet(sheet_name)
        _style_header(ws, DETAIL_COLUMNS)
        cur_row = 2

        for role_key, section_title in sections:
            section_rows = role_data[role_key]
            if not section_rows:
                continue

            cell = ws.cell(row=cur_row, column=1, value=f"▎{section_title}（{len(section_rows)} 笔）")
            cell.font = SECTION_FONT
            cell.fill = SECTION_FILL
            for c in range(2, len(DETAIL_COLUMNS) + 1):
                ws.cell(row=cur_row, column=c).fill = SECTION_FILL
            cur_row += 1

            data_start = cur_row
            for rd in section_rows:
                d, cname, spname, svname, sv2name, pdate, sfee = rd
                _write_detail_row(ws, cur_row, d, cname, spname, svname, sv2name, pdate, sfee)
                cur_row += 1

            _add_section_subtotal(ws, cur_row, data_start)
            cur_row += 2


def _query_details(db: Session, batch_id: int):
    """查询批次全部明细，含名称、回款日期、手续费"""
    SpUser = aliased(UserBasic)
    SvUser = aliased(UserBasic)
    Sv2User = aliased(UserBasic)

    return db.query(
        CommissionDetail,
        CustomerInfo.company_name.label("customer_name"),
        SpUser.full_name.label("salesperson_name"),
        SvUser.full_name.label("supervisor_name"),
        Sv2User.full_name.label("second_supervisor_name"),
        SyncedPayment.payment_date.label("payment_date"),
        SyncedPayment.service_fee.label("service_fee"),
    ).outerjoin(
        CustomerInfo, CommissionDetail.customer_id == CustomerInfo.company_id,
    ).outerjoin(
        SpUser, CommissionDetail.salesperson_id == SpUser.user_id,
    ).outerjoin(
        SvUser, CommissionDetail.supervisor_id == SvUser.user_id,
    ).outerjoin(
        Sv2User, CommissionDetail.second_supervisor_id == Sv2User.user_id,
    ).outerjoin(
        SyncedPayment, CommissionDetail.payment_id == SyncedPayment.payment_id,
    ).filter(
        CommissionDetail.batch_id == batch_id,
        CommissionDetail.status != "voided",
    ).order_by(
        SyncedPayment.payment_date,
    ).all()


def _to_streaming(wb: Workbook, filename: str) -> StreamingResponse:
    """将 Workbook 转为 StreamingResponse"""
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/commission/{batch_id}/export", summary="导出提成明细Excel")
def export_commission_details(
    batch_id: int = Path(..., description="批次ID"),
    group_by: str = Query("", description="分组方式: salesperson/supervisor/customer"),
    db: Session = Depends(get_db),
):
    """导出提成明细 Excel，支持按业务员/主管/客户分sheet"""
    batch = db.query(CommissionBatch).filter(CommissionBatch.id == batch_id).first()
    if not batch:
        return {"code": 404, "message": f"批次 {batch_id} 不存在", "data": None}

    rows = _query_details(db, batch_id)

    wb = Workbook()
    wb.remove(wb.active)

    if not group_by:
        ws = wb.create_sheet("全部明细")
        _style_header(ws, DETAIL_COLUMNS)
        for idx, (d, cname, spname, svname, sv2name, pdate, sfee) in enumerate(rows, start=2):
            _write_detail_row(ws, idx, d, cname, spname, svname, sv2name, pdate, sfee)
        if rows:
            _add_subtotal_row(ws, len(rows) + 2, len(rows))
    elif group_by == "salesperson":
        _write_salesperson_grouped_sheets(wb, rows)
    else:
        groups = defaultdict(list)
        for d, cname, spname, svname, sv2name, pdate, sfee in rows:
            if group_by == "supervisor":
                key = svname or (d.supervisor_id or "无主管")
            else:
                key = cname or d.customer_id
            groups[key].append((d, cname, spname, svname, sv2name, pdate, sfee))

        for group_name, group_rows in groups.items():
            sheet_name = str(group_name)[:31]
            ws = wb.create_sheet(sheet_name)
            _style_header(ws, DETAIL_COLUMNS)
            for idx, (d, cname, spname, svname, sv2name, pdate, sfee) in enumerate(group_rows, start=2):
                _write_detail_row(ws, idx, d, cname, spname, svname, sv2name, pdate, sfee)
            _add_subtotal_row(ws, len(group_rows) + 2, len(group_rows))

    if not wb.sheetnames:
        ws = wb.create_sheet("空")
        ws.cell(row=1, column=1, value="无数据")

    return _to_streaming(wb, f"commission_details_{batch.batch_name}.xlsx")


@router.get("/commission/{batch_id}/salesperson-summary", summary="导出业务员提成汇总Excel")
def export_salesperson_summary(
    batch_id: int = Path(..., description="批次ID"),
    db: Session = Depends(get_db),
):
    """导出业务员提成汇总 Excel"""
    batch = db.query(CommissionBatch).filter(CommissionBatch.id == batch_id).first()
    if not batch:
        return {"code": 404, "message": f"批次 {batch_id} 不存在", "data": None}

    results = db.query(
        CommissionDetail.salesperson_id,
        UserBasic.full_name,
        func.count(CommissionDetail.id).label("payment_count"),
        func.sum(CommissionDetail.payment_amount).label("total_amount"),
        CommissionDetail.salesperson_rate,
        func.sum(CommissionDetail.salesperson_commission).label("total_commission"),
    ).outerjoin(
        UserBasic, CommissionDetail.salesperson_id == UserBasic.user_id,
    ).filter(
        CommissionDetail.batch_id == batch_id,
        CommissionDetail.status != "voided",
    ).group_by(
        CommissionDetail.salesperson_id,
        UserBasic.full_name,
        CommissionDetail.salesperson_rate,
    ).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "业务员提成汇总"

    columns = ["业务员ID", "姓名", "回款笔数", "回款总额", "提成比例", "提成金额"]
    _style_header(ws, columns)

    for idx, r in enumerate(results, start=2):
        ws.cell(row=idx, column=1, value=r.salesperson_id)
        ws.cell(row=idx, column=2, value=r.full_name or "")
        ws.cell(row=idx, column=3, value=r.payment_count)
        ws.cell(row=idx, column=4, value=float(r.total_amount or 0))
        ws.cell(row=idx, column=5, value=float(r.salesperson_rate or 0))
        ws.cell(row=idx, column=6, value=float(r.total_commission or 0))

    return _to_streaming(wb, f"salesperson_summary_{batch.batch_name}.xlsx")


@router.get("/commission/{batch_id}/supervisor-summary", summary="导出主管提成汇总Excel")
def export_supervisor_summary(
    batch_id: int = Path(..., description="批次ID"),
    db: Session = Depends(get_db),
):
    """导出主管提成汇总 Excel"""
    batch = db.query(CommissionBatch).filter(CommissionBatch.id == batch_id).first()
    if not batch:
        return {"code": 404, "message": f"批次 {batch_id} 不存在", "data": None}

    results = db.query(
        CommissionDetail.supervisor_id,
        UserBasic.full_name,
        func.count(func.distinct(CommissionDetail.salesperson_id)).label("sp_count"),
        func.sum(CommissionDetail.payment_amount).label("total_amount"),
        CommissionDetail.supervisor_rate,
        func.sum(CommissionDetail.supervisor_commission).label("total_commission"),
    ).outerjoin(
        UserBasic, CommissionDetail.supervisor_id == UserBasic.user_id,
    ).filter(
        CommissionDetail.batch_id == batch_id,
        CommissionDetail.status != "voided",
        CommissionDetail.supervisor_id.isnot(None),
    ).group_by(
        CommissionDetail.supervisor_id,
        UserBasic.full_name,
        CommissionDetail.supervisor_rate,
    ).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "一级主管提成汇总"

    columns = ["主管ID", "姓名", "管辖业务员数", "管辖回款总额", "提成比例", "提成金额"]
    _style_header(ws, columns)

    for idx, r in enumerate(results, start=2):
        ws.cell(row=idx, column=1, value=r.supervisor_id or "")
        ws.cell(row=idx, column=2, value=r.full_name or "")
        ws.cell(row=idx, column=3, value=r.sp_count)
        ws.cell(row=idx, column=4, value=float(r.total_amount or 0))
        ws.cell(row=idx, column=5, value=float(r.supervisor_rate or 0))
        ws.cell(row=idx, column=6, value=float(r.total_commission or 0))

    return _to_streaming(wb, f"supervisor_summary_{batch.batch_name}.xlsx")
