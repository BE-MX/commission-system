"""客户归属重置、快照补全、Excel 导入服务"""

import logging
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal
from typing import Optional

from openpyxl import load_workbook
from sqlalchemy import text
from sqlalchemy.orm import Session

from app.core.config import get_settings
from app.models.customer import CustomerCommissionSnapshot
from app.models.employee import EmployeeAttributeHistory, SupervisorRelationHistory
from app.models.business import OkkiReceipt
from app.services.rate_utils import calc_commission_rates, get_employee_attribute_at_date

logger = logging.getLogger("commission.customer")


@dataclass
class ImportResult:
    """Excel 导入结果"""
    total_rows: int = 0
    success: int = 0
    failed: int = 0
    failures: list = field(default_factory=list)


def _get_employee_attribute(db: Session, employee_id: str) -> Optional[str]:
    """查询员工当前属性"""
    row = db.query(EmployeeAttributeHistory).filter(
        EmployeeAttributeHistory.employee_id == employee_id,
        EmployeeAttributeHistory.is_current == True,
    ).first()
    return row.attribute_type if row else None


def _get_supervisor_id(db: Session, salesperson_id: str) -> Optional[str]:
    """查询业务员当前一级主管"""
    row = db.query(SupervisorRelationHistory).filter(
        SupervisorRelationHistory.salesperson_id == salesperson_id,
        SupervisorRelationHistory.is_current == True,
    ).first()
    return row.supervisor_id if row else None


def _get_second_supervisor_id(db: Session, salesperson_id: str) -> Optional[str]:
    """查询业务员当前二级主管"""
    row = db.query(SupervisorRelationHistory).filter(
        SupervisorRelationHistory.salesperson_id == salesperson_id,
        SupervisorRelationHistory.is_current == True,
    ).first()
    return row.second_supervisor_id if row else None


def _invalidate_current_snapshot(db: Session, customer_id: str) -> None:
    """将客户的当前快照标记失效"""
    db.query(CustomerCommissionSnapshot).filter(
        CustomerCommissionSnapshot.customer_id == customer_id,
        CustomerCommissionSnapshot.is_current == True,
    ).update({"is_current": False}, synchronize_session="fetch")


def _exists_in_business_table(db: Session, table: str, pk_field: str, pk_value: str) -> bool:
    """检查业务库中是否存在指定记录"""
    schema = get_settings().BUSINESS_DB_NAME
    sql = f"SELECT 1 FROM `{schema}`.`{table}` WHERE `{pk_field}` = :v LIMIT 1"
    return db.execute(text(sql), {"v": pk_value}).first() is not None


def reset_customer_attribution(
    db: Session,
    customer_id: str,
    new_salesperson_id: str,
    new_supervisor_id: Optional[str],
    salesperson_attribute: str,
    supervisor_attribute: Optional[str],
    reason: str,
    operator: str,
    new_second_supervisor_id: Optional[str] = None,
    remark: Optional[str] = None,
) -> CustomerCommissionSnapshot:
    """
    人工重置客户归属。

    旧快照 is_current → False，新快照按规则算比例。
    仅影响后续回款。
    """
    # 旧快照失效
    _invalidate_current_snapshot(db, customer_id)

    # 计算比例
    sp_rate, sv_rate, sv2_rate, _ = calc_commission_rates(
        new_salesperson_id, salesperson_attribute,
        new_supervisor_id, supervisor_attribute,
        new_second_supervisor_id,
    )

    now = datetime.now()
    snapshot = CustomerCommissionSnapshot(
        customer_id=customer_id,
        salesperson_id=new_salesperson_id,
        salesperson_attribute=salesperson_attribute,
        salesperson_rate=sp_rate,
        supervisor_id=new_supervisor_id,
        supervisor_attribute=supervisor_attribute,
        supervisor_rate=sv_rate,
        second_supervisor_id=new_second_supervisor_id,
        second_supervisor_rate=sv2_rate,
        remark=remark,
        is_complete=True,
        is_current=True,
        source="manual",
        is_manual_reset=True,
        reset_reason=reason,
        operator=operator,
        operated_at=now,
    )
    db.add(snapshot)
    db.flush()
    logger.info(f"客户 {customer_id} 归属已重置 (by {operator})")
    return snapshot


def complete_snapshot(
    db: Session,
    snapshot_id: int,
    salesperson_attribute: str,
    supervisor_attribute: Optional[str],
    operator: str,
    salesperson_rate_override: Optional[Decimal] = None,
    supervisor_rate_override: Optional[Decimal] = None,
    second_supervisor_rate_override: Optional[Decimal] = None,
) -> CustomerCommissionSnapshot:
    """
    补全不完整的归属快照。

    填入属性值，按规则计算提成比例（可由用户覆盖），标记 is_complete=True。
    """
    snapshot = db.query(CustomerCommissionSnapshot).filter(
        CustomerCommissionSnapshot.id == snapshot_id,
    ).first()
    if not snapshot:
        raise ValueError(f"快照 {snapshot_id} 不存在")
    if snapshot.is_complete:
        raise ValueError(f"快照 {snapshot_id} 已完整，无需补全")

    sp_rate, sv_rate, sv2_rate, _ = calc_commission_rates(
        snapshot.salesperson_id, salesperson_attribute,
        snapshot.supervisor_id, supervisor_attribute,
        snapshot.second_supervisor_id,
    )

    if salesperson_rate_override is not None:
        sp_rate = salesperson_rate_override
    if supervisor_rate_override is not None:
        sv_rate = supervisor_rate_override
    if second_supervisor_rate_override is not None:
        sv2_rate = second_supervisor_rate_override

    snapshot.salesperson_attribute = salesperson_attribute
    snapshot.salesperson_rate = sp_rate
    snapshot.supervisor_attribute = supervisor_attribute
    snapshot.supervisor_rate = sv_rate
    snapshot.second_supervisor_rate = sv2_rate
    snapshot.is_complete = True
    snapshot.operator = operator
    snapshot.operated_at = datetime.now()
    db.flush()

    logger.info(f"快照 {snapshot_id} (客户 {snapshot.customer_id}) 已补全 (by {operator})")
    return snapshot


@dataclass
class AutoMatchResult:
    matched: int = 0
    remaining: int = 0


def _get_first_receipt_date(db: Session, customer_id: str) -> Optional[date]:
    """查询客户的首次成交日期（首笔回款日期）"""
    from sqlalchemy import func
    row = db.query(func.min(OkkiReceipt.collection_date)).filter(
        OkkiReceipt.company_id == customer_id,
    ).scalar()
    if row and isinstance(row, datetime):
        return row.date()
    return row


def auto_match_incomplete_snapshots(db: Session, operator: str) -> AutoMatchResult:
    """
    自动匹配所有待补充快照。

    根据客户首次成交日期查找业务员在该时间点的属性，计算提成比例。
    """
    incomplete = db.query(CustomerCommissionSnapshot).filter(
        CustomerCommissionSnapshot.is_current == True,
        CustomerCommissionSnapshot.is_complete == False,
    ).all()

    matched = 0
    now = datetime.now()

    for snap in incomplete:
        first_receipt = _get_first_receipt_date(db, snap.customer_id)
        sp_attr = get_employee_attribute_at_date(db, snap.salesperson_id, first_receipt)
        if not sp_attr:
            continue

        sv_id = snap.supervisor_id or _get_supervisor_id(db, snap.salesperson_id)
        sv_attr = None
        if sv_id:
            sv_attr = get_employee_attribute_at_date(db, sv_id, first_receipt)

        sv2_id = snap.second_supervisor_id or _get_second_supervisor_id(db, snap.salesperson_id)

        sp_rate, sv_rate, sv2_rate, _ = calc_commission_rates(
            snap.salesperson_id, sp_attr, sv_id, sv_attr, sv2_id,
        )

        snap.salesperson_attribute = sp_attr
        snap.salesperson_rate = sp_rate
        snap.supervisor_id = sv_id
        snap.supervisor_attribute = sv_attr
        snap.supervisor_rate = sv_rate
        snap.second_supervisor_id = sv2_id
        snap.second_supervisor_rate = sv2_rate
        snap.is_complete = True
        snap.operator = operator
        snap.operated_at = now
        matched += 1

    db.flush()
    remaining = len(incomplete) - matched
    logger.info(f"自动匹配完成: 成功 {matched}, 剩余 {remaining} (by {operator})")
    return AutoMatchResult(matched=matched, remaining=remaining)


def import_snapshots_from_excel(db: Session, file_path: str, operator: str) -> ImportResult:
    """
    从 Excel 文件批量导入客户归属快照。

    Excel 模板列（从第2行开始，第1行为表头）：
      A: 客户ID (company_id)
      B: 业务员ID (user_id)
      C: 业务员属性 (开发/分配)
      D: 一级主管ID (user_id)
      E: 一级主管属性 (开发/分配)
      F: 二级主管ID (user_id)

    Args:
        db: Session
        file_path: Excel 文件路径
        operator: 操作人

    Returns:
        ImportResult 导入结果
    """
    result = ImportResult()

    wb = load_workbook(file_path, data_only=True)
    ws = wb.active

    attr_map = {"开发": "develop", "分配": "distribute"}

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or not row[0]:
            break
        result.total_rows += 1

        try:
            company_id = str(row[0]).strip()
            sp_id = str(row[1]).strip()
            sp_attr_raw = str(row[2]).strip()
            sv_id = str(row[3]).strip() if row[3] else None
            sv_attr_raw = str(row[4]).strip() if row[4] else None
            sv2_id = str(row[5]).strip() if len(row) > 5 and row[5] else None

            # 属性映射
            sp_attr = attr_map.get(sp_attr_raw, sp_attr_raw)
            sv_attr = attr_map.get(sv_attr_raw, sv_attr_raw) if sv_attr_raw else None

            # 校验 company_id
            if not _exists_in_business_table(db, "customer_info", "company_id", company_id):
                raise ValueError(f"客户 {company_id} 在业务库中不存在")

            # 校验业务员
            if not _exists_in_business_table(db, "user_basic", "user_id", sp_id):
                raise ValueError(f"业务员 {sp_id} 在业务库中不存在")

            # 校验主管
            if sv_id and not _exists_in_business_table(db, "user_basic", "user_id", sv_id):
                raise ValueError(f"一级主管 {sv_id} 在业务库中不存在")

            # 校验二级主管
            if sv2_id and not _exists_in_business_table(db, "user_basic", "user_id", sv2_id):
                raise ValueError(f"二级主管 {sv2_id} 在业务库中不存在")

            # 属性值校验
            if sp_attr not in ("develop", "distribute"):
                raise ValueError(f"业务员属性无效: '{sp_attr_raw}'，应为 开发/分配")
            if sv_attr and sv_attr not in ("develop", "distribute"):
                raise ValueError(f"主管属性无效: '{sv_attr_raw}'，应为 开发/分配")

            # 计算比例
            sp_rate, sv_rate, sv2_rate, _ = calc_commission_rates(sp_id, sp_attr, sv_id, sv_attr, sv2_id)

            # 旧快照失效
            _invalidate_current_snapshot(db, company_id)

            # 创建新快照
            snapshot = CustomerCommissionSnapshot(
                customer_id=company_id,
                salesperson_id=sp_id,
                salesperson_attribute=sp_attr,
                salesperson_rate=sp_rate,
                supervisor_id=sv_id,
                supervisor_attribute=sv_attr,
                supervisor_rate=sv_rate,
                second_supervisor_id=sv2_id,
                second_supervisor_rate=sv2_rate,
                is_complete=True,
                is_current=True,
                source="import",
                operator=operator,
                operated_at=datetime.now(),
            )
            db.add(snapshot)
            result.success += 1

        except Exception as e:
            err = f"第 {row_idx} 行: {e}"
            result.failures.append(err)
            result.failed += 1
            logger.warning(f"Excel导入失败 - {err}")

    db.flush()
    wb.close()
    logger.info(f"Excel导入完成: 成功 {result.success}, 失败 {result.failed}")
    return result


@dataclass
class RefreshResult:
    total: int = 0
    updated: int = 0
    skipped: int = 0


def refresh_snapshots_by_employees(
    db: Session, employee_ids: list[str], operator: str,
) -> RefreshResult:
    """
    根据员工属性历史记录，刷新这些员工名下所有当前客户快照的属性和比例。

    按客户首次成交日期匹配员工属性，重新计算提成比例。
    """
    result = RefreshResult()
    now = datetime.now()

    snapshots = db.query(CustomerCommissionSnapshot).filter(
        CustomerCommissionSnapshot.is_current == True,
        CustomerCommissionSnapshot.salesperson_id.in_(employee_ids),
    ).all()

    result.total = len(snapshots)

    for snap in snapshots:
        first_receipt = _get_first_receipt_date(db, snap.customer_id)
        sp_attr = get_employee_attribute_at_date(db, snap.salesperson_id, first_receipt)
        if not sp_attr:
            result.skipped += 1
            continue

        sv_id = snap.supervisor_id or _get_supervisor_id(db, snap.salesperson_id)
        sv_attr = None
        if sv_id:
            sv_attr = get_employee_attribute_at_date(db, sv_id, first_receipt)

        sv2_id = snap.second_supervisor_id or _get_second_supervisor_id(db, snap.salesperson_id)

        sp_rate, sv_rate, sv2_rate, _ = calc_commission_rates(
            snap.salesperson_id, sp_attr, sv_id, sv_attr, sv2_id,
        )

        snap.salesperson_attribute = sp_attr
        snap.salesperson_rate = sp_rate
        snap.supervisor_id = sv_id
        snap.supervisor_attribute = sv_attr
        snap.supervisor_rate = sv_rate
        snap.second_supervisor_id = sv2_id
        snap.second_supervisor_rate = sv2_rate
        snap.is_complete = True
        snap.operator = operator
        snap.operated_at = now
        result.updated += 1

    db.flush()
    logger.info(
        f"快照刷新完成: 共 {result.total}, 更新 {result.updated}, "
        f"跳过 {result.skipped} (by {operator})"
    )
    return result
