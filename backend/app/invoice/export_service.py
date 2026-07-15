"""Invoice export helpers."""

from io import BytesIO
from html import escape
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from app.invoice.models import Invoice
from app.invoice.service import summarize_items


def build_invoice_workbook(invoice: Invoice) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Invoice"

    ws.merge_cells("A1:K1")
    ws["A1"] = "COMMERCIAL INVOICE"
    ws["A1"].font = Font(size=18, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center")

    header_pairs = [
        ("Invoice No.", invoice.invoice_no, "Date", invoice.invoice_date),
        ("To", invoice.contact_name or invoice.customer_name, "From", invoice.sales_user_name or ""),
        ("TEL/Fax", invoice.contact_phone or "", "TEL", invoice.sales_phone or ""),
        ("E-mail", invoice.contact_email or "", "E-mail", invoice.sales_email or ""),
        ("Delivery address", invoice.delivery_address or "", "Express", invoice.express_channel or ""),
        ("Currency", invoice.currency, "Payment Term", invoice.payment_term or ""),
    ]
    for row_idx, (label_l, value_l, label_r, value_r) in enumerate(header_pairs, start=3):
        ws.cell(row=row_idx, column=1, value=label_l).font = Font(bold=True)
        ws.cell(row=row_idx, column=2, value=str(value_l or ""))
        ws.cell(row=row_idx, column=6, value=label_r).font = Font(bold=True)
        ws.cell(row=row_idx, column=7, value=str(value_r or ""))

    headers = [
        "Product_name", "Product", "Net Weight Grams", "Curl", "Model",
        "Color", "Length", "Quantity", "Price/Piece", "Discount", "TotalPrice",
    ]
    start_row = 10
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2F5597")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    hair_items = [item for item in invoice.items if item.product_kind != "accessory"]
    accessory_items = [item for item in invoice.items if item.product_kind == "accessory"]
    for row_idx, item in enumerate(hair_items, start=start_row + 1):
        values = [
            item.product_name,
            item.product_display,
            item.net_weight_grams,
            item.curl,
            item.model,
            item.color,
            item.length,
            item.quantity,
            float(item.price_per_piece or 0),
            float(item.discount_amount or 0),
            float(item.total_price or 0),
        ]
        for col, value in enumerate(values, start=1):
            ws.cell(row=row_idx, column=col, value=value)

    current_row = start_row + len(hair_items)
    if accessory_items:
        title_row = current_row + 2
        ws.merge_cells(start_row=title_row, start_column=1, end_row=title_row, end_column=8)
        ws.cell(row=title_row, column=1, value="Accessories").font = Font(size=13, bold=True)
        accessory_headers = [
            "Name", "Model", "Color", "Standard Price", "Customer Price",
            "Quantity", "Discount", "TotalPrice",
        ]
        accessory_header_row = title_row + 1
        for col, header in enumerate(accessory_headers, start=1):
            cell = ws.cell(row=accessory_header_row, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="2F5597")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for row_idx, item in enumerate(accessory_items, start=accessory_header_row + 1):
            values = [
                item.product_display or item.product_name,
                item.model,
                item.color,
                float(item.standard_price or 0),
                float(item.customer_price or 0),
                item.quantity,
                float(item.discount_amount or 0),
                float(item.total_price or 0),
            ]
            for col, value in enumerate(values, start=1):
                ws.cell(row=row_idx, column=col, value=value)
        current_row = accessory_header_row + len(accessory_items)

    summary = summarize_items(invoice)
    fee_rows = [
        ("Hair Price", float(_gross_hair_amount(invoice))),
        ("Hair Discount", float(summary["hair_discount"])),
        ("Accessory Amount", float(summary["accessory_amount"])),
        ("Accessory Discount", float(summary["accessory_discount"])),
        ("Packaging Quantity", int(invoice.packaging_quantity or 0)),
        ("Packaging", float(invoice.internal_accessory or 0)),
        ("Shipping Fee", float(invoice.shipping_fee or 0)),
        ("Handling Fee", float(invoice.surcharge_amount or 0)),
        ("Total", float(invoice.total_amount or 0)),
    ]

    total_row = current_row
    for label, amount in fee_rows:
        total_row += 1
        ws.cell(row=total_row, column=10, value=label).font = Font(bold=True)
        ws.cell(row=total_row, column=11, value=amount).font = Font(bold=(label == "Total"))

    thin = Side(style="thin", color="D9E2F3")
    for row in ws.iter_rows(min_row=start_row, max_row=total_row, min_col=1, max_col=11):
        for cell in row:
            cell.border = Border(top=thin, right=thin, bottom=thin, left=thin)
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    widths = [28, 18, 18, 12, 14, 14, 12, 12, 14, 14, 14]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + idx)].width = width

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream


def build_print_html(invoice: Invoice) -> str:
    rows = []
    for item in invoice.items:
        if item.product_kind == "accessory":
            continue
        rows.append(f"""
        <tr>
          <td>{escape(item.product_name or "")}</td>
          <td>{escape(item.product_display or "")}</td>
          <td>{escape(item.net_weight_grams or "")}</td>
          <td>{escape(item.curl or "")}</td>
          <td>{escape(item.model or "")}</td>
          <td>{escape(item.color or "")}</td>
          <td>{escape(item.length or "")}</td>
          <td class="num">{item.quantity}</td>
          <td class="num">{item.price_per_piece or ""}</td>
          <td class="num">{item.discount_amount or 0}</td>
          <td class="num">{item.total_price or ""}</td>
        </tr>
        """)
    accessory_rows = []
    for item in invoice.items:
        if item.product_kind != "accessory":
            continue
        accessory_rows.append(f"""
        <tr>
          <td>{escape(item.product_display or item.product_name or "")}</td>
          <td>{escape(item.model or "")}</td>
          <td>{escape(item.color or "")}</td>
          <td class="num">{"" if item.standard_price is None else item.standard_price}</td>
          <td class="num">{"" if item.customer_price is None else item.customer_price}</td>
          <td class="num">{item.quantity}</td>
          <td class="num">{item.discount_amount or 0}</td>
          <td class="num">{item.total_price or ""}</td>
        </tr>
        """)
    accessory_section = ""
    if accessory_rows:
        accessory_section = f"""
  <h2>Accessories</h2>
  <table>
    <thead>
      <tr><th>Name</th><th>Model</th><th>Color</th><th>Standard Price</th><th>Customer Price</th><th>Quantity</th><th>Discount</th><th>TotalPrice</th></tr>
    </thead>
    <tbody>{''.join(accessory_rows)}</tbody>
  </table>
        """
    summary = summarize_items(invoice)
    return f"""<!doctype html>
<html>
<head>
  <meta charset="utf-8" />
  <title>{escape(invoice.invoice_no)} Invoice</title>
  <style>
    body {{ font-family: Arial, sans-serif; margin: 32px; color: #1f2937; }}
    h1 {{ text-align: center; letter-spacing: 0; }}
    .meta {{ display: grid; grid-template-columns: 140px 1fr; gap: 8px; margin: 24px 0; }}
    table {{ width: 100%; border-collapse: collapse; font-size: 12px; }}
    th {{ background: #2f5597; color: white; }}
    th, td {{ border: 1px solid #d9e2f3; padding: 8px; text-align: left; }}
    .num {{ text-align: right; }}
    .total {{ margin-top: 16px; text-align: right; font-weight: 700; }}
    @media print {{ button {{ display: none; }} }}
  </style>
</head>
<body>
  <button onclick="window.print()">Print / Save as PDF</button>
  <h1>COMMERCIAL INVOICE</h1>
  <div class="meta">
    <strong>Invoice No.</strong><span>{escape(invoice.invoice_no)}</span>
    <strong>Date</strong><span>{invoice.invoice_date}</span>
    <strong>To</strong><span>{escape(invoice.contact_name or invoice.customer_name or "")}</span>
    <strong>From</strong><span>{escape(invoice.sales_user_name or "")}</span>
    <strong>TEL/Fax</strong><span>{escape(invoice.contact_phone or "")}</span>
    <strong>TEL</strong><span>{escape(invoice.sales_phone or "")}</span>
    <strong>E-mail</strong><span>{escape(invoice.contact_email or "")}</span>
    <strong>Delivery address</strong><span>{escape(invoice.delivery_address or "")}</span>
    <strong>Express</strong><span>{escape(invoice.express_channel or "")}</span>
    <strong>Currency</strong><span>{escape(invoice.currency or "")}</span>
    <strong>Payment Term</strong><span>{escape(invoice.payment_term or "")}</span>
  </div>
  <table>
    <thead>
      <tr>
        <th>Product_name</th><th>Product</th><th>Net Weight Grams</th><th>Curl</th><th>Model</th>
        <th>Color</th><th>Length</th><th>Quantity</th><th>Price/Piece</th><th>Discount</th><th>TotalPrice</th>
      </tr>
    </thead>
    <tbody>{''.join(rows)}</tbody>
  </table>
  {accessory_section}
  <div class="total">
    <div>Hair Price: {invoice.currency} {_gross_hair_amount(invoice)}</div>
    <div>Hair Discount: {invoice.currency} {summary['hair_discount']}</div>
    <div>Accessory Amount: {invoice.currency} {summary['accessory_amount']}</div>
    <div>Accessory Discount: {invoice.currency} {summary['accessory_discount']}</div>
    <div>Packaging Quantity: {invoice.packaging_quantity or 0}</div>
    <div>Packaging: {invoice.currency} {invoice.internal_accessory or 0}</div>
    <div>Shipping Fee: {invoice.currency} {invoice.shipping_fee or 0}</div>
    <div>Handling Fee: {invoice.currency} {invoice.surcharge_amount or 0}</div>
    <div>Total: {invoice.currency} {invoice.total_amount}</div>
  </div>
</body>
</html>"""


def build_invoice_pdf(invoice: Invoice) -> BytesIO:
    pages: list[list[tuple[str, int, int, int]]] = []

    def new_page(*, first: bool = False) -> tuple[list[tuple[str, int, int, int]], int]:
        page: list[tuple[str, int, int, int]] = []
        pages.append(page)
        if first:
            page.extend([
                ("COMMERCIAL INVOICE", 220, 800, 18),
                (f"Invoice No.: {invoice.invoice_no}", 50, 760, 11),
                (f"Date: {invoice.invoice_date}", 50, 742, 11),
                (f"Customer: {invoice.customer_name}", 50, 724, 11),
                (f"Currency: {invoice.currency}", 50, 706, 11),
            ])
            return page, 668
        page.append((f"COMMERCIAL INVOICE - {invoice.invoice_no} (continued)", 50, 800, 12))
        return page, 770

    page, y = new_page(first=True)
    hair_items = [item for item in invoice.items if item.product_kind != "accessory"]
    accessory_items = [item for item in invoice.items if item.product_kind == "accessory"]

    page.append(("HAIR PRODUCTS", 50, y, 11))
    y -= 18
    page.append(("Product | Model | Color | Length | Qty | Price | Discount | Total", 50, y, 9))
    y -= 18
    for item in hair_items:
        if y < 64:
            page, y = new_page()
            page.append(("HAIR PRODUCTS (continued)", 50, y, 11))
            y -= 18
            page.append(("Product | Model | Color | Length | Qty | Price | Discount | Total", 50, y, 9))
            y -= 18
        product = (item.product_display or item.product_name or "")[:24]
        row = (
            f"{product} | {item.model or ''} | {item.color or ''} | {item.length or ''} | "
            f"{item.quantity} | {item.price_per_piece or ''} | {item.discount_amount or 0} | {item.total_price or ''}"
        )
        page.append((row[:105], 50, y, 8))
        y -= 16

    if accessory_items:
        if y < 100:
            page, y = new_page()
        else:
            y -= 8
        page.append(("ACCESSORIES", 50, y, 11))
        y -= 18
        page.append(("Name | Model | Color | Standard Price | Customer Price | Quantity | Discount | TotalPrice", 50, y, 8))
        y -= 18
        for item in accessory_items:
            if y < 64:
                page, y = new_page()
                page.append(("ACCESSORIES (continued)", 50, y, 11))
                y -= 18
                page.append(("Name | Model | Color | Standard Price | Customer Price | Quantity | Discount | TotalPrice", 50, y, 8))
                y -= 18
            name = (item.product_display or item.product_name or "")[:18]
            model = (item.model or "")[:10]
            color = (item.color or "")[:10]
            row = (
                f"{name} | {model} | {color} | "
                f"{'' if item.standard_price is None else item.standard_price} | "
                f"{'' if item.customer_price is None else item.customer_price} | {item.quantity} | "
                f"{item.discount_amount or 0} | {item.total_price or ''}"
            )
            page.append((row, 50, y, 7))
            y -= 16

    item_summary = summarize_items(invoice)
    summary = [
        ("Hair Price", item_summary["hair_amount"]),
        ("Hair Discount", item_summary["hair_discount"]),
        ("Accessory Amount", item_summary["accessory_amount"]),
        ("Accessory Discount", item_summary["accessory_discount"]),
        ("Packaging Quantity", invoice.packaging_quantity or 0),
        ("Packaging", invoice.internal_accessory or 0),
        ("Shipping Fee", invoice.shipping_fee or 0),
        ("Handling Fee", invoice.surcharge_amount or 0),
        ("Total", invoice.total_amount or 0),
    ]
    required_height = len(summary) * 16 + 12
    if y - required_height < 40:
        page, y = new_page()
        page.append(("SUMMARY", 360, y, 11))
        y -= 20
    else:
        y -= 8
    for label, amount in summary:
        value = str(amount) if label == "Packaging Quantity" else f"{invoice.currency} {amount}"
        page.append((f"{label}: {value}", 330, y, 10 if label != "Total" else 12))
        y -= 16

    return _write_pdf_pages(pages)


def _gross_hair_amount(invoice: Invoice) -> Decimal:
    return summarize_items(invoice)["hair_amount"]


def _pdf_escape(value: str) -> str:
    return str(value).replace("\\", "\\\\").replace("(", "\\(").replace(")", "\\)")


def _write_pdf_pages(pages: list[list[tuple[str, int, int, int]]]) -> BytesIO:
    page_count = len(pages)
    page_ids = list(range(3, 3 + page_count))
    font_id = 3 + page_count
    content_ids = list(range(font_id + 1, font_id + 1 + page_count))
    kids = " ".join(f"{page_id} 0 R" for page_id in page_ids)
    objects: list[bytes] = [
        b"<< /Type /Catalog /Pages 2 0 R >>",
        f"<< /Type /Pages /Kids [{kids}] /Count {page_count} >>".encode("ascii"),
    ]
    for content_id in content_ids:
        objects.append(
            f"<< /Type /Page /Parent 2 0 R /MediaBox [0 0 595 842] "
            f"/Resources << /Font << /F1 {font_id} 0 R >> >> /Contents {content_id} 0 R >>".encode("ascii")
        )
    objects.append(b"<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>")
    for lines in pages:
        commands = ["BT"]
        for value, x, y_pos, size in lines:
            commands.append(f"/F1 {size} Tf")
            commands.append(f"1 0 0 1 {x} {y_pos} Tm")
            commands.append(f"({_pdf_escape(value)}) Tj")
        commands.append("ET")
        content = "\n".join(commands).encode("latin-1", errors="replace")
        objects.append(
            b"<< /Length " + str(len(content)).encode("ascii") + b" >>\nstream\n"
            + content + b"\nendstream"
        )
    return _write_pdf(objects)


def _write_pdf(objects: list[bytes]) -> BytesIO:
    stream = BytesIO()
    stream.write(b"%PDF-1.4\n")
    offsets = [0]
    for index, obj in enumerate(objects, start=1):
        offsets.append(stream.tell())
        stream.write(f"{index} 0 obj\n".encode("ascii"))
        stream.write(obj)
        stream.write(b"\nendobj\n")
    xref_offset = stream.tell()
    stream.write(f"xref\n0 {len(objects) + 1}\n".encode("ascii"))
    stream.write(b"0000000000 65535 f \n")
    for offset in offsets[1:]:
        stream.write(f"{offset:010d} 00000 n \n".encode("ascii"))
    stream.write(
        f"trailer\n<< /Size {len(objects) + 1} /Root 1 0 R >>\nstartxref\n{xref_offset}\n%%EOF".encode("ascii")
    )
    stream.seek(0)
    return stream
