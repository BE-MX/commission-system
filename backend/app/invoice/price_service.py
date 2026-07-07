"""Pricing engine for order invoices.

Standard price is a matrix keyed by (series_grade, length, weight_unit,
color_type) mirroring the business price sheet, so stock products and
free-form production products share one price source. Customer price is
derived from a per-customer rule: fixed signed delta or signed percent.
"""

import logging
import re
from decimal import Decimal, ROUND_HALF_UP
from io import BytesIO

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.invoice.models import CustomerPriceRule, PriceColorType, StdPrice

logger = logging.getLogger(__name__)

COLOR_TYPES = ("solid", "piano", "ombre", "balayage")

_QUOTE_CHARS = "“”‘’′″\"'`"


# ── normalization ─────────────────────────────────────────────

def normalize_text(value: str | None) -> str:
    if not value:
        return ""
    text = str(value).strip()
    for ch in _QUOTE_CHARS:
        text = text.replace(ch, "")
    text = re.sub(r"\s+", " ", text)
    return text.lower()


def normalize_length(value: str | None) -> str:
    """'16\"' / '16寸' / ' 16 ' -> '16'."""
    text = normalize_text(value)
    return text.replace("寸", "").replace("inch", "").strip()


def normalize_color(value: str | None) -> str:
    return normalize_text(value).lstrip("#")


def make_match_key(product_display: str, model: str | None, color: str, size: str, unit: str) -> str:
    return "|".join([
        normalize_text(product_display),
        normalize_text(model),
        normalize_color(color),
        normalize_length(size),
        normalize_text(unit),
    ])


# ── color type ────────────────────────────────────────────────

def resolve_color_type(db: Session, color: str) -> tuple[str | None, str]:
    """Return (color_type, source) where source is exact/inferred/missing."""
    code = normalize_color(color)
    if not code:
        return None, "missing"
    row = (
        db.query(PriceColorType)
        .filter(PriceColorType.color_code.in_([code, f"#{code}"]))
        .first()
    )
    if row:
        return row.color_type, "exact"
    inferred = _infer_color_type(code)
    return inferred, "inferred" if inferred else "missing"


def _infer_color_type(code: str) -> str | None:
    """Heuristic mirroring the color mapping sheet's naming convention."""
    upper = code.upper()
    if "TP" in upper:
        return "balayage"
    # piano codes look like P2/8, P18/60: P followed by a digit (not English words like "Pink")
    if re.match(r"^P\d", upper):
        return "piano"
    # ombre codes look like 4T60 / 5AT60 / 18T60: digits(+letters) then T then digits
    if re.match(r"^\d+[A-Z]*T\d", upper):
        return "ombre"
    if re.match(r"^\d+[A-Z]*$", upper):
        return "solid"
    return None


# ── price resolution ──────────────────────────────────────────

def resolve_price(
    db: Session,
    *,
    customer_id: str | None,
    product_display: str,
    length: str,
    unit: str,
    color: str,
) -> dict:
    color_type, color_source = resolve_color_type(db, color)
    standard = _lookup_std_price(db, product_display, length, unit, color_type) if color_type else None

    rule = None
    if customer_id:
        rule = (
            db.query(CustomerPriceRule)
            .filter(CustomerPriceRule.customer_id == str(customer_id), CustomerPriceRule.enabled == 1)
            .first()
        )

    customer_price = None
    if standard is not None:
        customer_price = apply_rule(standard.price, rule)

    return {
        "standard_price": standard.price if standard else None,
        "customer_price": customer_price,
        "currency": standard.currency if standard else "USD",
        "color_type": color_type,
        "color_type_source": color_source,
        "rule": _serialize_rule(rule) if rule else None,
        "rule_desc": describe_rule(rule),
        "price_source": "customer_rule" if customer_price is not None else "missing_std",
    }


def apply_rule(standard_price: Decimal, rule: CustomerPriceRule | None) -> Decimal:
    price = Decimal(standard_price)
    if rule:
        value = Decimal(rule.adjust_value)
        if rule.adjust_type == "percent":
            price = price * (Decimal("1") + value / Decimal("100"))
        else:
            price = price + value
    if price < 0:
        price = Decimal("0")
    return price.quantize(Decimal("0.0001"), rounding=ROUND_HALF_UP)


def describe_rule(rule: CustomerPriceRule | None) -> str:
    if not rule:
        return "无客户规则，按标准价"
    sign = "+" if rule.adjust_value >= 0 else ""
    if rule.adjust_type == "percent":
        return f"标准价 {sign}{rule.adjust_value:.2f}%"
    return f"标准价 {sign}{rule.adjust_value:.2f}"


def _lookup_std_price(db: Session, product_display: str, length: str, unit: str, color_type: str) -> StdPrice | None:
    """Prefix-tolerant series match: 'Super Double Drawn Genius' matches
    display 'Super Double Drawn Genius Weft' and vice versa.

    Money rule: if more than one distinct series matches, the price is
    ambiguous — return None (missing) instead of silently picking one.
    """
    candidates = (
        db.query(StdPrice)
        .filter(
            StdPrice.length == normalize_length(length),
            StdPrice.weight_unit == normalize_text(unit),
            StdPrice.color_type == color_type,
        )
        .all()
    )
    display_norm = normalize_text(product_display)
    matched: list[StdPrice] = []
    for row in candidates:
        series_norm = normalize_text(row.series_grade)
        if not series_norm:
            continue
        if display_norm.startswith(series_norm) or series_norm.startswith(display_norm):
            matched.append(row)
    if not matched:
        return None
    distinct_series = {normalize_text(r.series_grade) for r in matched}
    if len(distinct_series) > 1:
        logger.warning("标准价系列匹配歧义 display=%s 命中 %s", product_display, sorted(distinct_series))
        return None
    return matched[0]


# ── admin CRUD ────────────────────────────────────────────────

def list_std_prices(db: Session, *, series_grade: str | None = None) -> list[dict]:
    query = db.query(StdPrice)
    if series_grade:
        query = query.filter(StdPrice.series_grade == series_grade)
    rows = query.order_by(StdPrice.series_grade, StdPrice.weight_unit, StdPrice.length, StdPrice.color_type).all()
    return [_serialize_std(r) for r in rows]


def upsert_std_price(
    db: Session,
    *,
    series_grade: str,
    length: str,
    weight_unit: str,
    color_type: str,
    price: Decimal,
    currency: str = "USD",
    user_id: int | None = None,
    price_id: int | None = None,
) -> StdPrice:
    if color_type not in COLOR_TYPES:
        raise ValueError(f"color_type 必须是 {COLOR_TYPES}")
    length_n = normalize_length(length)
    unit_n = normalize_text(weight_unit)
    if price_id:
        # explicit edit: mutate that row in place so changing a key field
        # does not leave the old cell behind in the matrix
        row = db.query(StdPrice).filter(StdPrice.id == price_id).first()
        if not row:
            raise ValueError("价格记录不存在")
        duplicate = (
            db.query(StdPrice)
            .filter(
                StdPrice.id != price_id,
                StdPrice.series_grade == series_grade.strip(),
                StdPrice.length == length_n,
                StdPrice.weight_unit == unit_n,
                StdPrice.color_type == color_type,
            )
            .first()
        )
        if duplicate:
            raise ValueError("同键价格已存在，请直接编辑那一行")
        row.series_grade = series_grade.strip()
        row.length = length_n
        row.weight_unit = unit_n
        row.color_type = color_type
        row.price = price
        row.currency = currency
        row.updated_by = user_id
        return row
    row = (
        db.query(StdPrice)
        .filter(
            StdPrice.series_grade == series_grade.strip(),
            StdPrice.length == length_n,
            StdPrice.weight_unit == unit_n,
            StdPrice.color_type == color_type,
        )
        .first()
    )
    if row:
        row.price = price
        row.currency = currency
        row.updated_by = user_id
    else:
        row = StdPrice(
            series_grade=series_grade.strip(),
            length=length_n,
            weight_unit=unit_n,
            color_type=color_type,
            price=price,
            currency=currency,
            updated_by=user_id,
        )
        db.add(row)
    return row


def delete_std_price(db: Session, price_id: int) -> bool:
    row = db.query(StdPrice).filter(StdPrice.id == price_id).first()
    if not row:
        return False
    db.delete(row)
    return True


def list_color_types(db: Session) -> list[dict]:
    rows = db.query(PriceColorType).order_by(PriceColorType.color_type, PriceColorType.color_code).all()
    return [{"id": r.id, "color_code": r.color_code, "color_type": r.color_type} for r in rows]


def upsert_color_type(db: Session, *, color_code: str, color_type: str) -> PriceColorType:
    if color_type not in COLOR_TYPES:
        raise ValueError(f"color_type 必须是 {COLOR_TYPES}")
    code = normalize_color(color_code)
    if not code:
        raise ValueError("color_code 不能为空")
    row = db.query(PriceColorType).filter(PriceColorType.color_code == code).first()
    if row:
        row.color_type = color_type
    else:
        row = PriceColorType(color_code=code, color_type=color_type)
        db.add(row)
    return row


def delete_color_type(db: Session, entry_id: int) -> bool:
    row = db.query(PriceColorType).filter(PriceColorType.id == entry_id).first()
    if not row:
        return False
    db.delete(row)
    return True


def list_customer_rules(db: Session, *, keyword: str | None = None) -> list[dict]:
    query = db.query(CustomerPriceRule)
    if keyword:
        like = f"%{keyword}%"
        query = query.filter(
            (CustomerPriceRule.customer_id.like(like)) | (CustomerPriceRule.customer_name.like(like))
        )
    rows = query.order_by(CustomerPriceRule.updated_at.desc()).all()
    return [_serialize_rule(r) for r in rows]


def upsert_customer_rule(
    db: Session,
    *,
    customer_id: str,
    customer_name: str | None,
    adjust_type: str,
    adjust_value: Decimal,
    enabled: bool = True,
    preferred_template: str | None = None,
    remark: str | None = None,
    user_id: int | None = None,
) -> CustomerPriceRule:
    if adjust_type not in ("fixed", "percent"):
        raise ValueError("adjust_type 必须是 fixed 或 percent")
    row = db.query(CustomerPriceRule).filter(CustomerPriceRule.customer_id == str(customer_id)).first()
    if row:
        row.customer_name = customer_name or row.customer_name
        row.adjust_type = adjust_type
        row.adjust_value = adjust_value
        row.enabled = 1 if enabled else 0
        row.preferred_template = preferred_template
        row.remark = remark
        row.updated_by = user_id
    else:
        row = CustomerPriceRule(
            customer_id=str(customer_id),
            customer_name=customer_name,
            adjust_type=adjust_type,
            adjust_value=adjust_value,
            enabled=1 if enabled else 0,
            preferred_template=preferred_template,
            remark=remark,
            updated_by=user_id,
        )
        db.add(row)
    return row


def delete_customer_rule(db: Session, rule_id: int) -> bool:
    row = db.query(CustomerPriceRule).filter(CustomerPriceRule.id == rule_id).first()
    if not row:
        return False
    db.delete(row)
    return True


def get_customer_rule(db: Session, customer_id: str) -> dict | None:
    row = (
        db.query(CustomerPriceRule)
        .filter(CustomerPriceRule.customer_id == str(customer_id), CustomerPriceRule.enabled == 1)
        .first()
    )
    return _serialize_rule(row) if row else None


# ── Excel import (business price sheet format) ────────────────

_PRICE_LIST_MARK = "price list----"


def import_price_workbook(db: Session, content: bytes, user_id: int | None = None) -> dict:
    """Import the business 基础价格表 workbook.

    Sheet '价格表': blocks of [Price List----<series>] > <grade> > header >
    rows (length, weight, 4 prices for solid/piano/ombre/balayage).
    Sheet '颜色对照表': 4 columns of color codes, one per color type.
    """
    wb = load_workbook(BytesIO(content), data_only=True)
    prices_imported = 0
    colors_imported = 0
    skipped: list[str] = []

    for ws in wb.worksheets:
        header = [str(c or "").strip().lower() for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())]
        if header and header[0].startswith("solid"):
            colors_imported += _import_color_sheet(db, ws)
            continue
        imported, sheet_skipped = _import_price_sheet(db, ws, user_id)
        prices_imported += imported
        skipped.extend(sheet_skipped)

    return {"prices_imported": prices_imported, "colors_imported": colors_imported, "skipped": skipped[:50]}


def _import_price_sheet(db: Session, ws, user_id: int | None) -> tuple[int, list[str]]:
    series: str | None = None
    grade: str | None = None
    imported = 0
    skipped: list[str] = []
    for row in ws.iter_rows(values_only=True):
        cells = [c for c in row]
        first = str(cells[0] or "").strip()
        if not first:
            continue
        lowered = first.lower()
        if _PRICE_LIST_MARK in lowered:
            series = first.split("----", 1)[-1].strip()
            grade = None
            continue
        if lowered.startswith("length"):
            continue
        rest = [c for c in cells[1:6] if c not in (None, "")]
        if not rest:
            grade = first
            continue
        if series is None or grade is None:
            skipped.append(f"{ws.title}: 行 '{first}' 缺少系列/工艺档上下文")
            continue
        prices = list(cells[2:6])
        if len([p for p in prices if p not in (None, "")]) < 4:
            skipped.append(f"{ws.title}: '{first}' 价格列不完整")
            continue
        series_grade = f"{grade} {series}".strip()
        weight = str(cells[1] or "").strip()
        try:
            for color_type, price in zip(COLOR_TYPES, prices):
                upsert_std_price(
                    db,
                    series_grade=series_grade,
                    length=first,
                    weight_unit=weight,
                    color_type=color_type,
                    price=Decimal(str(price)),
                    user_id=user_id,
                )
                imported += 1
        except Exception as exc:  # noqa: BLE001 - keep importing remaining rows
            logger.warning("价格表导入行失败 %s/%s: %s", series_grade, first, exc)
            print(f"[invoice] price import row failed {series_grade}/{first}: {exc}", flush=True)
            skipped.append(f"{series_grade} {first}: {exc}")
    return imported, skipped


def _import_color_sheet(db: Session, ws) -> int:
    imported = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        for idx, color_type in enumerate(COLOR_TYPES):
            value = row[idx] if idx < len(row) else None
            code = str(value or "").strip()
            if not code:
                continue
            upsert_color_type(db, color_code=code, color_type=color_type)
            imported += 1
    return imported


# ── serializers ───────────────────────────────────────────────

def _serialize_std(row: StdPrice) -> dict:
    return {
        "id": row.id,
        "series_grade": row.series_grade,
        "length": row.length,
        "weight_unit": row.weight_unit,
        "color_type": row.color_type,
        "price": row.price,
        "currency": row.currency,
        "updated_at": row.updated_at,
    }


def _serialize_rule(row: CustomerPriceRule) -> dict:
    return {
        "id": row.id,
        "customer_id": row.customer_id,
        "customer_name": row.customer_name,
        "adjust_type": row.adjust_type,
        "adjust_value": row.adjust_value,
        "enabled": bool(row.enabled),
        "preferred_template": row.preferred_template,
        "remark": row.remark,
        "updated_at": row.updated_at,
    }
