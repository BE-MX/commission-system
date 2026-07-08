"""回款日期修复：把外部维护的只读业务镜像 okki_receipts.collection_date
按田雯的工作表纠正为正确日期。

匹配锚点（经线上真实数据验证后确定）：
  客户名(B) + 订单总额USD(D) → okki_orders 唯一订单 → 改该订单名下回款单日期。
  E(已回款额) 不参与匹配，因回款单 amount_usd 被扣手续费，与 E 不相等。
  单笔订单：唯一回款单日期 = A(订单日期即首笔回款日)。
  多笔订单：按回款单号(cash_collection_no)升序 = 时间序，依次分配 A / H / J / L。

流程：解析 → 只读试跑分类(build_plan) → 前端预览确认 → 写库(apply_changes)。
写业务镜像的每一次改动都落审计表 ark_receipt_repair_log(old→new)，可回滚。

写库前提（亮哥 2026-07-08 确认 + SHOW GRANTS 验证）：本连接账号对 lsordertest 有
UPDATE 权限；OKKI 同步作业不回写已存在回款单的 collection_date，故本工具是该字段
的唯一写者，修复不会被下次同步覆盖。此假设若变更，本功能即失效——改前先复核。
"""

import logging
import re
import uuid
from datetime import date, datetime
from io import BytesIO
from typing import Optional

from openpyxl import Workbook, load_workbook
from sqlalchemy import text
from sqlalchemy.orm import Session

from app.core.config import get_settings
from app.invoice.models import ReceiptRepairLog

logger = logging.getLogger("invoice.receipt_repair")

# Excel 列(1-based)：A首笔日期 B客户名 C国家 D总额 E首笔额 ... H/J/L后续日期 I/K/M后续额
COL_DATE_FIRST = 1
COL_COMPANY = 2
COL_TOTAL = 4
COL_AMOUNT_FIRST = 5
# (amount_col, date_col) for 2nd/3rd/4th receipts
SUBSEQUENT = [(9, 8), (11, 10), (13, 12)]
HEADER_ROWS = 2  # 第1、2行是表头

REASON_LABELS = {
    "company_not_found": "客户名在回款表中不存在",
    "no_order_amount_match": "该客户下无总额匹配的订单",
    "multi_order_match": "匹配到多个订单，无法唯一确定",
    "order_no_receipt": "订单存在但系统无回款记录",
    "receipt_count_mismatch": "回款笔数与系统不一致",
    "no_target_date": "Excel 未提供有效回款日期",
}


def _r2(v) -> Optional[float]:
    if v is None or v == "":
        return None
    try:
        return round(float(v), 2)
    except (TypeError, ValueError):
        return None


def _to_date(v) -> Optional[date]:
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d"):
        try:
            return datetime.strptime(str(v).strip()[:10], fmt).date()
        except ValueError:
            continue
    return None


def _norm_name(s) -> str:
    """公司名归一：去首尾空格 + 内部连续空白折叠为单空格 + 小写。两侧一致才能匹配。"""
    return re.sub(r"\s+", " ", str(s or "").strip()).lower()


def _receipt_sort_key(no):
    s = "" if no is None else str(no).strip()
    return (0, int(s)) if s.isdigit() else (1, s)


def parse_workbook(content: bytes) -> list[dict]:
    """解析工作表为行列表。每行含 company / order_total / 首笔起的 (date, amount) 序列，
    并保留原始单元格用于导出无法匹配清单。"""
    wb = load_workbook(BytesIO(content), data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows: list[dict] = []
    for r in range(HEADER_ROWS + 1, ws.max_row + 1):
        company = ws.cell(row=r, column=COL_COMPANY).value
        if company is None or not str(company).strip():
            continue
        first_amt = _r2(ws.cell(row=r, column=COL_AMOUNT_FIRST).value)
        # 序列：首笔用 A 列日期 + E 列金额；后续用 H/J/L + I/K/M
        seq = [(_to_date(ws.cell(row=r, column=COL_DATE_FIRST).value), first_amt)]
        for amt_c, dt_c in SUBSEQUENT:
            amt = _r2(ws.cell(row=r, column=amt_c).value)
            if amt is not None:
                seq.append((_to_date(ws.cell(row=r, column=dt_c).value), amt))
        raw = {chr(64 + c): ws.cell(row=r, column=c).value for c in range(1, 14)}
        rows.append({
            "excel_row": r,
            "company": str(company).strip(),
            "order_total": _r2(ws.cell(row=r, column=COL_TOTAL).value),
            "seq": seq,               # [(date, amount), ...] 时间序
            "n_receipts": len(seq),
            "raw": raw,
        })
    return rows


def _match_row(row: dict, name2ids: dict, orders_by_cid: dict, receipts_by_order: dict) -> tuple[str, dict]:
    """纯匹配函数（不碰 DB），返回 (status, payload)。status 见 REASON_LABELS 或
    'will_change' / 'already_ok'。"""
    ids = name2ids.get(_norm_name(row["company"]))
    if not ids:
        return "company_not_found", {}

    total = row["order_total"]
    candidates = [
        oid for cid in ids for oid, amt in orders_by_cid.get(cid, {}).items()
        if total is not None and amt is not None and abs(amt - total) < 0.005
    ]
    candidates = list(dict.fromkeys(candidates))
    if not candidates:
        return "no_order_amount_match", {}
    if len(candidates) > 1:
        return "multi_order_match", {"order_count": len(candidates)}

    order_id = candidates[0]
    receipts = receipts_by_order.get(order_id, [])
    order_no = receipts[0]["order_no"] if receipts else None
    if not receipts:
        return "order_no_receipt", {"order_id": str(order_id)}
    if len(receipts) != row["n_receipts"]:
        return "receipt_count_mismatch", {
            "db_count": len(receipts), "excel_count": row["n_receipts"], "order_no": order_no,
        }

    receipts_sorted = sorted(receipts, key=lambda x: _receipt_sort_key(x["cash_collection_no"]))
    changes, unchanged = [], 0
    have_target = False
    for (target_date, excel_amt), rec in zip(row["seq"], receipts_sorted):
        cur = rec["collection_date"]
        if target_date is None:
            unchanged += 1
            continue
        have_target = True
        detail = {
            "cash_collection_id": str(rec["cash_collection_id"]),
            "cash_collection_no": rec["cash_collection_no"],
            "receipt_amount": float(rec["amount_usd"]) if rec["amount_usd"] is not None else None,
            "excel_amount": excel_amt,
            "old_date": cur.isoformat() if cur else None,
            "new_date": target_date.isoformat(),
        }
        if cur == target_date:
            unchanged += 1
        else:
            changes.append(detail)

    if not have_target:
        return "no_target_date", {"order_no": order_no}
    if not changes:
        return "already_ok", {"order_no": order_no}
    return "will_change", {
        "order_no": order_no,
        "order_id": str(order_id),
        "changes": changes,
    }


def build_plan(db: Session, rows: list[dict], source_file: str = "") -> dict:
    """只读试跑：从业务库拉取 → 分类为 待修改/已正确/无法匹配。不写任何库。"""
    schema = get_settings().BUSINESS_DB_NAME
    names = sorted({_norm_name(r["company"]) for r in rows})

    # Q1: 归一化公司名(去首尾+折叠内部空白+小写) -> company_ids。
    # SQL 侧用 REGEXP_REPLACE 折叠空白，与 Python _norm_name 口径一致（MySQL 8）。
    name2ids: dict[str, set] = {}
    for i in range(0, len(names), 500):
        chunk = names[i:i + 500]
        ph = ",".join(f":n{j}" for j in range(len(chunk)))
        pr = {f"n{j}": v for j, v in enumerate(chunk)}
        q = text(f"SELECT DISTINCT company_name, company_id FROM `{schema}`.okki_receipts "
                 f"WHERE LOWER(REGEXP_REPLACE(TRIM(company_name), '\\\\s+', ' ')) IN ({ph})")
        for cn, cid in db.execute(q, pr).all():
            if cn is not None:
                name2ids.setdefault(_norm_name(cn), set()).add(cid)

    all_ids = [i for s in name2ids.values() for i in s if i is not None]

    # Q2: orders by company_id -> {order_id: amount}
    orders_by_cid: dict = {}
    for i in range(0, len(all_ids), 500):
        chunk = all_ids[i:i + 500]
        ph = ",".join(f":i{j}" for j in range(len(chunk)))
        pr = {f"i{j}": v for j, v in enumerate(chunk)}
        q = text(f"SELECT order_id, company_id, amount_usd FROM `{schema}`.okki_orders "
                 f"WHERE company_id IN ({ph})")
        for oid, cid, amt in db.execute(q, pr).all():
            orders_by_cid.setdefault(cid, {})[oid] = _r2(amt)

    # Q3: receipts by order_id
    receipts_by_order: dict = {}
    for i in range(0, len(all_ids), 500):
        chunk = all_ids[i:i + 500]
        ph = ",".join(f":i{j}" for j in range(len(chunk)))
        pr = {f"i{j}": v for j, v in enumerate(chunk)}
        q = text(f"SELECT cash_collection_id, cash_collection_no, amount_usd, collection_date, "
                 f"order_id, order_no FROM `{schema}`.okki_receipts WHERE company_id IN ({ph})")
        for d in db.execute(q, pr).mappings().all():
            receipts_by_order.setdefault(d["order_id"], []).append({
                "cash_collection_id": d["cash_collection_id"],
                "cash_collection_no": d["cash_collection_no"],
                "amount_usd": d["amount_usd"],
                "collection_date": _to_date(d["collection_date"]),
                "order_no": d["order_no"],
            })

    will_change, already_ok, unmatched = [], 0, []
    for row in rows:
        status, payload = _match_row(row, name2ids, orders_by_cid, receipts_by_order)
        if status == "will_change":
            will_change.append({
                "excel_row": row["excel_row"],
                "company": row["company"],
                "order_total": row["order_total"],
                "order_no": payload["order_no"],
                "changes": payload["changes"],
            })
        elif status == "already_ok":
            already_ok += 1
        else:
            unmatched.append({
                "excel_row": row["excel_row"],
                "company": row["company"],
                "order_total": row["order_total"],
                "reason_code": status,
                "reason": REASON_LABELS.get(status, status),
                "raw": row["raw"],
            })

    change_count = sum(len(w["changes"]) for w in will_change)
    return {
        "source_file": source_file,
        "summary": {
            "total_rows": len(rows),
            "will_change_orders": len(will_change),
            "will_change_receipts": change_count,
            "already_ok": already_ok,
            "unmatched": len(unmatched),
        },
        "will_change": will_change,
        "unmatched": unmatched,
    }


def apply_changes(db: Session, items: list[dict], operator_id: Optional[int], source_file: str = "") -> dict:
    """写库：只更新前端确认的 (cash_collection_id, new_date)。逐条 savepoint 隔离，
    old_date 以库内实时值为准写审计表。"""
    schema = get_settings().BUSINESS_DB_NAME
    batch_id = uuid.uuid4().hex
    applied, skipped, failed = 0, 0, 0
    errors: list[str] = []

    for it in items:
        ccid = str(it.get("cash_collection_id", "")).strip()
        new_date = _to_date(it.get("new_date"))
        if not ccid or new_date is None:
            skipped += 1
            continue
        try:
            with db.begin_nested():
                cur = db.execute(
                    text(f"SELECT collection_date, order_no, company_name FROM `{schema}`.okki_receipts "
                         f"WHERE cash_collection_id = :id"),
                    {"id": ccid},
                ).mappings().first()
                if cur is None:
                    skipped += 1
                    continue
                old_date = _to_date(cur["collection_date"])
                if old_date == new_date:
                    skipped += 1
                    continue
                db.execute(
                    text(f"UPDATE `{schema}`.okki_receipts SET collection_date = :d "
                         f"WHERE cash_collection_id = :id"),
                    {"d": new_date, "id": ccid},
                )
                db.add(ReceiptRepairLog(
                    batch_id=batch_id,
                    cash_collection_id=ccid,
                    order_no=cur["order_no"],
                    company_name=cur["company_name"],
                    old_date=old_date,
                    new_date=new_date,
                    source_file=source_file[:256] if source_file else None,
                    operator_id=operator_id,
                ))
            applied += 1
        except Exception as e:  # noqa: BLE001 —— 单条失败不拖垮整批
            failed += 1
            msg = f"回款单 {ccid} 更新失败: {e}"
            errors.append(msg)
            logger.warning(msg)
            print(f"[receipt_repair] {msg}", flush=True)

    if applied:
        db.commit()
    else:
        db.rollback()
    logger.info(f"回款日期修复批次 {batch_id}: 应用 {applied}, 跳过 {skipped}, 失败 {failed}")
    return {
        "batch_id": batch_id,
        "applied": applied,
        "skipped": skipped,
        "failed": failed,
        "errors": errors[:20],
    }


MAX_EXPORT_ROWS = 5000


def _safe_cell(v):
    """防 CSV/公式注入：以 = + - @ 开头的字符串加前导单引号，Excel 打开不会当公式执行。"""
    if isinstance(v, str) and v[:1] in ("=", "+", "-", "@"):
        return "'" + v
    return v


def build_unmatched_workbook(unmatched: list[dict]) -> BytesIO:
    """把无法匹配的行导出为新工作表（保留原始列 + 原因列）。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "无法匹配"
    headers = ["订单日期", "客户名", "国家", "总金额USD", "已回款金额", "实际运费", "手续费",
               "回款日期2", "回款金额2", "回款日期3", "回款金额3", "回款日期4", "回款金额4", "无法匹配原因"]
    ws.append(headers)
    for u in unmatched[:MAX_EXPORT_ROWS]:
        raw = u.get("raw", {})
        line = [_safe_cell(raw.get(chr(64 + c))) for c in range(1, 14)]
        line.append(_safe_cell(u.get("reason", "")))
        ws.append(line)
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream
