"""内贸订单 Excel 导出。"""

from datetime import date, datetime
from io import BytesIO
from math import ceil

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins

from app.domestic.constants import PRODUCT_TYPES


_HEADERS = (
    "明细号", "产品类型", "产品名称", "工艺/尺寸", "发长", "网帽颜色",
    "头套尺寸", "发量", "发型系列", "数量", "发型", "颜色", "发型要求", "备注",
)
_WIDTHS = (
    9, 11, 24, 15, 12, 14, 12, 11, 13, 10, 18, 18, 26, 22,
)
_FONT_NAME = "宋体"
_THIN = Side(style="thin", color="000000")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_LONG_TEXT_THRESHOLD = 80
_TEXT_FIELDS = (
    ("hairstyle", "发型"),
    ("color", "颜色"),
    ("style_requirement", "发型要求"),
    ("remark", "备注"),
)


def _safe_text(value) -> str:
    text = "" if value is None else str(value).strip()
    if text.startswith(("=", "+", "-", "@")):
        return f"'{text}"
    return text


def _safe_raw_text(value) -> str:
    text = "" if value is None else str(value)
    if text.startswith(("=", "+", "-", "@")):
        return f"'{text}"
    return text


def _display(value) -> str:
    return _safe_text(value) or "—"


def _date_text(value) -> str:
    if not value:
        return ""
    if isinstance(value, str):
        try:
            value = datetime.fromisoformat(value)
        except ValueError:
            return _safe_text(value)
    if isinstance(value, datetime):
        return value.strftime("%Y/%m/%d")
    if isinstance(value, date):
        return value.strftime("%Y/%m/%d")
    return _safe_text(value)


def _wrapped_lines(value, width: int) -> int:
    text = str(value or "")
    chars_per_line = max(6, int(width * 1.4))
    return sum(max(1, ceil(len(part) / chars_per_line)) for part in text.splitlines() or [""])


def _item_row_height(values: tuple) -> float:
    text_columns = ((1, 11), (2, 24), (10, 18), (11, 18), (12, 26), (13, 22))
    lines = max(_wrapped_lines(values[index], width) for index, width in text_columns)
    return min(300, max(75, lines * 15 + 15))


def _print_chunks(
    text: str, width: int = 100, max_lines: int = 18, max_chars: int = 600,
) -> list[str]:
    chars_per_line = max(6, int(width * 1.4))
    pieces = []
    for line in text.splitlines(keepends=True) or [text]:
        pieces.extend(
            line[index:index + chars_per_line]
            for index in range(0, len(line), chars_per_line)
        )
    chunks = []
    current = ""
    for piece in pieces:
        candidate = current + piece
        if current and (
            len(candidate) > max_chars or _wrapped_lines(candidate, width) > max_lines
        ):
            chunks.append(current)
            current = piece
        else:
            current = candidate
    if current or not chunks:
        chunks.append(current)
    return chunks


def _add_full_requirements_sheet(wb: Workbook, detail: dict) -> None:
    rows = []
    for item in detail.get("items") or []:
        for key, label in _TEXT_FIELDS:
            text = str(item.get(key) or "").strip()
            if len(text) > _LONG_TEXT_THRESHOLD:
                for index, chunk in enumerate(_print_chunks(text)):
                    chunk_label = label if index == 0 else f"{label}（续）"
                    rows.append((_display(item.get("line_code")), chunk_label, _safe_raw_text(chunk)))
    order_remark = str(detail.get("remark") or "").strip()
    if len(order_remark) > _LONG_TEXT_THRESHOLD:
        for index, chunk in enumerate(_print_chunks(order_remark)):
            label = "订单备注" if index == 0 else "订单备注（续）"
            rows.append(("订单", label, _safe_raw_text(chunk)))
    if not rows:
        return

    ws = wb.create_sheet("完整要求")
    for col, value in enumerate(("明细号", "字段", "完整内容"), start=1):
        cell = ws.cell(1, col, value)
        cell.font = Font(name=_FONT_NAME, size=12, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _BORDER
    for row_idx, (line_code, label, content) in enumerate(rows, start=2):
        for col, value in enumerate((line_code, label, content), start=1):
            cell = ws.cell(row_idx, col, value)
            cell.font = Font(name=_FONT_NAME, size=11)
            cell.alignment = Alignment(
                horizontal="left" if col == 3 else "center",
                vertical="top",
                wrap_text=True,
            )
            cell.border = _BORDER
        ws.row_dimensions[row_idx].height = min(
            409, max(45, _wrapped_lines(content, 100) * 15 + 15)
        )
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 100
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"
    ws.print_title_rows = "1:1"
    ws.print_area = f"A1:C{len(rows) + 1}"
    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins = PageMargins(left=0.35, right=0.35, top=0.5, bottom=0.5)


def build_order_workbook(detail: dict, applicant_name: str = "") -> BytesIO:
    """按内贸领货单模板生成单张订单工作簿。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "内贸订单领货单"

    ws.merge_cells("B1:N1")
    ws["B1"] = "内贸订单领货单"
    ws["B1"].font = Font(name=_FONT_NAME, size=18, bold=True)
    ws["B1"].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A2:N2")
    ws["A2"] = (
        f"下单日期：{_date_text(detail.get('order_date'))}     "
        f"要求发货日期：{_date_text(detail.get('required_ship_date'))}     "
        f"客户订单号：{_safe_text(detail.get('order_no'))}     "
        f"系统单号：{_safe_text(detail.get('domestic_no'))}     "
        f"申请人：{_safe_text(applicant_name)}     "
        f"客户：{_safe_text(detail.get('customer_name'))}"
    )
    ws.merge_cells("A3:N3")
    ws["A3"] = (
        "审批人签字：____________________     "
        f"订单类别：{_safe_text(detail.get('order_category_label'))}     "
        f"订单类型：{_safe_text(detail.get('order_type_label'))}     "
        f"订单渠道：{_safe_text(detail.get('order_channel_label'))}"
    )
    for cell in (ws["A2"], ws["A3"]):
        cell.font = Font(name=_FONT_NAME, size=12, bold=True)
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.border = _BORDER

    for col, header in enumerate(_HEADERS, start=1):
        cell = ws.cell(4, col, header)
        cell.font = Font(name=_FONT_NAME, size=12, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _BORDER

    items = detail.get("items") or []
    for row_idx, item in enumerate(items, start=5):
        attrs = item.get("attrs") or {}
        is_piece = attrs.get("product_type") == "piece"
        values = (
            _display(item.get("line_code")),
            PRODUCT_TYPES.get(attrs.get("product_type"), _display(attrs.get("product_type"))),
            _display(item.get("product_name")),
            _display(attrs.get("craft")),
            _display(attrs.get("length")),
            "" if is_piece else _display(attrs.get("net_color")),
            "" if is_piece else _display(attrs.get("size")),
            "" if is_piece else _display(attrs.get("density")),
            "" if is_piece else _display(attrs.get("hair_style_series")),
            item.get("order_qty") or 0,
            _display(item.get("hairstyle")),
            _display(item.get("color")),
            _display(item.get("style_requirement")),
            _display(item.get("remark")),
        )
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row_idx, col, value)
            cell.font = Font(name=_FONT_NAME, size=11)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = _BORDER
        ws.row_dimensions[row_idx].height = _item_row_height(values)

    notes_row = 5 + len(items) + 1
    ws.merge_cells(start_row=notes_row, start_column=1, end_row=notes_row, end_column=14)
    notes = "注意事项：\n！导出内容以方舟内贸订单记录为准。\n！领货与签字流程按内贸部门现行规定执行。"
    if detail.get("remark"):
        notes += f"\n订单备注：{_safe_text(detail['remark'])}"
    ws.cell(notes_row, 1, notes)
    ws.cell(notes_row, 1).font = Font(name=_FONT_NAME, size=11, bold=True)
    ws.cell(notes_row, 1).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws.cell(notes_row, 1).border = _BORDER

    for col, width in enumerate(_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[1].height = 32
    ws.row_dimensions[2].height = 26
    ws.row_dimensions[3].height = 42
    ws.row_dimensions[4].height = 36
    ws.row_dimensions[notes_row].height = min(
        240, max(75, _wrapped_lines(notes, sum(_WIDTHS)) * 15 + 15)
    )

    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins = PageMargins(left=0.24, right=0.24, top=0.35, bottom=0.35)
    ws.print_title_rows = "1:4"
    ws.print_area = f"A1:N{notes_row}"
    _add_full_requirements_sheet(wb, detail)

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream
