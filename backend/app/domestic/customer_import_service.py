"""内贸客户 Excel 导入 —— 《莱莎客户信息录入表》解析与落库。

解析与写库分离：parse_workbook 纯函数可测试；import_customers 负责 upsert。
upsert 口径：先按客户编码(custom_code)命中→整档覆盖；否则按店名(shop_name)命中→
只补空档并记归属冲突（店名全局唯一，Excel 内同人/跨人重复店名都走这里）；都没有→新建。
财务字段（余额/会员等级/启用状态）导入一律不碰。
"""

import io
import logging
import re
from collections import Counter
from datetime import date, datetime
from decimal import Decimal, InvalidOperation

import openpyxl
from sqlalchemy.exc import IntegrityError, SQLAlchemyError
from sqlalchemy.orm import Session

from app.auth.models import ArkUser
from app.domestic.models import DomesticCustomer

logger = logging.getLogger("commission")

CODE_PATTERN = re.compile(r"^ls-[a-z]+-\d+$", re.IGNORECASE)
TEMPLATE_SHEET = "模版"
EXPECTED_HEADERS = (
    "客户编号", "客户正式名称", "联系人", "手机号", "省份", "城市",
    "客户来源", "归属销售", "客户等级", "首次联系日期", "首次下单日期",
    "最近下单日期", "累计订单数", "累计销售额", "客户状态", "门店类型", "备注",
)

PROVINCE_ALIASES = {
    "内蒙古": "内蒙古自治区", "内蒙古省": "内蒙古自治区",
    "广西": "广西壮族自治区", "广西省": "广西壮族自治区",
    "新疆": "新疆维吾尔自治区", "宁夏": "宁夏回族自治区", "西藏": "西藏自治区",
}
MUNICIPALITIES = {"北京市", "天津市", "上海市", "重庆市"}
# 与 ark_domestic_customers 列宽一致；超长按字段级脏数据处理，不拖到 MySQL 严格模式报错
FIELD_LIMITS = {
    "shop_name": 120, "contact": 60, "phone": 40, "province": 64, "city": 64,
    "customer_source": 32, "customer_level": 8, "lifecycle_status": 16,
    "store_type": 32, "remark": 500,
}
MAX_MONEY = Decimal("999999999999.99")
# 只收板上钉钉的笔误/别名；(省份,城市) 键优先于纯城市键
CITY_FIXES = {
    "深棕市": "深圳市", "绍通市": "昭通市",
    "巴彦卓尔市": "巴彦淖尔市", "巴盟": "巴彦淖尔市",
    "攀枝花": "攀枝花市", "朔州": "朔州市", "库尔勒": "库尔勒市",
    "乌兰浩特": "乌兰浩特市", "呼和浩特": "呼和浩特市",
    ("湖南省", "土家族"): "湘西土家族苗族自治州",
    ("四川省", "重庆市"): ("重庆市", "重庆市"),  # 省份也一并纠正
}
# 城市栏明显是脏数据（与省份重复或是省份残片）时丢弃城市
CITY_DISCARD = {"维吾尔自治区", "福建省"}


def _text(value) -> str | None:
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    text = str(value).strip()
    return text or None


def _phone(value) -> str | None:
    return _text(value)


def _date(value) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip().replace("/", "-").replace(".", "-")
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    raise ValueError(f"日期格式无法识别: {value!r}")


def _int(value) -> int | None:
    if value is None or value == "":
        return None
    try:
        return int(float(value))
    except (TypeError, ValueError) as exc:
        raise ValueError(f"数量格式无法识别: {value!r}") from exc


def _money(value) -> Decimal | None:
    if value is None or value == "":
        return None
    try:
        return Decimal(str(value)).quantize(Decimal("0.01"))
    except (InvalidOperation, ValueError) as exc:
        raise ValueError(f"金额格式无法识别: {value!r}") from exc


def _normalize_region(province: str | None, city: str | None) -> tuple[str | None, str | None]:
    if province:
        province = PROVINCE_ALIASES.get(province, province)
    if city and city in CITY_DISCARD:
        city = None
    if city:
        pair_fix = CITY_FIXES.get((province, city))
        if isinstance(pair_fix, tuple):
            return pair_fix
        city = pair_fix or CITY_FIXES.get(city, city)
    if province in MUNICIPALITIES and not city:
        city = province
    return province, city


def parse_workbook(file_bytes: bytes) -> tuple[list[dict], list[dict], list[dict]]:
    """解析《莱莎客户信息录入表》。返回 (有效行, 跳过行及原因, 字段级警告)。

    行级脏数据（编码不合规/缺店名）才进 skipped；单元格级脏数据（坏日期/坏数字）
    只把该字段置空并记 warning，不丢整条客户。
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    rows: list[dict] = []
    skipped: list[dict] = []
    warnings: list[dict] = []

    def _lenient(fn, value, sheet_name: str, row_no: int, code: str, field: str):
        try:
            return fn(value)
        except ValueError as exc:
            warnings.append({
                "sheet": sheet_name, "row_no": row_no, "code": code,
                "reason": f"{field}已置空：{exc}",
            })
            return None

    def _bounded(value: str | None, field: str, sheet_name: str, row_no: int, code: str):
        """超过列宽的字段：备注截断保留信息，其余置空，都记 warning。"""
        if value is None:
            return None
        limit = FIELD_LIMITS[field]
        if len(value) <= limit:
            return value
        if field == "remark":
            warnings.append({
                "sheet": sheet_name, "row_no": row_no, "code": code,
                "reason": f"备注超过 {limit} 字，已截断",
            })
            return value[:limit]
        warnings.append({
            "sheet": sheet_name, "row_no": row_no, "code": code,
            "reason": f"{field}超过 {limit} 字（{value!r}），已置空",
        })
        return None

    for sheet_name in wb.sheetnames:
        if sheet_name.strip() == TEMPLATE_SHEET:
            continue
        sheet = wb[sheet_name]
        header: list[str] | None = None
        for row_no, cells in enumerate(sheet.iter_rows(values_only=True), start=1):
            first = _text(cells[0]) if cells else None
            if header is None:
                if first and first.startswith("客户编号"):
                    # 表头带 * 必填后缀（客户编号*），统一剥掉再作键
                    header = [(_text(c) or "").rstrip("*").strip() for c in cells]
                continue
            if first is None and not any(_text(c) for c in cells):
                continue
            record = dict(zip(header, cells))
            code = _text(record.get("客户编号"))
            shop_name = _text(record.get("客户正式名称"))
            if not code or not CODE_PATTERN.match(code) or not shop_name:
                # 只有编号、其余全空的预编号行不算脏数据，静默跳过
                if code and CODE_PATTERN.match(code) and not any(
                    _text(record.get(h)) for h in header[1:]
                ):
                    continue
                skipped.append({
                    "sheet": sheet_name, "row_no": row_no,
                    "code": code, "shop_name": shop_name,
                    "reason": "客户编号不合规或未填客户正式名称",
                })
                continue
            if len(shop_name) > FIELD_LIMITS["shop_name"]:
                skipped.append({
                    "sheet": sheet_name, "row_no": row_no,
                    "code": code, "shop_name": shop_name,
                    "reason": f"客户店名超过 {FIELD_LIMITS['shop_name']} 字",
                })
                continue
            province, city = _normalize_region(
                _text(record.get("省份")), _text(record.get("城市")),
            )
            sales_amount = _lenient(_money, record.get("累计销售额"), sheet_name, row_no, code, "累计销售额")
            if sales_amount is not None and sales_amount > MAX_MONEY:
                warnings.append({
                    "sheet": sheet_name, "row_no": row_no, "code": code,
                    "reason": "累计销售额超出金额量程，已置空",
                })
                sales_amount = None
            rows.append({
                "sheet": sheet_name,
                "row_no": row_no,
                "custom_code": code.lower(),
                "shop_name": shop_name,
                "contact": _bounded(_text(record.get("联系人")), "contact", sheet_name, row_no, code),
                "phone": _bounded(_phone(record.get("手机号")), "phone", sheet_name, row_no, code),
                "province": _bounded(province, "province", sheet_name, row_no, code),
                "city": _bounded(city, "city", sheet_name, row_no, code),
                "customer_source": _bounded(_text(record.get("客户来源")), "customer_source", sheet_name, row_no, code),
                "owner_name": _text(record.get("归属销售")) or sheet_name.strip(),
                "customer_level": _bounded(_text(record.get("客户等级")), "customer_level", sheet_name, row_no, code),
                "first_contact_date": _lenient(_date, record.get("首次联系日期"), sheet_name, row_no, code, "首次联系日期"),
                "first_order_date": _lenient(_date, record.get("首次下单日期"), sheet_name, row_no, code, "首次下单日期"),
                "last_order_date": _lenient(_date, record.get("最近下单日期"), sheet_name, row_no, code, "最近下单日期"),
                "total_order_count": _lenient(_int, record.get("累计订单数"), sheet_name, row_no, code, "累计订单数"),
                "total_sales_amount": sales_amount,
                "lifecycle_status": _bounded(_text(record.get("客户状态")), "lifecycle_status", sheet_name, row_no, code),
                "store_type": _bounded(_text(record.get("门店类型")), "store_type", sheet_name, row_no, code),
                "remark": _bounded(_text(record.get("备注")), "remark", sheet_name, row_no, code),
            })
    wb.close()
    return rows, skipped, warnings


_PROFILE_FIELDS = (
    "contact", "phone", "province", "city",
    "customer_source", "customer_level", "lifecycle_status", "store_type",
    "first_contact_date", "first_order_date", "last_order_date",
    "total_order_count", "total_sales_amount", "remark",
)


def _apply_row(customer: DomesticCustomer, row: dict, owner_id: int | None, *, fill_only: bool) -> None:
    for field in _PROFILE_FIELDS:
        if fill_only and getattr(customer, field) is not None:
            continue
        setattr(customer, field, row[field])
    if owner_id is not None and not (fill_only and customer.owner_user_id is not None):
        customer.owner_user_id = owner_id


def import_customers(
    db: Session, rows: list[dict], operator_id: int
) -> dict:
    """upsert 客户档案。每行独立 savepoint，单行失败不拖垮整批。"""
    owner_rows = (
        db.query(ArkUser.id, ArkUser.real_name)
        .filter(ArkUser.is_active.is_(True), ArkUser.deleted_at.is_(None))
        .all()
    )
    # real_name 无唯一约束：同名在职用户静默 last-wins 会把客户错挂到人，宁可整行拒绝
    name_counts = Counter(name for _, name in owner_rows)
    ambiguous_names = {name for name, count in name_counts.items() if count > 1}
    owner_ids = {name: uid for uid, name in owner_rows if name not in ambiguous_names}
    owner_names = {uid: name for name, uid in owner_ids.items()}
    result = {"created": 0, "updated": 0, "merged": 0, "errors": [], "collisions": []}
    if ambiguous_names:
        result["errors"].append({
            "row": "-",
            "reason": f"系统存在同名在职用户：{'、'.join(sorted(ambiguous_names))}，相关行已全部拒绝导入",
        })
    for row in rows:
        label = f"{row['sheet']}第{row['row_no']}行[{row['custom_code']}]{row['shop_name']}"
        if row["owner_name"] in ambiguous_names:
            result["errors"].append({"row": label, "reason": f"归属销售「{row['owner_name']}」存在同名在职用户，拒绝自动归属"})
            continue
        owner_id = owner_ids.get(row["owner_name"]) if row["owner_name"] else None
        if row["owner_name"] and owner_id is None:
            result["errors"].append({"row": label, "reason": f"归属销售「{row['owner_name']}」不是系统在职用户"})
            continue
        savepoint = db.begin_nested()
        try:
            existing = db.query(DomesticCustomer).filter(
                DomesticCustomer.custom_code == row["custom_code"]
            ).first()
            if existing is not None and existing.shop_name != row["shop_name"]:
                result["errors"].append({
                    "row": label,
                    "reason": f"客户编码已被「{existing.shop_name}」占用，未导入",
                })
                savepoint.rollback()
                continue
            if existing is None:
                existing = db.query(DomesticCustomer).filter(
                    DomesticCustomer.shop_name == row["shop_name"]
                ).first()
            if existing is None:
                customer = DomesticCustomer(
                    shop_name=row["shop_name"],
                    custom_code=row["custom_code"],
                    balance=0,
                    status=1,
                    created_by=operator_id,
                )
                _apply_row(customer, row, owner_id, fill_only=False)
                db.add(customer)
                db.flush()
                result["created"] += 1
            elif existing.custom_code == row["custom_code"]:
                _apply_row(existing, row, owner_id, fill_only=False)
                result["updated"] += 1
            else:
                # 店名命中但编码不同：老数据(下单就地建档)无编码时收养本次编码，
                # 已有编码则保持先来先得，重跑时按编码命中走覆盖更新，幂等一致
                if existing.custom_code is None:
                    existing.custom_code = row["custom_code"]
                _apply_row(existing, row, owner_id, fill_only=True)
                result["merged"] += 1
                owner_note = ""
                if owner_id is not None and existing.owner_user_id not in (None, owner_id):
                    owner_note = f"；归属保持「{owner_names.get(existing.owner_user_id, existing.owner_user_id)}」"
                result["collisions"].append({
                    "row": label,
                    "reason": f"店名与 [{existing.custom_code}] 重复，已并入并只补空档{owner_note}",
                })
            savepoint.commit()
        except IntegrityError as exc:
            savepoint.rollback()
            logger.warning("domestic customer import integrity error on %s: %s", label, exc)
            print(f"[domestic] customer import integrity error {label}: {exc}", flush=True)
            result["errors"].append({"row": label, "reason": "店名或编码与已有客户冲突"})
        except SQLAlchemyError as exc:
            # DataError(超长/超量程)等也按行隔离，不拖垮整批
            savepoint.rollback()
            logger.warning("domestic customer import db error on %s: %s", label, exc)
            print(f"[domestic] customer import db error {label}: {exc}", flush=True)
            result["errors"].append({"row": label, "reason": f"数据库写入失败：{type(exc).__name__}"})
    db.commit()
    return result
