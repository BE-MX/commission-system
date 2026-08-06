"""薪资模块 — 社保 / 公积金 Excel 解析（M2-b，纯函数，不碰 session）。

三条设计红线，改这个文件前先读完：

1. **不信表头文案，只信数据**。丽丝发 sheet 的表头写「工伤上升至 0.2%」，
   两个主体的数据实际都是 0.32%（设计文档 §2.2）。所以费率一律从数据列反推，
   解析结果里带 `rates_observed` 供异常面板展示，代码不拿表头里的百分比做任何计算。

2. **列按表头文字定位，且要求唯一命中**。源表列序换过一次就会让「单位缴纳」
   被读成「个人费用」——那是直接把公司缴的钱扣到员工头上。命中 0 列或 >1 列
   一律抛 `SalaryImportError`，不做「猜一个」。

3. **两道互相独立的结构性守卫**，任一不过就是列映射错了，不是数据脏：
   - 分项恒等式：养老个人 + 失业个人 + 医保个人 == 个人总计（实测 5345 基数 →
     427.6 + 16.04 + 106.9 = 550.54，分位吻合）；公积金：个人 + 单位 == 总计。
     少数行不符 = 补缴/滞纳金，记行级 warning；过半不符 = 列映射错，直接抛错。
   - 合计行对账：解析出的个人合计 vs 源表最后一行的合计值。差额不为 0 说明漏读了行。

PII：身份证在解析当场就 hash + 加密，**返回值里永不出现明文**（只出 masked）。
下游落库、匹配档案都用 `id_card_hash`（档案侧同一把 HMAC 钥匙，见 pii.py）。
"""

from __future__ import annotations

import logging
from dataclasses import dataclass, field
from decimal import ROUND_HALF_UP, Decimal, InvalidOperation
from typing import Any, Optional, Sequence

from app.salary import pii

logger = logging.getLogger("commission")

# 参保主体规范名：与 salary_employee_profile.insurance_entity 同一套字面量。
# 源表「交费单位」列写的是简称（"青岛丽丝发贸易"），必须归一才能跟档案对上。
ENTITY_LISI = "青岛丽丝发贸易有限公司"
ENTITY_JUANCHENG = "鄄城莱莎发制品有限公司青岛分公司"

# 判定顺序有意义：「鄄城…青岛分公司」同时含「青岛」，先判分公司才不会被丽丝发规则截胡。
_ENTITY_RULES = (
    (("鄄城", "分公司"), ENTITY_JUANCHENG),
    (("丽丝",), ENTITY_LISI),
)

# 合计/签字行的姓名特征。这些行混进人员清单会污染人数与个人合计。
NON_PERSON_TOKENS = ("合计", "总计", "小计", "制表", "审核", "批准", "复核", "签字", "备注")

MAX_SHEET_ROWS = 5000

# 恒等式容差：源表是 Excel 浮点，分位允许 1 分钱的进位误差。
_TOLERANCE = Decimal("0.02")

# 个人扣款项的费率合理区间（反推用，超出只告警不拦截——费率本来就会年度调整）。
_PERSONAL_RATE_BANDS = {
    "pension_personal": (Decimal("0.06"), Decimal("0.10")),
    "unemployment_personal": (Decimal("0.001"), Decimal("0.01")),
    "medical_personal": (Decimal("0.010"), Decimal("0.035")),
}
_FUND_RATE_BAND = (Decimal("0.04"), Decimal("0.13"))


class SalaryImportError(ValueError):
    """结构性解析失败：表头找不到、列命中不唯一、恒等式过半不成立。

    这类错误必须让用户看见并换文件重传，绝不能降级成「先导进去再说」——
    列映射错一次，66 个人的社保扣款就全错了。
    """


# ---------------------------------------------------------------------------
# 数据结构
# ---------------------------------------------------------------------------

@dataclass
class ParsedPersonRow:
    """一个真人的一行。**没有身份证明文字段，这是有意的。**"""

    sheet: str
    row_no: int  # 源表 1-based 行号，异常面板要能告诉 HR「第几行」
    entity: Optional[str]
    name: str
    dept_text: Optional[str]
    id_card_hash: Optional[str]
    id_card_cipher: Optional[str]
    id_card_masked: str
    base_amount: Optional[Decimal]
    personal_total: Optional[Decimal]
    company_total: Optional[Decimal]
    detail: dict[str, str] = field(default_factory=dict)
    warnings: list[str] = field(default_factory=list)


@dataclass
class ParsedSheet:
    sheet: str
    entity: Optional[str]
    person_count: int
    personal_sum: Decimal  # 解析出的个人合计
    personal_total_in_file: Optional[Decimal]  # 源表合计行的值
    rates_observed: dict[str, str]  # 反推费率，展示用
    warnings: list[str] = field(default_factory=list)


@dataclass
class ParseResult:
    rows: list[ParsedPersonRow]
    sheets: list[ParsedSheet]
    warnings: list[str] = field(default_factory=list)


# ---------------------------------------------------------------------------
# 通用工具
# ---------------------------------------------------------------------------

def _norm_header(value: Any) -> str:
    """表头归一化：去掉所有空白（源表「社保缴费\\n基数」是换行分两行写的）。"""
    if value is None:
        return ""
    return "".join(str(value).split())


def _dec(value: Any) -> Optional[Decimal]:
    """单元格转 Decimal，统一量化到分。空/非数字返回 None（由调用方决定是否算异常）。

    量化用 ROUND_HALF_UP 而不是 Python 内建 round()：后者是银行家舍入
    （0.005 → 0.00），跟工资表的四舍五入口径不一致，会在半分位上稳定差 1 分。
    源表的 Excel 浮点（550.5400000000001）也在这里被抹平。
    """
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, Decimal):
        return value.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    if isinstance(value, (int, float)):
        return Decimal(str(value)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    text = str(value).strip().replace(",", "").replace("¥", "").replace("￥", "")
    if not text:
        return None
    try:
        return Decimal(text).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    except (InvalidOperation, ArithmeticError):
        return None


def _text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def is_person_row(name: str) -> bool:
    """判定是否真人行。空姓名与合计/签字行一律排除。"""
    if not name:
        return False
    return not any(tok in name for tok in NON_PERSON_TOKENS)


def normalize_entity(raw: str, sheet_name: str = "") -> Optional[str]:
    """把「交费单位」文本（或 sheet 名兜底）归一成档案里的参保主体规范名。"""
    haystack = f"{raw}{sheet_name}"
    for tokens, canonical in _ENTITY_RULES:
        if any(tok in haystack for tok in tokens):
            return canonical
    return None


def _load_sheets(file_bytes: bytes) -> list[tuple[str, list[list[Any]]]]:
    """读工作簿，返回 [(sheet 名, 二维单元格)]。按魔数分派 .xls / .xlsx。

    HR 给的是 .xls（BIFF，openpyxl 读不了），但不排除以后另存成 .xlsx，
    所以两条路都留着——靠魔数判断而不是扩展名，改个后缀不该让导入炸掉。
    """
    if not file_bytes:
        raise SalaryImportError("文件为空，请重新上传")
    if file_bytes[:4] == b"\xd0\xcf\x11\xe0":  # OLE2 复合文档 = 老版 .xls
        import xlrd

        try:
            book = xlrd.open_workbook(file_contents=file_bytes)
        except Exception as exc:  # noqa: BLE001
            raise SalaryImportError(f"读取 .xls 失败：{exc}") from exc
        out = []
        for sheet in book.sheets():
            nrows = min(sheet.nrows, MAX_SHEET_ROWS)
            out.append((sheet.name, [list(sheet.row_values(r)) for r in range(nrows)]))
        return out
    if file_bytes[:2] == b"PK":  # zip 容器 = .xlsx
        from io import BytesIO

        from openpyxl import load_workbook

        try:
            book = load_workbook(filename=BytesIO(file_bytes), data_only=True, read_only=True)
        except Exception as exc:  # noqa: BLE001
            raise SalaryImportError(f"读取 .xlsx 失败：{exc}") from exc
        out = []
        for sheet in book.worksheets:
            rows = []
            for idx, row in enumerate(sheet.iter_rows(values_only=True)):
                if idx >= MAX_SHEET_ROWS:
                    break
                rows.append(list(row))
            out.append((sheet.title, rows))
        return out
    raise SalaryImportError("无法识别的文件格式，请上传 .xls 或 .xlsx")


def _find_header(rows: Sequence[Sequence[Any]]) -> tuple[int, list[str]]:
    """定位表头行：第一行同时含「姓名」和「身份证」的行。

    源表第 0 行是标题（跨列合并），第 1 行才是表头。写死行号会在 HR 加一行说明时
    静默错位，所以按内容找。
    """
    for idx, row in enumerate(rows[:10]):
        header = [_norm_header(c) for c in row]
        joined = "".join(header)
        if "姓名" in joined and "身份证" in joined:
            return idx, header
    raise SalaryImportError("找不到表头行（需同时包含「姓名」与「身份证号」两列）")


def _resolve(
    header: Sequence[str],
    candidates: Sequence[tuple[str, ...]],
    *,
    label: str,
    taken: set[int],
    required: bool = True,
) -> Optional[int]:
    """按候选 token 组定位列，**要求唯一命中**。

    candidates 按优先级排序：先试精确的（"单位缴纳"），命中不唯一或为空再试宽松的
    （"单位"）。宽松规则会撞上「交费单位」这类文本列，所以已被认领的列（taken）
    先排除掉——文本列在金额列之前解析就是为了这个。
    """
    for tokens in candidates:
        hits = [
            i
            for i, h in enumerate(header)
            if i not in taken and h and all(tok in h for tok in tokens)
        ]
        if len(hits) == 1:
            taken.add(hits[0])
            return hits[0]
        if len(hits) > 1:
            names = "、".join(header[i] for i in hits)
            raise SalaryImportError(
                f"表头「{label}」命中多列（{names}），无法确定取哪一列，请检查源表列名"
            )
    if required:
        raise SalaryImportError(f"表头缺少「{label}」列")
    return None


def _cell(row: Sequence[Any], idx: Optional[int]) -> Any:
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def _rate(part: Optional[Decimal], base: Optional[Decimal]) -> Optional[Decimal]:
    if part is None or not base:
        return None
    try:
        return (part / base).quantize(Decimal("0.0001"))
    except (InvalidOperation, ArithmeticError, ZeroDivisionError):
        return None


def _median(values: list[Decimal]) -> Optional[Decimal]:
    if not values:
        return None
    ordered = sorted(values)
    return ordered[len(ordered) // 2]


def _apply_pii(row: ParsedPersonRow, id_card_plain: str) -> None:
    """身份证落到行上：只留 hash / cipher / masked，明文就地丢弃。"""
    norm = pii.normalize_id_card(id_card_plain)
    if not norm:
        row.warnings.append("身份证为空，无法匹配档案")
        return
    if len(norm) not in (15, 18):
        row.warnings.append(f"身份证位数异常（{len(norm)} 位）")
    row.id_card_hash = pii.hash_pii(norm)
    row.id_card_cipher = pii.encrypt_pii(norm)
    row.id_card_masked = pii.mask_pii(norm, 3, 4)


def _check_duplicates(rows: list[ParsedPersonRow]) -> list[str]:
    """同一份文件里身份证撞号 = 誊抄错误（3 月公积金表刘也/姜婷即此，§2.5 错误 3）。"""
    seen: dict[str, list[str]] = {}
    for r in rows:
        if r.id_card_hash:
            seen.setdefault(r.id_card_hash, []).append(f"{r.sheet} 第{r.row_no}行 {r.name}")
    return [
        f"身份证重复：{' / '.join(who)}——源表誊抄错误，需人工核对"
        for who in seen.values()
        if len(who) > 1
    ]


def _reconcile(
    sheet_name: str, personal_sum: Decimal, in_file: Optional[Decimal]
) -> list[str]:
    """解析合计 vs 源表合计行。差额说明漏读或多读了行。"""
    if in_file is None:
        return [f"{sheet_name}：源表没有合计行，无法对账（已按解析结果 {personal_sum} 计）"]
    diff = (personal_sum - in_file).copy_abs()
    if diff > _TOLERANCE:
        return [
            f"{sheet_name}：个人合计对不上——解析 {personal_sum}，源表合计行 {in_file}，"
            f"差额 {personal_sum - in_file}。可能有行未被识别为人员"
        ]
    return []


def _guard_identity(mismatch: int, total: int, label: str) -> None:
    """恒等式过半不成立 = 列映射错了，不是数据脏。这时导进去等于把钱算错。"""
    if total >= 3 and mismatch * 2 > total:
        raise SalaryImportError(
            f"{label}：{total} 行中有 {mismatch} 行分项之和与合计列对不上，"
            "判定为列映射错误（很可能源表列序或列名变了），已终止导入"
        )


# ---------------------------------------------------------------------------
# 社保
# ---------------------------------------------------------------------------

# 各险种列的候选 token。顺序 = 优先级，前面的更精确。
_INS_COLUMNS: tuple[tuple[str, str, tuple[tuple[str, ...], ...], bool], ...] = (
    # (键, 中文标签, 候选 tokens, 是否必需)
    ("social_base", "社保缴费基数", (("社保缴费基数",), ("社保", "基数")), True),
    ("pension_personal", "养老个人", (("养老", "个人"),), True),
    ("pension_company", "养老单位", (("养老", "单位"),), False),
    ("unemployment_personal", "失业个人", (("失业", "个人"),), True),
    ("unemployment_company", "失业单位", (("失业", "单位"),), False),
    ("injury_company", "工伤单位", (("工伤",),), False),
    ("social_late_fee", "社保滞纳金", (("社保滞纳金",), ("社保", "滞纳")), False),
    ("medical_base", "医保缴费基数", (("医保缴费基数",), ("医保", "基数")), False),
    ("medical_company", "医保单位", (("医保", "单位"),), False),
    ("medical_personal", "医保个人", (("医保", "个人"),), True),
    ("medical_late_fee", "医保滞纳金", (("医保滞纳金",), ("医保", "滞纳")), False),
    ("personal_total", "个人总计", (("个人总计",), ("个人", "合计")), True),
    ("company_total", "单位总计", (("单位总计",), ("单位", "合计")), False),
)

# 构成「个人总计」的三项。恒等式与费率反推都只认这三项。
_INS_PERSONAL_PARTS = ("pension_personal", "unemployment_personal", "medical_personal")


def parse_insurance(file_bytes: bytes) -> ParseResult:
    """解析社保明细（两个 sheet = 两个参保主体）。"""
    return _parse_workbook(file_bytes, kind="insurance")


def parse_fund(file_bytes: bytes) -> ParseResult:
    """解析公积金明细（结构同社保，列更少）。"""
    return _parse_workbook(file_bytes, kind="fund")


_FUND_COLUMNS: tuple[tuple[str, str, tuple[tuple[str, ...], ...], bool], ...] = (
    ("fund_base", "缴存基数", (("缴费基数",), ("缴存基数",), ("基数",)), True),
    # 「总计缴纳」必须比「个人费用/单位缴纳」先认领：宽松 token 会互相撞。
    ("fund_total", "总计缴纳", (("总计缴纳",), ("合计金额",), ("总计",), ("合计",)), False),
    ("personal_amount", "个人缴存", (("个人费用",), ("个人份额",), ("个人",)), True),
    ("company_amount", "单位缴存", (("单位缴纳",), ("单位份额",), ("单位",)), True),
)


def _parse_workbook(file_bytes: bytes, *, kind: str) -> ParseResult:
    sheets_raw = _load_sheets(file_bytes)
    all_rows: list[ParsedPersonRow] = []
    sheet_summaries: list[ParsedSheet] = []
    warnings: list[str] = []

    for sheet_name, rows in sheets_raw:
        if not rows:
            continue
        try:
            header_idx, header = _find_header(rows)
        except SalaryImportError as exc:
            # 单个 sheet 没表头（说明页/空页）只告警，别让整份文件导不进来。
            logger.warning("薪资导入：sheet %s 跳过（%s）", sheet_name, exc)
            print(f"[salary.import] skip sheet {sheet_name}: {exc}", flush=True)
            warnings.append(f"{sheet_name}：{exc}，该 sheet 已跳过")
            continue
        parsed, summary = _parse_sheet(
            sheet_name, header, rows[header_idx + 1 :], kind=kind, first_row_no=header_idx + 2
        )
        all_rows.extend(parsed)
        sheet_summaries.append(summary)

    if not sheet_summaries:
        raise SalaryImportError("文件里没有可解析的 sheet，请确认上传的是社保/公积金明细原件")
    warnings.extend(_check_duplicates(all_rows))
    return ParseResult(rows=all_rows, sheets=sheet_summaries, warnings=warnings)


def _parse_sheet(
    sheet_name: str,
    header: list[str],
    body: Sequence[Sequence[Any]],
    *,
    kind: str,
    first_row_no: int = 1,
) -> tuple[list[ParsedPersonRow], ParsedSheet]:
    taken: set[int] = set()
    # 文本列先认领——否则「交费单位」会被金额列的宽松 token「单位」抢走。
    i_name = _resolve(header, (("姓名",),), label="姓名", taken=taken)
    i_id = _resolve(header, (("身份证",),), label="身份证号", taken=taken)
    i_dept = _resolve(header, (("部门",),), label="部门", taken=taken, required=False)
    i_entity = _resolve(
        header, (("交费单位",), ("缴费单位",)), label="交费单位", taken=taken, required=False
    )
    for tokens, label in ((("序号",), "序号"), (("缴费类型",), "缴费类型")):
        _resolve(header, (tokens,), label=label, taken=taken, required=False)
    for tokens, label in ((("业务年",), "业务年月"), (("缴费年",), "缴费年月")):
        _resolve(header, (tokens,), label=label, taken=taken, required=False)

    specs = _INS_COLUMNS if kind == "insurance" else _FUND_COLUMNS
    cols: dict[str, Optional[int]] = {}
    for key, label, candidates, required in specs:
        cols[key] = _resolve(header, candidates, label=label, taken=taken, required=required)

    base_key = "social_base" if kind == "insurance" else "fund_base"
    total_key = "personal_total" if kind == "insurance" else "fund_total"

    rows: list[ParsedPersonRow] = []
    personal_sum = Decimal("0.00")
    total_in_file: Optional[Decimal] = None
    mismatch = 0
    rate_samples: dict[str, list[Decimal]] = {}
    sheet_entity: Optional[str] = None

    for offset, raw in enumerate(body):
        # 源表真实行号（1-based，与用户在 Excel 里看到的一致）。异常面板要靠它定位。
        row_no = first_row_no + offset
        name = _text(_cell(raw, i_name))
        if not is_person_row(name):
            # 合计行：姓名为空但金额列有值，取来做对账。
            candidate = _dec(_cell(raw, cols.get("personal_total") or cols.get("personal_amount")))
            if candidate is not None:
                total_in_file = candidate
            continue

        base = _dec(_cell(raw, cols[base_key]))
        row = ParsedPersonRow(
            sheet=sheet_name,
            row_no=row_no,
            entity=None,
            name=name,
            dept_text=_text(_cell(raw, i_dept)) or None,
            id_card_hash=None,
            id_card_cipher=None,
            id_card_masked="",
            base_amount=base,
            personal_total=None,
            company_total=None,
        )
        _apply_pii(row, _text(_cell(raw, i_id)))
        row.entity = normalize_entity(_text(_cell(raw, i_entity)), sheet_name)
        if row.entity and sheet_entity is None:
            sheet_entity = row.entity

        if kind == "insurance":
            _fill_insurance(row, raw, cols, rate_samples)
        else:
            _fill_fund(row, raw, cols, rate_samples)

        if row.personal_total is None:
            row.warnings.append("个人扣款金额为空")
        else:
            personal_sum += row.personal_total
        if any("分项之和" in w for w in row.warnings):
            mismatch += 1
        rows.append(row)

    _guard_identity(mismatch, len(rows), f"{sheet_name}（{'社保' if kind == 'insurance' else '公积金'}）")

    rates_observed, rate_warnings = _summarize_rates(sheet_name, rate_samples, kind=kind)
    summary = ParsedSheet(
        sheet=sheet_name,
        entity=sheet_entity or normalize_entity("", sheet_name),
        person_count=len(rows),
        personal_sum=personal_sum,
        personal_total_in_file=total_in_file,
        rates_observed=rates_observed,
        warnings=rate_warnings + _reconcile(sheet_name, personal_sum, total_in_file),
    )
    if mismatch:
        summary.warnings.append(f"{mismatch} 行分项之和与合计列不符（可能是补缴或滞纳金），已逐行标注")
    return rows, summary


def _fill_insurance(
    row: ParsedPersonRow,
    raw: Sequence[Any],
    cols: dict[str, Optional[int]],
    rate_samples: dict[str, list[Decimal]],
) -> None:
    detail: dict[str, Decimal] = {}
    for key, _label, _cands, _req in _INS_COLUMNS:
        if key in ("social_base", "personal_total", "company_total"):
            continue
        value = _dec(_cell(raw, cols.get(key)))
        if value is not None:
            detail[key] = value
    row.detail = {k: str(v) for k, v in detail.items()}
    row.personal_total = _dec(_cell(raw, cols.get("personal_total")))
    row.company_total = _dec(_cell(raw, cols.get("company_total")))

    parts = [detail.get(k) for k in _INS_PERSONAL_PARTS]
    if all(p is not None for p in parts) and row.personal_total is not None:
        derived = sum(parts, Decimal("0.00"))
        if (derived - row.personal_total).copy_abs() > _TOLERANCE:
            row.warnings.append(
                f"分项之和 {derived} 与个人总计 {row.personal_total} 不符（差 {derived - row.personal_total}）"
            )
    for key in _INS_PERSONAL_PARTS:
        rate = _rate(detail.get(key), row.base_amount)
        if rate is not None:
            rate_samples.setdefault(key, []).append(rate)


def _fill_fund(
    row: ParsedPersonRow,
    raw: Sequence[Any],
    cols: dict[str, Optional[int]],
    rate_samples: dict[str, list[Decimal]],
) -> None:
    personal = _dec(_cell(raw, cols.get("personal_amount")))
    company = _dec(_cell(raw, cols.get("company_amount")))
    total = _dec(_cell(raw, cols.get("fund_total")))
    row.personal_total = personal
    row.company_total = company
    row.detail = {
        k: str(v)
        for k, v in (("personal_amount", personal), ("company_amount", company), ("total_amount", total))
        if v is not None
    }
    if personal is not None and company is not None and total is not None:
        derived = personal + company
        if (derived - total).copy_abs() > _TOLERANCE:
            row.warnings.append(f"分项之和 {derived} 与总计缴纳 {total} 不符（差 {derived - total}）")
    rate = _rate(personal, row.base_amount)
    if rate is not None:
        rate_samples.setdefault("fund_personal", []).append(rate)


def _summarize_rates(
    sheet_name: str, rate_samples: dict[str, list[Decimal]], *, kind: str
) -> tuple[dict[str, str], list[str]]:
    """反推费率并做区间体检。**只告警不拦截**——费率年度调整是正常的。"""
    observed: dict[str, str] = {}
    warnings: list[str] = []
    bands = _PERSONAL_RATE_BANDS if kind == "insurance" else {"fund_personal": _FUND_RATE_BAND}
    for key, samples in rate_samples.items():
        med = _median(samples)
        if med is None:
            continue
        observed[key] = f"{med * 100:.2f}%"
        band = bands.get(key)
        if band and not (band[0] <= med <= band[1]):
            warnings.append(
                f"{sheet_name}：{key} 反推费率 {med * 100:.2f}% 超出常规区间"
                f"（{band[0] * 100:.1f}%~{band[1] * 100:.1f}%），请确认列是否取对"
            )
    return observed, warnings
