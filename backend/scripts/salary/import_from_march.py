"""薪资档案反导入脚本（M1 首任务，A3 决策落地）。

从 3 份人工 Excel + 平台 users 表反推 66 人薪资档案，产出：
  1. backend/tmp/salary_profile_draft.xlsx  — 档案初稿（交 HR 逐条复核）
  2. backend/tmp/salary_anomalies.xlsx      — 异常清单（交 HR 批注修正）

**只读不写库**：本脚本不落 DB。HR 在初稿上批注确认后，由 M1 的 seed 流程导入。
这样设计是因为 §2.5 已证明人工表存在 9 处以上错误，直接落库会把错误固化进主数据。

异常检测覆盖 §2.5 的 6 类错误 + 反导入过程中新发现的：
  A 工号重复（含归一化撞号，如 "3" vs "003"）
  B 银行卡号重复（资金风险级）
  C 身份证号跨表不一致（社保表 vs 公积金表）
  D 入职日期两表冲突（明细 vs 汇总）
  E 职务两表不一致（B4 决策后以明细为准，仍需 HR 确认）
  F 社保/公积金参保但不发薪（白名单候选）
  G 发薪但未参保（新入职正常，需 HR 确认）
  H users 表匹配失败（平台账号未关联，影响钉钉考勤绑定）

用法：
    cd backend && .venv\\Scripts\\python.exe scripts/salary/import_from_march.py
    # 可选：--materials <dir> 覆盖材料目录
"""
from __future__ import annotations

import argparse
import datetime as dt
import pathlib
import re
import sys
from collections import defaultdict
from dataclasses import dataclass, field
from typing import Any

import xlrd  # type: ignore
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

DEFAULT_MATERIALS = pathlib.Path(
    r"C:\Users\windb\Downloads\项目\_Review\2026-07-29_薪资系统所需材料"
)
OUT_DIR = pathlib.Path(__file__).resolve().parents[2] / "tmp"

F_PAYROLL = "2026年3月工资表.xls"
F_INSURANCE = "2026年3月社保&医保明细.xls"
F_FUND = "26年3月公积金.xls"

# 参保主体标识（sheet 名关键词 → 主体简称），用于 profile.insurance_entity
ENTITY_LISI = "青岛丽丝发贸易有限公司"
ENTITY_JUANCHENG = "鄄城莱莎发制品有限公司青岛分公司"

# 非人员行：合计行 + 表尾签名行（"制表：" 那行曾被误判成员工，2026-08-06）
NON_PERSON_TOKENS = ("合计", "总计", "小计", "制表", "审核", "批准", "复核", "签字", "备注")

# 信息提示类：系统性口径差异而非逐人数据错误，单独成页，不占异常清单版面
INFO_CODES = frozenset({"E"})


def is_person_row(name: str) -> bool:
    """判定该行是否真人。签名/合计行一律排除，否则会污染人数与异常清单。"""
    if not name:
        return False
    return not any(tok in name for tok in NON_PERSON_TOKENS)

# 明细表 23 列的列序（r00 为表头，已由 _explore 确认）
D_SEQ, D_HIRE, D_EMPNO, D_NAME, D_DEPT, D_POSITION = 0, 1, 2, 3, 4, 5
D_DUE_DAYS, D_ACTUAL_DAYS, D_BASE = 6, 7, 8
D_BONUS, D_PERF, D_SENIORITY, D_ATTEND, D_ADD_SUB = 9, 10, 11, 12, 13
D_SOCIAL, D_FUND, D_ABSENCE, D_OTHER, D_DEDUCT_SUB = 14, 15, 16, 17, 18
D_SUBSIDY, D_NET, D_TAX, D_NET_AFTER_TAX = 19, 20, 21, 22

# 汇总表 9 列（r01 为表头）
S_SEQ, S_DEPT, S_NAME, S_POSITION, S_HIRE, S_BANK, S_LEAVE, S_NET, S_REMARK = range(9)

# 社保表列序（r01 为表头）
I_NAME, I_ID, I_BASE, I_PERSONAL_TOTAL = 6, 7, 8, 19
I_DEPT = 5
# 公积金表列序
G_NAME, G_ID, G_BASE, G_PERSONAL = 6, 7, 8, 10
G_DEPT = 5


def excel_serial_to_date(v: Any) -> dt.date | None:
    """Excel 日期序列号 → date。人工表里入职日期混用序列号和文本，两种都要吃。"""
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)) and v > 1000:
        # xlrd 的 1900 日期系统：序列号 1 = 1900-01-01，但 Excel 有 1900 闰年 bug
        try:
            y, m, d, *_ = xlrd.xldate_as_tuple(v, 0)
            return dt.date(y, m, d)
        except Exception:
            return None
    s = str(v).strip()
    # 2025.2.11 / 2025/02/21 / 2025-02-11 三种人工写法
    m = re.match(r"^(\d{4})[.\-/](\d{1,2})[.\-/](\d{1,2})", s)
    if m:
        try:
            return dt.date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        except ValueError:
            return None
    return None


def norm_emp_no(v: Any) -> str:
    """工号归一化：'3' / '003' / 3.0 → '3'，用于撞号检测（§2.5 错误 4）。"""
    if v is None or v == "":
        return ""
    s = str(v).strip()
    if s.endswith(".0"):
        s = s[:-2]
    return s.lstrip("0") or "0"


def norm_id_card(v: Any) -> str:
    """身份证归一化：去空格、末位 x → X。"""
    if v is None or v == "":
        return ""
    return str(v).strip().replace(" ", "").upper()


def norm_bank(v: Any) -> str:
    if v is None or v == "":
        return ""
    s = str(v).strip().replace(" ", "")
    if s.endswith(".0"):
        s = s[:-2]
    return s


def mask_secret(s: str) -> str:
    """脱敏：只保留前 3 末 4。输出文件给 HR 看，不落明文。"""
    if not s or len(s) < 8:
        return s
    return f"{s[:3]}{'*' * (len(s) - 7)}{s[-4:]}"


def num(v: Any, default: float = 0.0) -> float:
    if v is None or v == "":
        return default
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).strip().replace(",", ""))
    except ValueError:
        return default


def text(v: Any) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    if s.endswith(".0") and s[:-2].isdigit():
        s = s[:-2]
    return s
@dataclass
class DetailRow:
    """明细表一行（工资表真相源）。"""

    seq: int
    emp_no_raw: str
    emp_no: str
    name: str
    dept_detail: str
    position: str
    hire_date: dt.date | None
    due_days: float
    actual_days: float
    base_salary: float
    bonus: float
    performance: float
    seniority: float
    attendance_bonus: float
    add_subtotal: float
    social: float
    fund: float
    absence: float
    other: float
    deduct_subtotal: float
    subsidy: float
    net: float
    tax: float


@dataclass
class SummaryRow:
    """汇总表一行（发放/公示口径）。"""

    seq: int
    dept_group: str
    name: str
    position: str
    hire_date: dt.date | None
    bank_card: str
    leave_text: str
    net: float
    remark: str


@dataclass
class InsuranceRow:
    entity: str
    dept: str
    name: str
    id_card: str
    base: float
    personal_total: float


@dataclass
class FundRow:
    entity: str
    dept: str
    name: str
    id_card: str
    base: float
    personal: float


@dataclass
class Anomaly:
    code: str
    category: str
    subject: str
    detail: str
    suggestion: str


@dataclass
class Profile:
    """反推出的档案初稿（一人一行）。"""

    emp_no: str
    name: str
    dept_detail: str = ""
    dept_group: str = ""
    position: str = ""
    hire_date: dt.date | None = None
    id_card: str = ""
    bank_card: str = ""
    insurance_entity: str = ""
    payroll_included: bool = True
    fund_included: bool = False
    base_salary_march: float = 0.0
    social_march: float = 0.0
    fund_march: float = 0.0
    user_id: int | None = None
    user_login: str = ""
    notes: list[str] = field(default_factory=list)


def load_payroll(path: pathlib.Path) -> tuple[list[DetailRow], list[SummaryRow]]:
    wb = xlrd.open_workbook(str(path))
    summary_sh = wb.sheet_by_index(0)
    detail_sh = wb.sheet_by_index(1)

    details: list[DetailRow] = []
    for r in range(1, detail_sh.nrows):
        name = text(detail_sh.cell_value(r, D_NAME))
        if not is_person_row(name):
            continue
        emp_raw = text(detail_sh.cell_value(r, D_EMPNO))
        details.append(
            DetailRow(
                seq=int(num(detail_sh.cell_value(r, D_SEQ))),
                emp_no_raw=emp_raw,
                emp_no=norm_emp_no(emp_raw),
                name=name,
                dept_detail=text(detail_sh.cell_value(r, D_DEPT)),
                position=text(detail_sh.cell_value(r, D_POSITION)),
                hire_date=excel_serial_to_date(detail_sh.cell_value(r, D_HIRE)),
                due_days=num(detail_sh.cell_value(r, D_DUE_DAYS)),
                actual_days=num(detail_sh.cell_value(r, D_ACTUAL_DAYS)),
                base_salary=num(detail_sh.cell_value(r, D_BASE)),
                bonus=num(detail_sh.cell_value(r, D_BONUS)),
                performance=num(detail_sh.cell_value(r, D_PERF)),
                seniority=num(detail_sh.cell_value(r, D_SENIORITY)),
                attendance_bonus=num(detail_sh.cell_value(r, D_ATTEND)),
                add_subtotal=num(detail_sh.cell_value(r, D_ADD_SUB)),
                social=num(detail_sh.cell_value(r, D_SOCIAL)),
                fund=num(detail_sh.cell_value(r, D_FUND)),
                absence=num(detail_sh.cell_value(r, D_ABSENCE)),
                other=num(detail_sh.cell_value(r, D_OTHER)),
                deduct_subtotal=num(detail_sh.cell_value(r, D_DEDUCT_SUB)),
                subsidy=num(detail_sh.cell_value(r, D_SUBSIDY)),
                net=num(detail_sh.cell_value(r, D_NET)),
                tax=num(detail_sh.cell_value(r, D_TAX)),
            )
        )

    summaries: list[SummaryRow] = []
    current_dept = ""
    for r in range(2, summary_sh.nrows):
        name = text(summary_sh.cell_value(r, S_NAME))
        dept = text(summary_sh.cell_value(r, S_DEPT))
        if dept:
            # 合并单元格：部门只在分组首行出现，向下填充
            current_dept = dept
        if not is_person_row(name):
            continue
        summaries.append(
            SummaryRow(
                seq=int(num(summary_sh.cell_value(r, S_SEQ))),
                dept_group=current_dept,
                name=name,
                position=text(summary_sh.cell_value(r, S_POSITION)),
                hire_date=excel_serial_to_date(summary_sh.cell_value(r, S_HIRE)),
                bank_card=norm_bank(summary_sh.cell_value(r, S_BANK)),
                leave_text=text(summary_sh.cell_value(r, S_LEAVE)),
                net=num(summary_sh.cell_value(r, S_NET)),
                remark=text(summary_sh.cell_value(r, S_REMARK)),
            )
        )
    return details, summaries


def load_insurance(path: pathlib.Path) -> list[InsuranceRow]:
    wb = xlrd.open_workbook(str(path))
    rows: list[InsuranceRow] = []
    for sh in wb.sheets():
        entity = ENTITY_JUANCHENG if "鄄城" in sh.name else ENTITY_LISI
        for r in range(2, sh.nrows):
            name = text(sh.cell_value(r, I_NAME))
            if not is_person_row(name):
                continue
            rows.append(
                InsuranceRow(
                    entity=entity,
                    dept=text(sh.cell_value(r, I_DEPT)),
                    name=name,
                    id_card=norm_id_card(sh.cell_value(r, I_ID)),
                    base=num(sh.cell_value(r, I_BASE)),
                    personal_total=num(sh.cell_value(r, I_PERSONAL_TOTAL)),
                )
            )
    return rows


def load_fund(path: pathlib.Path) -> list[FundRow]:
    wb = xlrd.open_workbook(str(path))
    rows: list[FundRow] = []
    for sh in wb.sheets():
        entity = ENTITY_JUANCHENG if ("分公司" in sh.name or "鄄城" in sh.name) else ENTITY_LISI
        for r in range(2, sh.nrows):
            name = text(sh.cell_value(r, G_NAME))
            if not is_person_row(name):
                continue
            rows.append(
                FundRow(
                    entity=entity,
                    dept=text(sh.cell_value(r, G_DEPT)),
                    name=name,
                    id_card=norm_id_card(sh.cell_value(r, G_ID)),
                    base=num(sh.cell_value(r, G_BASE)),
                    personal=num(sh.cell_value(r, G_PERSONAL)),
                )
            )
    return rows
def load_platform_users() -> dict[str, tuple[int, str]]:
    """从平台 users 表取 姓名 → (id, 登录名)。DB 连不上就返回空 dict（不阻塞脚本）。"""
    try:
        sys.path.insert(0, str(pathlib.Path(__file__).resolve().parents[2]))
        from sqlalchemy import text as sa_text  # type: ignore

        from app.core.database import SessionLocal  # type: ignore
    except Exception as exc:  # pragma: no cover - 环境缺失时降级
        print(f"[warn] 无法导入平台 DB 层，跳过 users 匹配：{exc}", flush=True)
        return {}

    mapping: dict[str, tuple[int, str]] = {}
    try:
        with SessionLocal() as db:
            rows = db.execute(
                sa_text(
                    "SELECT id, username, real_name, dingtalk_id FROM ark_users "
                    "WHERE is_active = 1 AND deleted_at IS NULL"
                )
            ).fetchall()
        for row in rows:
            real = (row[2] or "").strip()
            if real:
                # 同名账号取第一个，重名交由 H 类异常人工判定
                mapping.setdefault(real, (int(row[0]), row[1] or ""))
    except Exception as exc:
        print(f"[warn] 查询 ark_users 失败，跳过匹配：{exc}", flush=True)
    return mapping


def build_profiles(
    details: list[DetailRow],
    summaries: list[SummaryRow],
    insurance: list[InsuranceRow],
    fund: list[FundRow],
    users: dict[str, tuple[int, str]],
) -> tuple[list[Profile], list[Anomaly]]:
    anomalies: list[Anomaly] = []
    summary_by_name = {s.name: s for s in summaries}
    ins_by_name: dict[str, list[InsuranceRow]] = defaultdict(list)
    for i in insurance:
        ins_by_name[i.name].append(i)
    fund_by_name: dict[str, list[FundRow]] = defaultdict(list)
    for f in fund:
        fund_by_name[f.name].append(f)

    profiles: list[Profile] = []
    for d in details:
        p = Profile(
            emp_no=d.emp_no,
            name=d.name,
            dept_detail=d.dept_detail,
            position=d.position,  # B4：明细表职务为单一真相源
            hire_date=d.hire_date,
            base_salary_march=d.base_salary,
            social_march=abs(d.social),
            fund_march=abs(d.fund),
        )

        s = summary_by_name.get(d.name)
        if s:
            p.dept_group = s.dept_group
            p.bank_card = s.bank_card
            if s.hire_date and d.hire_date and s.hire_date != d.hire_date:
                anomalies.append(
                    Anomaly(
                        "D",
                        "入职日期两表冲突",
                        d.name,
                        f"明细={d.hire_date} / 汇总={s.hire_date}",
                        "以 HR 档案原件为准，确认后填 profile.hire_date（影响工龄与月中入职计薪）",
                    )
                )
            if s.position and d.position and s.position != d.position:
                anomalies.append(
                    Anomaly(
                        "E",
                        "职务两表口径不同代",
                        d.name,
                        f"明细={d.position} / 汇总={s.position}",
                        "明细用新职级体系、汇总用旧粗分类，属系统性差异；取明细为准即可，无需逐人处理",
                    )
                )
        else:
            anomalies.append(
                Anomaly(
                    "D",
                    "汇总表缺失该人",
                    d.name,
                    "明细表有此人，汇总表未匹配到同名行",
                    "确认是否漏列（影响银行卡号与大部门归属反推）",
                )
            )

        ins_rows = ins_by_name.get(d.name, [])
        if ins_rows:
            p.insurance_entity = ins_rows[0].entity
            p.id_card = ins_rows[0].id_card
        elif abs(d.social) > 0:
            anomalies.append(
                Anomaly(
                    "G",
                    "工资表扣社保但社保表无此人",
                    d.name,
                    f"明细社保={d.social}，社保明细表未匹配",
                    "核实参保状态（可能同名不同人或社保表漏列）",
                )
            )
        else:
            anomalies.append(
                Anomaly(
                    "G",
                    "发薪未参保",
                    d.name,
                    "工资表社保为 0 且社保表无此人",
                    "新入职常见，确认是否应参保；确认后 profile 标记参保主体",
                )
            )

        fund_rows = fund_by_name.get(d.name, [])
        if fund_rows:
            p.fund_included = True
            if not p.id_card:
                p.id_card = fund_rows[0].id_card
            if not p.insurance_entity:
                p.insurance_entity = fund_rows[0].entity
            # C：身份证跨表不一致（§2.5 错误 3，刘也 vs 姜婷）
            if ins_rows and ins_rows[0].id_card and fund_rows[0].id_card:
                if ins_rows[0].id_card != fund_rows[0].id_card:
                    anomalies.append(
                        Anomaly(
                            "C",
                            "身份证号跨表不一致",
                            d.name,
                            f"社保表={mask_secret(ins_rows[0].id_card)} / "
                            f"公积金表={mask_secret(fund_rows[0].id_card)}",
                            "以身份证原件为准；错误方需通知代缴机构更正",
                        )
                    )
        elif abs(d.fund) > 0:
            anomalies.append(
                Anomaly(
                    "G",
                    "工资表扣公积金但公积金表无此人",
                    d.name,
                    f"明细公积金={d.fund}",
                    "核实公积金缴存状态",
                )
            )

        if d.name in users:
            p.user_id, p.user_login = users[d.name]
        elif users:
            anomalies.append(
                Anomaly(
                    "H",
                    "平台账号未匹配",
                    d.name,
                    "users 表按 real_name 未找到在职账号",
                    "无平台账号属正常（如质检岗）；有账号则手动绑定，影响钉钉 userid 回填",
                )
            )

        profiles.append(p)
    return profiles, anomalies
def detect_duplicates(
    profiles: list[Profile], details: list[DetailRow]
) -> list[Anomaly]:
    """A/B 类：工号撞号、银行卡撞号（§2.5 错误 2、4，资金风险级）。"""
    anomalies: list[Anomaly] = []

    by_norm: dict[str, list[DetailRow]] = defaultdict(list)
    for d in details:
        if d.emp_no:
            by_norm[d.emp_no].append(d)
    for emp_no, group in sorted(by_norm.items()):
        if len(group) > 1:
            raws = ", ".join(f"{g.name}(原值 {g.emp_no_raw})" for g in group)
            anomalies.append(
                Anomaly(
                    "A",
                    "工号重复",
                    " / ".join(g.name for g in group),
                    f"归一化后同为 {emp_no}：{raws}",
                    "重新分配工号并确保全局唯一（档案 emp_no 建唯一索引）",
                )
            )

    by_bank: dict[str, list[Profile]] = defaultdict(list)
    for p in profiles:
        if p.bank_card:
            by_bank[p.bank_card].append(p)
    for bank, group in by_bank.items():
        if len(group) > 1:
            anomalies.append(
                Anomaly(
                    "B",
                    "银行卡号重复（资金风险级）",
                    " / ".join(g.name for g in group),
                    f"共用卡号 {mask_secret(bank)}",
                    "立即核实真实卡号——按此发放会把钱打错人（§2.5 错误 2）",
                )
            )

    by_id: dict[str, list[Profile]] = defaultdict(list)
    for p in profiles:
        if p.id_card:
            by_id[p.id_card].append(p)
    for id_card, group in by_id.items():
        if len(group) > 1:
            anomalies.append(
                Anomaly(
                    "C",
                    "身份证号重复",
                    " / ".join(g.name for g in group),
                    f"共用证号 {mask_secret(id_card)}",
                    "核实原件；hash 列建唯一索引后此类数据无法入库",
                )
            )
    return anomalies


def detect_non_payroll(
    profiles: list[Profile], insurance: list[InsuranceRow], fund: list[FundRow]
) -> list[Anomaly]:
    """F 类：参保但不在发薪名单（§2.2 的 8 人白名单候选）。"""
    payroll_names = {p.name for p in profiles}
    anomalies: list[Anomaly] = []
    seen: set[str] = set()
    for row in insurance:
        if row.name not in payroll_names and row.name not in seen:
            seen.add(row.name)
            anomalies.append(
                Anomaly(
                    "F",
                    "参保未发薪（白名单候选）",
                    row.name,
                    f"社保表 {row.entity}（表内部门={row.dept}），"
                    f"个人总计={row.personal_total}，本工资表无此人",
                    "确认后建档并标记 payroll_included=False——"
                    "按人标记，不按部门字段判定（§2.2）",
                )
            )
    for row in fund:
        if row.name not in payroll_names and row.name not in seen:
            seen.add(row.name)
            anomalies.append(
                Anomaly(
                    "F",
                    "缴公积金未发薪（白名单候选）",
                    row.name,
                    f"公积金表 {row.entity}，个人费用={row.personal}",
                    "同上，建档并标记 payroll_included=False",
                )
            )
    return anomalies


def verify_march_arithmetic(details: list[DetailRow]) -> list[Anomaly]:
    """交叉验算 3 月表自身的公式一致性——M3 引擎的验收基线在这里先摸清。

    检查：小计（增项）、小计（减项）、实发 三个公式列是否与分项自洽。
    不自洽的行说明人工表内部有错或有未记录的口径，M3 复算时会撞上。
    """
    anomalies: list[Anomaly] = []
    for d in details:
        add_expect = d.bonus + d.performance + d.seniority + d.attendance_bonus
        if abs(add_expect - d.add_subtotal) > 0.01:
            anomalies.append(
                Anomaly(
                    "V",
                    "增项小计不自洽",
                    d.name,
                    f"奖励{d.bonus}+绩效{d.performance}+工龄{d.seniority}"
                    f"+全勤{d.attendance_bonus}={add_expect:.2f}，"
                    f"表内小计={d.add_subtotal}",
                    "M3 复算会撞上此差异，需 HR 说明口径",
                )
            )
        ded_expect = d.social + d.fund + d.absence + d.other
        if abs(ded_expect - d.deduct_subtotal) > 0.01:
            anomalies.append(
                Anomaly(
                    "V",
                    "减项小计不自洽",
                    d.name,
                    f"社保{d.social}+公积金{d.fund}+缺勤{d.absence}"
                    f"+其他{d.other}={ded_expect:.2f}，表内小计={d.deduct_subtotal}",
                    "同上",
                )
            )
        net_expect = round(d.base_salary + d.add_subtotal + d.deduct_subtotal + d.subsidy)
        if abs(net_expect - d.net) > 0.51:
            anomalies.append(
                Anomaly(
                    "V",
                    "实发不自洽",
                    d.name,
                    f"底薪{d.base_salary}+增项{d.add_subtotal}+减项{d.deduct_subtotal}"
                    f"+补贴{d.subsidy}=round({net_expect})，表内实发={d.net}",
                    "M3 验收基线差异，优先查清",
                )
            )
    return anomalies
HEAD_FILL = PatternFill("solid", fgColor="1F4E79")
HEAD_FONT = Font(color="FFFFFF", bold=True)
WARN_FILL = PatternFill("solid", fgColor="FFF2CC")
RISK_FILL = PatternFill("solid", fgColor="F8CBAD")


def _write_sheet(ws, headers: list[str], rows: list[list[Any]], widths: list[int]) -> None:
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = HEAD_FILL
        cell.font = HEAD_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in rows:
        ws.append(row)
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.freeze_panes = "A2"


def write_profile_draft(profiles: list[Profile], dest: pathlib.Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "档案初稿"
    headers = [
        "工号", "姓名", "明细部门", "大部门(汇总表)", "职务(明细表·真相源)",
        "入职日期", "身份证(脱敏)", "银行卡(脱敏)", "参保主体",
        "是否发薪", "是否缴公积金", "3月底薪", "3月社保(个人)", "3月公积金(个人)",
        "平台账号", "HR 复核批注",
    ]
    rows = []
    for p in profiles:
        rows.append([
            p.emp_no, p.name, p.dept_detail, p.dept_group, p.position,
            p.hire_date.isoformat() if p.hire_date else "",
            mask_secret(p.id_card), mask_secret(p.bank_card), p.insurance_entity,
            "是" if p.payroll_included else "否",
            "是" if p.fund_included else "否",
            p.base_salary_march, p.social_march, p.fund_march,
            p.user_login or "", "",
        ])
    _write_sheet(
        ws, headers, rows,
        [8, 10, 14, 14, 16, 12, 22, 24, 26, 9, 12, 11, 13, 15, 14, 30],
    )
    dest.parent.mkdir(parents=True, exist_ok=True)
    wb.save(dest)


def write_anomalies(anomalies: list[Anomaly], dest: pathlib.Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "异常清单"
    headers = ["类别码", "异常类型", "涉及人员", "detail", "建议处理", "HR 批注(修正结论)"]
    order = {"B": 0, "C": 1, "A": 2, "V": 3, "D": 4, "F": 5, "G": 6, "H": 7}
    actionable = [a for a in anomalies if a.code not in INFO_CODES]
    ranked = sorted(actionable, key=lambda a: (order.get(a.code, 99), a.subject))
    rows = [[a.code, a.category, a.subject, a.detail, a.suggestion, ""] for a in ranked]
    _write_sheet(ws, headers, rows, [9, 26, 20, 60, 46, 30])
    # 资金风险级(B/C/A)整行高亮，HR 打开先看红的
    for idx, a in enumerate(ranked, start=2):
        fill = RISK_FILL if a.code in ("B", "C", "A") else (WARN_FILL if a.code == "V" else None)
        if fill:
            for cell in ws[idx]:
                cell.fill = fill

    ws2 = wb.create_sheet("按类别统计")
    counts: dict[tuple[str, str], int] = defaultdict(int)
    for a in actionable:
        counts[(a.code, a.category)] += 1
    _write_sheet(
        ws2, ["类别码", "异常类型", "条数"],
        [[c, cat, n] for (c, cat), n in sorted(counts.items(), key=lambda kv: -kv[1])],
        [9, 30, 8],
    )

    # 信息提示单独一页：系统性口径差异，不是逐人错误，混进异常清单会淹没资金风险行
    infos = sorted((a for a in anomalies if a.code in INFO_CODES), key=lambda a: a.subject)
    ws3 = wb.create_sheet("信息提示")
    _write_sheet(
        ws3,
        ["类别码", "类型", "涉及人员", "detail", "说明"],
        [[a.code, a.category, a.subject, a.detail, a.suggestion] for a in infos],
        [9, 26, 20, 60, 52],
    )
    wb.save(dest)


def main() -> int:
    parser = argparse.ArgumentParser(description="薪资档案反导入（只读，产出 Excel 供 HR 复核）")
    parser.add_argument("--materials", type=pathlib.Path, default=DEFAULT_MATERIALS)
    parser.add_argument("--out", type=pathlib.Path, default=OUT_DIR)
    parser.add_argument("--skip-users", action="store_true", help="跳过平台 users 匹配")
    args = parser.parse_args()

    for fname in (F_PAYROLL, F_INSURANCE, F_FUND):
        if not (args.materials / fname).exists():
            print(f"[error] 缺少材料文件：{args.materials / fname}", flush=True)
            return 1

    details, summaries = load_payroll(args.materials / F_PAYROLL)
    insurance = load_insurance(args.materials / F_INSURANCE)
    fund = load_fund(args.materials / F_FUND)
    users = {} if args.skip_users else load_platform_users()

    print(
        f"解析完成：明细 {len(details)} 人 / 汇总 {len(summaries)} 人 / "
        f"社保 {len(insurance)} 条 / 公积金 {len(fund)} 条 / 平台账号 {len(users)} 个",
        flush=True,
    )

    profiles, anomalies = build_profiles(details, summaries, insurance, fund, users)
    anomalies += detect_duplicates(profiles, details)
    anomalies += detect_non_payroll(profiles, insurance, fund)
    anomalies += verify_march_arithmetic(details)

    draft_path = args.out / "salary_profile_draft.xlsx"
    anomaly_path = args.out / "salary_anomalies.xlsx"
    write_profile_draft(profiles, draft_path)
    write_anomalies(anomalies, anomaly_path)

    by_code: dict[str, int] = defaultdict(int)
    for a in anomalies:
        by_code[a.code] += 1
    n_action = sum(n for c, n in by_code.items() if c not in INFO_CODES)
    n_info = sum(n for c, n in by_code.items() if c in INFO_CODES)
    print(f"\n档案初稿：{len(profiles)} 人 → {draft_path}", flush=True)
    print(f"异常清单：{n_action} 条待处理 + {n_info} 条信息提示 → {anomaly_path}", flush=True)
    for code in sorted(by_code):
        tag = "(信息提示)" if code in INFO_CODES else ""
        print(f"  {code}: {by_code[code]} 条{tag}", flush=True)
    if not users:
        print("\n[note] 未匹配平台账号（DB 未连或 --skip-users），H 类异常未检测", flush=True)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
