"""导出「旧标签 → 新体系」映射确认表（给设计部逐行确认）

用法（backend/ 目录下）：
    python scripts/tag_taxonomy/export_mapping.py [输出路径]

默认输出 ../tmp/tag_mapping_review.xlsx。
黄色行 = 需设计部重点确认（REVIEW_NOTES）；每行有「确认结果」列（默认=同意）。
"""

import os
import sys

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from sqlalchemy import text

from app.core.database import SessionLocal
from scripts.tag_taxonomy.mapping_def import MAPPING, REVIEW_NOTES, normalize_color

YELLOW = PatternFill("solid", start_color="FFF3B0")
HEADER_FONT = Font(bold=True, name="Arial")
DIM_LABELS = {
    "content_category": "内容大类", "content_type": "内容子类", "product_type": "产品类型",
    "color_code": "色号", "color_family": "色系", "texture": "纹理造型",
    "shoot_style": "拍摄风格", "process_step": "工艺环节", "theme": "节日营销主题",
    "year": "年份", "media_trait": "媒体特性",
}
OLD_DIM_LABELS = {"asset_type": "素材类型", "asset_type_2": "素材子类", "color": "颜色标签", "others": "综合标签"}


def _usage_counts(db) -> dict[tuple[str, str], int]:
    rows = db.execute(text("""
        SELECT d.name AS dim_name, v.value, COUNT(at.asset_id) AS cnt
        FROM ark_tag_values v
        JOIN ark_tag_dimensions d ON d.id = v.dimension_id
        LEFT JOIN ark_asset_tags at ON at.tag_value_id = v.id
        WHERE d.name IN ('asset_type', 'asset_type_2', 'color', 'others')
        GROUP BY v.id
    """)).fetchall()
    return {(r.dim_name, r.value): r.cnt for r in rows}


def _fmt_targets(targets) -> str:
    if not targets:
        return "（不映射，信息已由其他维度/字段承载）"
    return "；".join(f"{DIM_LABELS.get(d, d)}={v}" for d, v, _p in targets)


def _sheet_header(ws, cols):
    ws.append(cols)
    for c in ws[1]:
        c.font = HEADER_FONT
    ws.freeze_panes = "A2"


def main():
    out = sys.argv[1] if len(sys.argv) > 1 else os.path.join(
        os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))), "tmp", "tag_mapping_review.xlsx")
    os.makedirs(os.path.dirname(out), exist_ok=True)

    db = SessionLocal()
    usage = _usage_counts(db)
    db.close()

    wb = Workbook()

    # Sheet1: 分类映射（素材类型/子类/综合标签）
    ws = wb.active
    ws.title = "分类映射"
    _sheet_header(ws, ["旧维度", "旧标签值", "使用素材数", "映射到新体系", "备注（需确认项）", "确认结果"])
    for (old_dim, old_val), targets in sorted(MAPPING.items(), key=lambda x: (x[0][0], -usage.get(x[0], 0))):
        cnt = usage.get((old_dim, old_val), 0)
        note = REVIEW_NOTES.get((old_dim, old_val), "")
        row = [OLD_DIM_LABELS.get(old_dim, old_dim), old_val, cnt, _fmt_targets(targets), note, "同意"]
        ws.append(row)
        if note:
            for cell in ws[ws.max_row]:
                cell.fill = YELLOW

    # Sheet2: 色号规范化
    ws2 = wb.create_sheet("色号规范化")
    _sheet_header(ws2, ["旧色号值", "使用素材数", "规范化结果", "备注（需确认项）", "确认结果"])
    color_vals = sorted([(k[1], v) for k, v in usage.items() if k[0] == "color"], key=lambda x: -x[1])
    for old_val, cnt in color_vals:
        targets = normalize_color(old_val)
        note = REVIEW_NOTES.get(("color", old_val), "")
        changed = [f"{DIM_LABELS[d]}={v}" for d, v, _p in targets]
        result = "；".join(changed)
        if len(targets) == 1 and targets[0][1] == old_val:
            result = "（不变）" + result
        ws2.append([old_val, cnt, result, note, "同意"])
        if note:
            for cell in ws2[ws2.max_row]:
                cell.fill = YELLOW

    # Sheet3: 放弃清单（有使用量但未进 MAPPING 的旧值）
    ws3 = wb.create_sheet("放弃清单")
    _sheet_header(ws3, ["旧维度", "旧标签值", "使用素材数", "放弃原因", "确认结果"])
    for (old_dim, old_val), cnt in sorted(usage.items(), key=lambda x: -x[1]):
        if old_dim == "color" or (old_dim, old_val) in MAPPING:
            continue
        reason = "使用量为 0，无素材受影响" if cnt == 0 else "未纳入映射（信息已由内容子类+年份承载，场次名可查 remark）"
        ws3.append([OLD_DIM_LABELS.get(old_dim, old_dim), old_val, cnt, reason, "同意"])
        if cnt > 0:
            for cell in ws3[ws3.max_row]:
                cell.fill = YELLOW

    for sheet in wb.worksheets:
        widths = {"A": 14, "B": 34, "C": 12, "D": 52, "E": 40, "F": 12}
        for col, w in widths.items():
            sheet.column_dimensions[col].width = w
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                cell.font = Font(name="Arial")

    wb.save(out)
    print(f"已导出: {out}", flush=True)


if __name__ == "__main__":
    main()
