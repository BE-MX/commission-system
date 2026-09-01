from io import BytesIO
from urllib.parse import unquote

from openpyxl import load_workbook

from app.domestic import router as domestic_router
from app.domestic.export_service import build_order_workbook


def _order_detail():
    return {
        "domestic_no": "DO20260819-001",
        "order_no": "322-3",
        "order_date": "2026-08-19",
        "customer_name": "尚都",
        "order_category": "special",
        "order_category_label": "特单",
        "order_type": "first_order",
        "order_type_label": "首单",
        "order_channel": "wechat",
        "order_channel_label": "微信",
        "customer_balance": 4506,
        "total_amount": 1698,
        "remark": "整单备注",
        "items": [
            {
                "line_code": "A1",
                "product_name": "头套 / 递顶 / L / 20厘米",
                "attrs": {
                    "product_type": "cap",
                    "craft": "递顶",
                    "net_color": "浅棕",
                    "size": "L",
                    "length": "20厘米",
                    "density": "中",
                    "hair_style_series": "直发",
                },
                "order_qty": 2,
                "unit_price": 849,
                "line_amount": 1698,
                "hairstyle": "短直发",
                "color": "自然色",
                "style_requirement": "前额和鬓角缝粘胶点",
                "remark": "明细备注",
                "ship_time": "2026-08-20T10:30:00",
                "ship_weight": 54.5,
            },
            {
                "line_code": "A2",
                "product_name": "发片 / 机制 / 15厘米",
                "attrs": {
                    "product_type": "piece",
                    "craft": "机制",
                    "net_color": None,
                    "size": "8×10",
                    "length": "15厘米",
                    "density": "轻",
                },
                "order_qty": 1,
                "unit_price": 300,
                "line_amount": 300,
                "hairstyle": None,
                "color": None,
                "style_requirement": None,
                "remark": None,
                "ship_time": None,
                "ship_weight": None,
            },
        ],
    }


def test_build_order_workbook_matches_requisition_layout_and_fields():
    workbook = load_workbook(build_order_workbook(_order_detail(), "Rice"), data_only=False)
    sheet = workbook.active

    assert sheet.title == "内贸订单领货单"
    assert sheet["B1"].value == "内贸订单领货单"
    assert "下单日期：2026/08/19" in sheet["A2"].value
    assert "客户订单号：322-3" in sheet["A2"].value
    assert "系统单号：DO20260819-001" in sheet["A2"].value
    assert "申请人：Rice" in sheet["A2"].value
    assert "客户：尚都" in sheet["A2"].value
    assert "订单类别：特单" in sheet["A3"].value
    assert "订单类型：首单" in sheet["A3"].value
    assert "订单渠道：微信" in sheet["A3"].value
    assert "订单金额：¥1,698.00" in sheet["A3"].value
    assert "客户余额：¥4,506.00" in sheet["A3"].value

    assert [sheet.cell(4, col).value for col in range(1, 17)] == [
        "明细号", "产品类型", "产品名称", "工艺/尺寸", "发长", "网帽颜色",
        "头套尺寸", "发量", "发型系列", "数量", "单价", "小计", "发型", "颜色",
        "发型要求", "备注",
    ]
    assert [sheet.cell(5, col).value for col in range(1, 17)] == [
        "A1", "头套", "头套 / 递顶 / L / 20厘米", "递顶", "20厘米", "浅棕", "L",
        "中", "直发", 2, 849, 1698, "短直发", "自然色", "前额和鬓角缝粘胶点", "明细备注",
    ]
    assert sheet["B6"].value == "发片"
    assert sheet["D6"].value == "机制"
    assert sheet["F6"].value is None
    assert sheet["G6"].value is None
    assert sheet["H6"].value is None
    assert sheet["I6"].value is None
    assert sheet.row_dimensions[5].height >= 75
    assert sheet.page_setup.orientation == "landscape"
    assert sheet.print_area == "'内贸订单领货单'!$A$1:$P$8"
    assert "整单备注" in sheet["A8"].value


def test_build_order_workbook_uses_safe_excel_text_for_user_content():
    detail = _order_detail()
    detail["items"][0]["remark"] = "=HYPERLINK(\"https://example.com\")"

    workbook = load_workbook(build_order_workbook(detail, "+SUM(1,1)"), data_only=False)
    sheet = workbook.active

    assert sheet["P5"].value == "'=HYPERLINK(\"https://example.com\")"
    assert "申请人：'+SUM(1,1)" in sheet["A2"].value
    assert sheet["A2"].data_type != "f"


def test_export_order_returns_named_xlsx(monkeypatch):
    detail = _order_detail() | {"created_by_name": "下单员"}
    captured = {}
    monkeypatch.setattr(domestic_router.order_service, "get_order_detail", lambda db, order_id: detail)

    def fake_workbook(data, applicant_name):
        captured["args"] = (data, applicant_name)
        return BytesIO(b"xlsx")

    monkeypatch.setattr(domestic_router.export_service, "build_order_workbook", fake_workbook)

    class FakeDb:
        committed = False

        def commit(self):
            self.committed = True

    db = FakeDb()
    response = domestic_router.export_order(7, db=db, _user={"sub": "9"})

    assert captured["args"] == (detail, "下单员")
    assert db.committed is True
    assert response.media_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert "内贸订单-DO20260819-001.xlsx" in unquote(response.headers["content-disposition"])


def test_build_order_workbook_preserves_long_requirements_for_printing():
    detail = _order_detail()
    detail["items"] = []
    long_requirement = "要求\n" * 666 + "要求"
    continuous_requirement = "要求" * 1000
    assert len(long_requirement) == 2000
    assert len(continuous_requirement) == 2000
    for index in range(1, 51):
        item = _order_detail()["items"][0] | {
            "line_code": f"A{index}",
            "style_requirement": (
                long_requirement if index == 1
                else continuous_requirement if index == 2
                else f"要求{index}"
            ),
        }
        detail["items"].append(item)

    workbook = load_workbook(build_order_workbook(detail, "下单员"), data_only=False)
    order_sheet = workbook["内贸订单领货单"]
    requirement_sheet = workbook["完整要求"]

    assert order_sheet.row_dimensions[5].height > 75
    assert order_sheet.print_area == "'内贸订单领货单'!$A$1:$P$56"
    assert order_sheet.print_title_rows == "$1:$4"
    assert requirement_sheet["A2"].value == "A1"
    assert requirement_sheet["B2"].value == "发型要求"
    requirement_rows = [
        row for row in range(2, requirement_sheet.max_row + 1)
        if requirement_sheet.cell(row, 1).value == "A1"
    ]
    assert len(requirement_rows) > 1
    assert "".join(requirement_sheet.cell(row, 3).value for row in requirement_rows) == long_requirement
    assert max(requirement_sheet.row_dimensions[row].height for row in requirement_rows) <= 300
    continuous_rows = [
        row for row in range(2, requirement_sheet.max_row + 1)
        if requirement_sheet.cell(row, 1).value == "A2"
    ]
    assert len(continuous_rows) > 1
    assert "".join(requirement_sheet.cell(row, 3).value for row in continuous_rows) == continuous_requirement
