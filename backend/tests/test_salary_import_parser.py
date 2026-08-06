"""社保/公积金 Excel 解析器的口径测试（M2-b）。

测的不是「能不能读出数」，而是**读错的时候会不会炸**：
列名换序、列名撞车、分项对不上、合计对不上、身份证撞号——
这几件事任何一件静默通过，66 个人的社保扣款就都错了。

用 openpyxl 在内存里造 .xlsx，不依赖 HR 原件（原件不进仓库）。
真实原件的验证在 M2 提交说明里另有记录。
"""

from decimal import ROUND_HALF_UP, Decimal
from io import BytesIO

import pytest
from openpyxl import Workbook

from app.core.config import get_settings
from app.salary import import_service as imp

_TEST_ENC_KEY = "dGVzdC1zYWxhcnktZW5jLWtleS0zMi1ieXRlcy0hIQ=="
_TEST_HASH_KEY = "test-salary-hash-key"


@pytest.fixture(autouse=True)
def salary_keys(monkeypatch):
    settings = get_settings()
    monkeypatch.setattr(settings, "ARK_SALARY_ENCRYPTION_KEY", _TEST_ENC_KEY)
    monkeypatch.setattr(settings, "ARK_SALARY_HASH_KEY", _TEST_HASH_KEY)


# ---------------------------------------------------------------------------
# 造表工具：列顺序与表头文案照抄 HR 原件（2026-03 实件核对过）
# ---------------------------------------------------------------------------

INS_HEADER = [
    "序号", "业务年费", "缴费年费", "缴费类型", "交费单位", "部门", "姓名", "身份证号",
    "社保缴费\n基数", "企业养老个人缴费", "企业养老单位缴费", "失业个人缴费", "失业单位缴费",
    "工伤单位缴费", "社保滞纳金", "医保缴费基数", "医保单位应缴", "医保个人应缴", "医保滞纳金",
    "个人总计", "单位总计",
]

FUND_HEADER = [
    "序号", "业务年月", "缴费年月", "缴费类型", "交费单位", "部门", "姓名", "身份证号",
    "月缴费\n基数", "总计缴纳", "个人费用", "单位缴纳",
]


def _cent(v: Decimal) -> Decimal:
    return v.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def ins_row(idx, name, id_card, base, entity="青岛丽丝发贸易", dept="外贸部"):
    """按实测费率造一行：养老个人 8%、失业个人 0.3%、医保个人 2%、工伤单位 0.32%。

    每个险种先各自舍入到分再相加——这是源表的真实行为：5345 × 0.3% = 16.035
    在表里是 16.04，个人总计 550.54 而不是按未舍入值加出来的 550.53。
    造数据时不照此办理，恒等式测试测的就是个假口径。
    """
    b = Decimal(str(base))
    pension_p = _cent(b * Decimal("0.08"))
    unemp_p = _cent(b * Decimal("0.003"))
    med_p = _cent(b * Decimal("0.02"))
    return [
        idx, 202603, 202603, "正常应缴", entity, dept, name, id_card,
        float(b), float(pension_p), float(_cent(b * Decimal("0.16"))),
        float(unemp_p), float(_cent(b * Decimal("0.007"))), float(_cent(b * Decimal("0.0032"))), 0.0,
        float(b), float(_cent(b * Decimal("0.08"))), float(med_p), 0.0,
        float(pension_p + unemp_p + med_p),
        float(_cent(b * Decimal("0.16")) + _cent(b * Decimal("0.007"))
              + _cent(b * Decimal("0.0032")) + _cent(b * Decimal("0.08"))),
    ]


def fund_row(idx, name, id_card, base=2200, entity="青岛丽丝发贸易"):
    b = Decimal(str(base))
    half = _cent(b * Decimal("0.05"))
    return [
        idx, 202603, 202603, "正常应缴", entity, "外贸部", name, id_card,
        float(b), float(half * 2), float(half), float(half),
    ]


def build(header, rows, *, sheet_name="社保", title="2026年3月社保医保明细", total_col=None):
    """标题行 + 表头行 + 数据行 + 合计行，与原件同构。"""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append([title] + [""] * (len(header) - 1))
    ws.append(header)
    for r in rows:
        ws.append(r)
    if total_col is not None:
        footer = [""] * len(header)
        footer[total_col] = float(sum(Decimal(str(r[total_col])) for r in rows))
        ws.append(footer)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# 正路：解析出的数与源表一致
# ---------------------------------------------------------------------------

def test_insurance_happy_path():
    data = build(INS_HEADER, [
        ins_row(1, "孙正华", "370213199103125218", 5345),
        ins_row(2, "王天", "220284199911103423", 4504),
    ], total_col=19)
    result = imp.parse_insurance(data)

    assert len(result.rows) == 2
    first = result.rows[0]
    # 实件核对值：5345 基数 → 427.6 + 16.04 + 106.9 = 550.54
    assert first.personal_total == Decimal("550.54")
    assert first.base_amount == Decimal("5345.00")
    assert first.name == "孙正华"
    assert first.dept_text == "外贸部"
    assert first.entity == imp.ENTITY_LISI
    assert first.warnings == []

    sheet = result.sheets[0]
    assert sheet.person_count == 2
    # 550.54 + 463.91 = 1014.45
    assert sheet.personal_sum == Decimal("1014.45")
    assert sheet.personal_total_in_file == Decimal("1014.45")
    assert sheet.warnings == []
    assert sheet.rates_observed["pension_personal"] == "8.00%"
    assert sheet.rates_observed["unemployment_personal"] == "0.30%"
    assert sheet.rates_observed["medical_personal"] == "2.00%"


def test_fund_happy_path():
    data = build(FUND_HEADER, [
        fund_row(1, "孙正华", "370213199103125218"),
        fund_row(2, "王天", "220284199911103423"),
    ], sheet_name="公积金", title="2026年3月公积金", total_col=10)
    result = imp.parse_fund(data)

    assert len(result.rows) == 2
    assert result.rows[0].personal_total == Decimal("110.00")
    assert result.rows[0].company_total == Decimal("110.00")
    assert result.sheets[0].personal_sum == Decimal("220.00")
    assert result.sheets[0].rates_observed["fund_personal"] == "5.00%"
    assert result.sheets[0].warnings == []


# ---------------------------------------------------------------------------
# PII：明文绝不出现在解析结果里
# ---------------------------------------------------------------------------

def test_id_card_never_returned_in_plaintext():
    plain = "370213199103125218"
    data = build(INS_HEADER, [ins_row(1, "孙正华", plain, 5345)], total_col=19)
    row = imp.parse_insurance(data).rows[0]

    assert plain not in repr(row)
    assert row.id_card_masked == "370***********5218"
    assert row.id_card_hash and len(row.id_card_hash) == 64
    assert row.id_card_cipher and plain not in row.id_card_cipher


def test_id_card_hash_matches_profile_hash():
    """解析出的哈希必须与档案侧同一把钥匙算出的一致，否则 M2-c 匹配全落空。"""
    from app.salary import pii

    plain = "37292919900916721x"  # 末位小写 x，归一化后应与大写 X 同哈希
    data = build(INS_HEADER, [ins_row(1, "刘美美", plain, 8000)], total_col=19)
    row = imp.parse_insurance(data).rows[0]

    assert row.id_card_hash == pii.hash_pii(pii.normalize_id_card("37292919900916721X"))


# ---------------------------------------------------------------------------
# 结构性守卫：列错了必须炸，不能静默导入
# ---------------------------------------------------------------------------

def test_column_order_change_is_tolerated():
    """列换序不该报错——定位靠表头文字，不靠位置。"""
    order = [0, 1, 2, 3, 4, 5, 6, 7, 8, 19, 20, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18]
    header = [INS_HEADER[i] for i in order]
    raw = ins_row(1, "孙正华", "370213199103125218", 5345)
    data = build(header, [[raw[i] for i in order]], total_col=9)

    row = imp.parse_insurance(data).rows[0]
    assert row.personal_total == Decimal("550.54")
    assert row.base_amount == Decimal("5345.00")


def test_missing_required_column_raises():
    header = list(INS_HEADER)
    header[19] = "备注"  # 抹掉「个人总计」
    data = build(header, [ins_row(1, "孙正华", "370213199103125218", 5345)])
    with pytest.raises(imp.SalaryImportError, match="个人总计"):
        imp.parse_insurance(data)


def test_ambiguous_column_raises():
    """两列都叫「个人总计」时必须拒绝，不许挑一个——挑错就是扣错钱。"""
    header = list(INS_HEADER)
    header[20] = "个人总计"
    data = build(header, [ins_row(1, "孙正华", "370213199103125218", 5345)])
    with pytest.raises(imp.SalaryImportError, match="命中多列"):
        imp.parse_insurance(data)


def test_no_header_raises():
    wb = Workbook()
    wb.active.append(["随便什么", "不相干的表"])
    buf = BytesIO()
    wb.save(buf)
    with pytest.raises(imp.SalaryImportError):
        imp.parse_insurance(buf.getvalue())


def test_identity_mismatch_majority_raises():
    """过半行分项之和对不上合计 = 列映射错，直接终止。"""
    rows = []
    for i in range(4):
        r = ins_row(i + 1, f"测试{i}", f"37021319910312521{i}", 5000)
        r[19] = 9999.0  # 个人总计被改成假值
        rows.append(r)
    data = build(INS_HEADER, rows, total_col=19)
    with pytest.raises(imp.SalaryImportError, match="列映射错误"):
        imp.parse_insurance(data)


def test_identity_mismatch_minority_warns_only():
    """个别行对不上是补缴/滞纳金，逐行标注即可，不该拦住整批。"""
    rows = [ins_row(i + 1, f"测试{i}", f"37021319910312521{i}", 5000) for i in range(4)]
    rows[0][19] = float(Decimal(str(rows[0][19])) + Decimal("50"))  # 一行补缴
    data = build(INS_HEADER, rows, total_col=19)

    result = imp.parse_insurance(data)
    assert len(result.rows) == 4
    assert any("分项之和" in w for w in result.rows[0].warnings)
    assert result.rows[1].warnings == []
    assert any("分项之和" in w for w in result.sheets[0].warnings)


def test_footer_reconciliation_detects_missing_rows():
    """源表合计行 ≠ 解析合计 → 说明有行没被识别，必须告警。"""
    rows = [ins_row(1, "孙正华", "370213199103125218", 5345)]
    data = build(INS_HEADER, rows, total_col=19)
    # 手工把合计行改大：模拟「有一行被当成非人员跳过了」
    from openpyxl import load_workbook

    wb = load_workbook(BytesIO(data))
    wb.active.cell(row=4, column=20, value=9999.0)
    buf = BytesIO()
    wb.save(buf)

    result = imp.parse_insurance(buf.getvalue())
    assert any("个人合计对不上" in w for w in result.sheets[0].warnings)


# ---------------------------------------------------------------------------
# 数据质量：非人员行、撞号、多主体
# ---------------------------------------------------------------------------

def test_signature_and_total_rows_excluded():
    rows = [
        ins_row(1, "孙正华", "370213199103125218", 5345),
        ins_row(2, "合计", "", 0),
        ins_row(3, "制表：张三", "", 0),
    ]
    data = build(INS_HEADER, rows)
    result = imp.parse_insurance(data)
    assert [r.name for r in result.rows] == ["孙正华"]


def test_duplicate_id_card_flagged():
    """3 月公积金表刘也/姜婷撞号（设计文档 §2.5 错误 3）的自动检出。"""
    same = "370213199103125218"
    data = build(FUND_HEADER, [
        fund_row(1, "姜婷", same),
        fund_row(2, "刘也", same),
    ], sheet_name="公积金", title="公积金", total_col=10)

    result = imp.parse_fund(data)
    assert any("身份证重复" in w for w in result.warnings)
    assert any("姜婷" in w and "刘也" in w for w in result.warnings)


def test_entity_normalized_from_text():
    """「鄄城…青岛分公司」同时含「青岛」，不能被丽丝发规则截胡。"""
    assert imp.normalize_entity("鄄城莱莎发制品有限公司青岛分公司") == imp.ENTITY_JUANCHENG
    assert imp.normalize_entity("青岛丽丝发贸易") == imp.ENTITY_LISI
    assert imp.normalize_entity("", "分公司公积金") == imp.ENTITY_JUANCHENG
    assert imp.normalize_entity("某不认识的公司") is None


def test_multi_sheet_parsed_as_two_entities():
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "丽丝发"
    ws1.append(["2026年3月社保"] + [""] * 20)
    ws1.append(INS_HEADER)
    ws1.append(ins_row(1, "孙正华", "370213199103125218", 5345))
    ws2 = wb.create_sheet("鄄城分公司")
    ws2.append(["2026年3月社保"] + [""] * 20)
    ws2.append(INS_HEADER)
    ws2.append(ins_row(1, "刘美美", "371329198904080024", 8000, entity="鄄城莱莎发制品有限公司青岛分公司"))
    buf = BytesIO()
    wb.save(buf)

    result = imp.parse_insurance(buf.getvalue())
    assert len(result.sheets) == 2
    assert {s.entity for s in result.sheets} == {imp.ENTITY_LISI, imp.ENTITY_JUANCHENG}
    assert len(result.rows) == 2


def test_row_no_points_at_real_excel_row():
    """异常面板要告诉 HR「第几行」，行号必须是 Excel 里看到的那个。"""
    data = build(INS_HEADER, [
        ins_row(1, "孙正华", "370213199103125218", 5345),
        ins_row(2, "王天", "220284199911103423", 4504),
    ])
    rows = imp.parse_insurance(data).rows
    # 第 1 行标题、第 2 行表头 → 第一个人在第 3 行
    assert [r.row_no for r in rows] == [3, 4]


def test_empty_file_raises():
    with pytest.raises(imp.SalaryImportError, match="文件为空"):
        imp.parse_insurance(b"")


def test_unknown_format_raises():
    with pytest.raises(imp.SalaryImportError, match="无法识别的文件格式"):
        imp.parse_insurance(b"not an excel file at all")
