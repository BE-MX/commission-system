from datetime import date, datetime
from decimal import Decimal
import re

import pytest
from openpyxl import load_workbook
from sqlalchemy import text

from app.invoice import export_service, product_service, service, xiaoman_service
from app.invoice.models import Invoice, InvoiceItem, StdPrice, XiaomanSettings  # noqa: F401 - register metadata for test create_all
from app.invoice.schemas import InvoiceCreate, InvoiceItemPayload


def _extract_pdf_text(content: bytes) -> str:
    unicode_chunks = [
        bytes.fromhex(chunk.decode("ascii")).decode("utf-16-be")
        for chunk in re.findall(rb"<([0-9A-F]+)> Tj", content)
    ]
    latin_chunks = [
        chunk.decode("latin-1").replace("\\(", "(").replace("\\)", ")").replace("\\\\", "\\")
        for chunk in re.findall(rb"\((.*?)\) Tj", content)
    ]
    return "\n".join([*unicode_chunks, *latin_chunks])


def _seed_products(db):
    db.execute(text("""
        CREATE TABLE IF NOT EXISTS lsordertest.okki_products (
            product_id INTEGER PRIMARY KEY,
            product_no TEXT,
            name TEXT,
            model TEXT,
            color TEXT,
            size TEXT,
            unit TEXT,
            disable_flag INTEGER
        )
    """))
    db.execute(text("""
        CREATE TABLE IF NOT EXISTS lsordertest.okki_inventory (
            product_id INTEGER,
            sku_id INTEGER,
            disable_flag INTEGER
        )
    """))
    db.execute(text("""
        INSERT INTO lsordertest.okki_products
        (product_id, product_no, name, model, color, size, unit, disable_flag)
        VALUES
        (1, 'P001', 'Raw Hair/Body Wave', 'M1', 'Natural', '18', '100g', 0),
        (2, 'P002', 'Raw Hair/Straight', 'M1', 'Natural', '20', '100g', 0),
        (3, 'P003', 'Raw Hair/Body Wave', 'M2', 'Piano', '18', '120g', 0)
    """))
    db.execute(text("""
        INSERT INTO lsordertest.okki_inventory (product_id, sku_id, disable_flag)
        VALUES (1, 9001, 0), (2, 9002, 0), (3, 9003, 0)
    """))
    db.commit()


def test_product_filter_options_and_unique_match(db):
    _seed_products(db)

    options = product_service.get_filter_options(db, model="M1", color="Natural")
    assert options["models"] == ["M1"]
    assert options["colors"] == ["Natural"]
    assert options["sizes"] == ["18", "20"]
    assert options["units"] == ["100g"]

    match = product_service.match_product(db, model="M1", color="Natural", size="18", unit="100g")
    assert match["is_unique"] is True
    assert match["item"]["product_id"] == 1
    assert match["item"]["sku_id"] == 9001
    assert match["item"]["product_name"] == "Raw Hair/Body Wave"
    assert match["item"]["product_display"] == "Raw Hair"


def test_invoice_create_totals_and_validation(db):
    _seed_products(db)
    payload = InvoiceCreate(
        customer_id="CUST001",
        customer_name="Customer A",
        invoice_date=date(2026, 7, 2),
        items=[
            InvoiceItemPayload(
                product_id=1,
                sku_id=9001,
                product_name="Raw Hair/Body Wave",
                product_display="Raw Hair",
                net_weight_grams="100g",
                curl="Body Wave",
                model="M1",
                color="Natural",
                length="18",
                quantity=3,
                price_per_piece=Decimal("12.50"),
                discount_amount=Decimal("-2.00"),
            )
        ],
        internal_accessory=Decimal("3.00"),
        packaging_quantity=4,
        shipping_fee=Decimal("4.00"),
        surcharge_amount=Decimal("1.00"),
    )

    invoice = service.create_invoice(db, payload, user_id=1)
    db.flush()

    assert invoice.items[0].total_price == Decimal("35.50")
    assert invoice.internal_discount == Decimal("-2.00")
    assert invoice.packaging_quantity == 4
    assert invoice.total_amount == Decimal("43.50")
    assert invoice.status == "ready"
    assert service.validate_invoice(invoice) == []
    workbook = load_workbook(export_service.build_invoice_workbook(invoice), data_only=True)
    sheet = workbook["Invoice"]
    assert [sheet.cell(10, col).value for col in range(1, 12)][-2:] == ["Discount", "TotalPrice"]
    assert sheet["J11"].value == -2
    assert sheet["K11"].value == 35.5
    assert [sheet.cell(row, 10).value for row in range(12, 21)] == [
        "Hair Price", "Hair Discount", "Accessory Amount", "Accessory Discount",
        "Packaging Quantity", "Packaging", "Shipping Fee", "Handling Fee", "Total",
    ]
    html = export_service.build_print_html(invoice)
    assert "<th>Discount</th><th>TotalPrice</th>" in html
    assert "Packaging Quantity: 4" in html
    assert "Packaging: USD 3.00" in html
    pdf = export_service.build_invoice_pdf(invoice).getvalue()
    assert pdf.startswith(b"%PDF-1.4")
    pdf_text = _extract_pdf_text(pdf)
    assert "Packaging Quantity: 4" in pdf_text
    assert "Packaging: USD 3.00" in pdf_text


def test_mixed_invoice_exports_independent_accessory_table_and_full_summary(db):
    _seed_products(db)
    db.execute(text("""
        CREATE TABLE lsordertest.okki_product_skus (
            product_id INTEGER, sku_id INTEGER, disable_flag INTEGER
        )
    """))
    db.execute(text("""
        INSERT INTO lsordertest.okki_products
            (product_id, product_no, name, model, color, size, unit, disable_flag)
        VALUES
            (4, 'ACC004', 'Hair Gripper', '魔术贴', '黑色', NULL, NULL, 0)
    """))
    db.execute(text("""
        INSERT INTO lsordertest.okki_inventory (product_id, sku_id, disable_flag)
        VALUES (4, 9004, 0)
    """))
    db.execute(text("""
        INSERT INTO lsordertest.okki_product_skus (product_id, sku_id, disable_flag)
        VALUES (4, 9004, 0)
    """))
    db.add(StdPrice(
        product_kind="accessory",
        product_id=4,
        sku_id=9004,
        accessory_name="Hair Gripper",
        accessory_model="魔术贴",
        accessory_color="黑色",
        price=Decimal("0.0000"),
        currency="USD",
    ))
    db.flush()
    invoice = service.create_invoice(
        db,
        InvoiceCreate(
            customer_id="CUST001",
            customer_name="Customer A",
            invoice_date=date(2026, 7, 15),
            packaging_quantity=4,
            internal_accessory=Decimal("3"),
            shipping_fee=Decimal("7"),
            surcharge_amount=Decimal("2"),
            items=[
                InvoiceItemPayload(
                    product_id=1,
                    sku_id=9001,
                    product_name="Raw Hair/Body Wave",
                    product_display="Raw Hair",
                    net_weight_grams="100g",
                    curl="Body Wave",
                    model="M1",
                    color="Natural",
                    length="18",
                    quantity=1,
                    price_per_piece=Decimal("100"),
                    discount_amount=Decimal("-10"),
                ),
                InvoiceItemPayload(
                    product_kind="accessory",
                    item_type="stock",
                    product_id=4,
                    sku_id=9004,
                    product_name="Hair Gripper",
                    product_display="Hair Gripper",
                    model="Forged Model",
                    color="Forged Color",
                    quantity=2,
                    price_per_piece=Decimal("15"),
                    discount_amount=Decimal("-2"),
                ),
            ],
        ),
        user_id=1,
    )

    workbook = load_workbook(export_service.build_invoice_workbook(invoice), data_only=True)
    sheet = workbook["Invoice"]
    values = list(sheet.iter_rows(values_only=True))
    accessory_header = (
        "Name", "Model", "Color", "Standard Price", "Customer Price", "Transaction Price",
        "Quantity", "Discount", "TotalPrice",
    )
    assert any(tuple(row[:9]) == accessory_header for row in values)
    assert any(tuple(row[:9]) == (
        "Hair Gripper", "魔术贴", "黑色", 0, 0, 15, 2, -2, 28,
    ) for row in values)
    summary_labels = [row[9] for row in values if row[9]]
    assert summary_labels[-9:] == [
        "Hair Price", "Hair Discount", "Accessory Amount", "Accessory Discount",
        "Packaging Quantity", "Packaging", "Shipping Fee", "Handling Fee", "Total",
    ]
    flattened = " ".join(str(value) for row in values for value in row if value is not None)
    assert "9004" not in flattened

    html = export_service.build_print_html(invoice)
    assert "<h2>Accessories</h2>" in html
    assert "<th>Name</th><th>Model</th><th>Color</th>" in html
    assert "<th>Transaction Price</th>" in html
    assert '<td class="num">15</td>' in html
    assert "Accessory Amount: USD 30.00" in html
    assert "Accessory Discount: USD -2.00" in html
    assert "<td class=\"num\">0.0000</td>" in html
    assert "9004" not in html

    pdf = export_service.build_invoice_pdf(invoice).getvalue()
    pdf_text = _extract_pdf_text(pdf)
    assert "ACCESSORIES" in pdf_text
    assert "Name: Hair Gripper" in pdf_text
    assert "Model: 魔术贴" in pdf_text
    assert "Color: 黑色" in pdf_text
    assert (
        "Standard Price: 0.0000 | Customer Price: 0.0000 | Transaction Price: 15 | Quantity: 2 | "
        "Discount: -2.00 | TotalPrice: 28.00"
    ) in pdf_text
    assert "Accessory Amount: USD 30.00" in pdf_text
    assert "Accessory Discount: USD -2.00" in pdf_text
    assert "9004" not in pdf_text
    assert "魔术贴".encode("utf-16-be").hex().upper().encode("ascii") in pdf
    assert b"/Subtype /Type0" in pdf
    assert b"/BaseFont /STSong-Light" in pdf
    assert b"/Encoding /UniGB-UTF16-H" in pdf
    assert b"/Subtype /Image" in pdf


def test_invoice_pdf_paginates_without_omitting_rows_or_negative_coordinates():
    invoice = Invoice(
        invoice_no="INV-LONG-001",
        customer_id="CUST-LONG",
        customer_name="Long Invoice Customer",
        invoice_date=date(2026, 7, 15),
        currency="USD",
        internal_accessory=Decimal("0"),
        shipping_fee=Decimal("0"),
        surcharge_amount=Decimal("0"),
        total_amount=Decimal("50"),
    )
    for index in range(35):
        invoice.items.append(InvoiceItem(
            sort_order=index + 1,
            product_kind="hair",
            item_type="stock",
            product_name=f"Hair-{index:02d}",
            product_display=f"Hair-{index:02d}",
            net_weight_grams="100g",
            model="M1",
            color="Black",
            length="18",
            quantity=1,
            standard_price=Decimal("1"),
            customer_price=Decimal("1"),
            price_per_piece=Decimal("1"),
            discount_amount=Decimal("0"),
            total_price=Decimal("1"),
        ))
    long_name = "Accessory-14-" + "超长配件名称" * 18
    long_model = "魔术贴-" + "超长型号" * 14
    long_color = "黑色-" + "超长颜色" * 14
    for index in range(15):
        is_last = index == 14
        invoice.items.append(InvoiceItem(
            sort_order=36 + index,
            product_kind="accessory",
            item_type="stock",
            product_name=long_name if is_last else f"Accessory-{index:02d}",
            product_display=long_name if is_last else f"Accessory-{index:02d}",
            model=long_model if is_last else "Tool",
            color=long_color if is_last else "Black",
            quantity=7 if is_last else 1,
            standard_price=Decimal("123.4567") if is_last else Decimal("1"),
            customer_price=Decimal("124.4567") if is_last else Decimal("1"),
            price_per_piece=Decimal("1"),
            discount_amount=Decimal("-1.23") if is_last else Decimal("0"),
            total_price=Decimal("869.97") if is_last else Decimal("1"),
        ))

    pdf = export_service.build_invoice_pdf(invoice).getvalue()
    content = pdf.decode("latin-1")
    text_content = _extract_pdf_text(pdf)
    compact_text = text_content.replace("\n", "")

    assert "Hair-34" in text_content
    assert long_name in compact_text
    assert long_model in compact_text
    assert long_color in compact_text
    assert (
        "Standard Price: 123.4567 | Customer Price: 124.4567 | Transaction Price: 1 | Quantity: 7 | "
        "Discount: -1.23 | TotalPrice: 869.97"
    ) in text_content
    assert "魔术贴".encode("utf-16-be").hex().upper().encode("ascii") in pdf
    assert "/Count 1" not in content
    y_positions = [int(value) for value in re.findall(r"1 0 0 1 \d+ (-?\d+) Tm", content)]
    assert y_positions and min(y_positions) > 0
    for page_stream in re.findall(rb"stream\n(.*?)\nendstream", pdf, re.DOTALL):
        page_y_positions = [
            int(value)
            for value in re.findall(rb"1 0 0 1 \d+ (-?\d+) Tm", page_stream)
        ]
        assert len(page_y_positions) == len(set(page_y_positions))


def test_invoice_workbook_neutralizes_formula_like_external_text():
    invoice = Invoice(
        invoice_no="=HYPERLINK(\"https://evil.invalid\",\"open\")",
        customer_id="CUST-FORMULA",
        customer_name="+Customer",
        contact_name="@Contact",
        delivery_address="-Address",
        invoice_date=date(2026, 7, 15),
        currency="USD",
        total_amount=Decimal("1"),
    )
    invoice.items.extend([
        InvoiceItem(
            product_kind="hair",
            product_name="=CMD()",
            product_display="+Hair",
            model="@Model",
            color="-Color",
            quantity=1,
            price_per_piece=Decimal("1"),
            discount_amount=Decimal("0"),
            total_price=Decimal("1"),
        ),
        InvoiceItem(
            product_kind="accessory",
            product_name="=Accessory",
            product_display="=Accessory",
            model="+Tool",
            color="@Black",
            quantity=1,
            price_per_piece=Decimal("1"),
            discount_amount=Decimal("0"),
            total_price=Decimal("1"),
        ),
    ])

    workbook = load_workbook(export_service.build_invoice_workbook(invoice), data_only=False)
    sheet = workbook["Invoice"]
    external_cells = ["B3", "B4", "B7", "A11", "B11", "E11", "F11", "A15", "B15", "C15"]
    for coordinate in external_cells:
        cell = sheet[coordinate]
        assert cell.data_type != "f"
        assert str(cell.value).startswith("'")

# ── 录入页自动填充 ────────────────────────────────────────────


def _header_payload(**overrides):
    payload = dict(
        customer_id="123456",
        customer_name="Customer A",
        invoice_date=date(2026, 7, 13),
        items=[],
    )
    payload.update(overrides)
    return InvoiceCreate(**payload)


def test_customer_contact_defaults_latest_snapshot(db):
    # 老单带联系信息
    old = service.create_invoice(db, _header_payload(
        contact_name="Alice", contact_phone="111", delivery_address="Old Addr",
    ), user_id=5)
    # 最新一单联系信息更新过 → 以它为准（同刻建单靠 id 倒序）
    new = service.create_invoice(db, _header_payload(
        contact_name="Alice Wang", contact_email="a@x.com", delivery_address="New Addr",
    ), user_id=6)
    # 最新但全空的单不参与（否则一张没填联系人的单会把历史信息洗掉）
    service.create_invoice(db, _header_payload(), user_id=5)
    # 其他客户的单不串
    service.create_invoice(db, _header_payload(
        customer_id="999", contact_name="Bob",
    ), user_id=5)
    db.flush()
    # 显式拉开建单时间，排序断言不依赖时钟刻度
    old.created_at = datetime(2026, 3, 1)
    new.created_at = datetime(2026, 7, 1)
    db.flush()

    defaults = service.get_customer_contact_defaults(db, "123456")
    assert defaults["contact_name"] == "Alice Wang"
    assert defaults["contact_email"] == "a@x.com"
    assert defaults["delivery_address"] == "New Addr"

    # 排序键是 created_at：老单被系统触碰（推单/校验 bump updated_at）不得顶掉新单
    old.updated_at = datetime(2026, 12, 31)
    db.flush()
    assert service.get_customer_contact_defaults(db, "123456")["contact_name"] == "Alice Wang"

    assert service.get_customer_contact_defaults(db, "no-such") == {"has_xiaoman_orders": False}


def test_okki_flags_stored_and_null_recomputes_on_update(db):
    from app.invoice.schemas import InvoiceUpdate

    # 创建时三标记 None → 服务端兜底后落库（无运费→包邮是；无历史→新成交是）
    invoice = service.create_invoice(db, _header_payload(), user_id=1)
    db.flush()
    assert (invoice.okki_new_deal, invoice.okki_free_shipping, invoice.okki_first_return) == (1, 1, 0)

    # 更新传 null=自动：运费改 50 → 包邮重算为否；显式值优先不重算
    update = InvoiceUpdate(**{**_header_payload().model_dump(), "shipping_fee": 50, "okki_new_deal": 0})
    invoice = service.update_invoice(db, invoice, update, user_id=1)
    db.flush()
    assert invoice.okki_free_shipping == 0  # null → 按运费重算
    assert invoice.okki_new_deal == 0       # 显式值保留


def test_create_invoice_salesperson_fallback(db):
    from app.auth.models import ArkUser

    user = ArkUser(username="zhang", password_hash="x", real_name="张三",
                   phone="13800000000", email="zhang@leshine.com")
    db.add(user)
    db.flush()

    # 前端没带业务员信息 → 按创建人兜底
    invoice = service.create_invoice(db, _header_payload(), user_id=user.id)
    db.flush()
    assert invoice.sales_user_name == "zhang"
    assert invoice.sales_phone == "13800000000"
    assert invoice.sales_email == "zhang@leshine.com"

    # 前端带了值 → 尊重传入，只补空位
    invoice2 = service.create_invoice(db, _header_payload(
        sales_user_name="李四",
    ), user_id=user.id)
    db.flush()
    assert invoice2.sales_user_name == "李四"
    assert invoice2.sales_phone == "13800000000"


# ── OKKI push settings ────────────────────────────────────────

def _base_settings_kwargs(**overrides):
    kwargs = dict(
        generic_product_no=None,
        generic_sku_id=None,
        default_order_status=None,
        default_currency="USD",
        access_token=None,
        token_expires_at=None,
        user_id=1,
    )
    kwargs.update(overrides)
    return kwargs


def test_xiaoman_settings_defaults_and_token_semantics(db):
    # 未初始化时返回默认结构，不隐式建行
    assert xiaoman_service.get_settings_row(db) is None
    default = xiaoman_service.serialize_settings(None)
    assert default["has_token"] is False
    assert default["default_currency"] == "USD"

    # 写入 token → 只回掩码；手动覆盖的过期时间不信表单残值，按刚签发 8h 计
    from datetime import timedelta

    expires = datetime(2026, 12, 31, 23, 59, 59)
    row = xiaoman_service.update_settings(db, **_base_settings_kwargs(
        access_token="tok_1234567890abcdef", token_expires_at=expires,
        default_order_status="待确认",
    ))
    db.commit()
    data = xiaoman_service.serialize_settings(row)
    assert data["has_token"] is True
    assert data["access_token_masked"] == "tok_****cdef"
    assert "tok_1234567890abcdef" not in str(data)
    assert row.token_expires_at > datetime.utcnow() + timedelta(hours=7)

    # access_token=None → 保持不变
    row = xiaoman_service.update_settings(db, **_base_settings_kwargs(
        access_token=None, token_expires_at=expires,
    ))
    db.commit()
    assert row.access_token == "tok_1234567890abcdef"

    # access_token="" → 清除，过期时间联动清空
    row = xiaoman_service.update_settings(db, **_base_settings_kwargs(
        access_token="", token_expires_at=expires,
    ))
    db.commit()
    assert row.access_token is None
    assert row.token_expires_at is None

    # 始终单行
    assert db.query(XiaomanSettings).count() == 1


def test_xiaoman_settings_generic_product_resolution(db):
    _seed_products(db)

    # 唯一 SKU 自动关联
    row = xiaoman_service.update_settings(db, **_base_settings_kwargs(generic_product_no="P001"))
    db.commit()
    assert row.generic_product_id == 1
    assert row.generic_sku_id == 9001

    # 不存在的产品编号拒绝保存
    with pytest.raises(ValueError):
        xiaoman_service.update_settings(db, **_base_settings_kwargs(generic_product_no="NOPE"))

    # SKU 不属于该产品拒绝保存
    with pytest.raises(ValueError):
        xiaoman_service.update_settings(db, **_base_settings_kwargs(
            generic_product_no="P001", generic_sku_id=9002,
        ))

    # 多 SKU 且未指定时不猜测
    db.execute(text("INSERT INTO lsordertest.okki_inventory (product_id, sku_id, disable_flag) VALUES (1, 9010, 0)"))
    db.commit()
    row = xiaoman_service.update_settings(db, **_base_settings_kwargs(generic_product_no="P001"))
    db.commit()
    assert row.generic_sku_id is None
    resolved = xiaoman_service.resolve_generic_product(db, "P001")
    assert resolved["skus"] == [9001, 9010]

    # 指定合法 SKU 通过
    row = xiaoman_service.update_settings(db, **_base_settings_kwargs(
        generic_product_no="P001", generic_sku_id=9010,
    ))
    db.commit()
    assert row.generic_sku_id == 9010

    # 清空产品编号 → 关联 ID 全部清空
    row = xiaoman_service.update_settings(db, **_base_settings_kwargs(generic_product_no="  "))
    db.commit()
    assert row.generic_product_no is None
    assert row.generic_product_id is None
    assert row.generic_sku_id is None
