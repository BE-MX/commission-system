"""PM Hub AI 差异管线测试：本地 diff 正确性 + 降级策略 + 看门狗回收。

AI 调用一律 mock——管线纪律是「本地先算精确 diff，AI 只转述」，
测试锁死的是 diff 的精确性与各降级分支，不锁模型输出。
"""

import pytest

from app.pm import diff_service, material_service
from app.pm.models import PmMaterial, PmMaterialVersion, PmProject, bj_now


@pytest.fixture
def pm_env(db, tmp_path, monkeypatch):
    monkeypatch.setattr("app.pm.service.PM_STORAGE_ROOT", tmp_path)
    project = PmProject(name="测试项目", code="alibaba-ai-agent")
    db.add(project)
    db.flush()
    material = PmMaterial(
        project_id=project.id, list_no=2, name="价格体系", category="产品与报价",
        importance="required", phase=1, delivery_type="file",
    )
    db.add(material)
    db.commit()
    captured: dict = {}

    def fake_chat(_db, preset, messages, caller_module, caller_user_id=None):
        captured["preset"] = preset
        captured["messages"] = messages
        return {"content": "· 新增：全色号样品包价格（$45/套）\n\n总评：小幅扩充"}

    monkeypatch.setattr("app.ai.service.chat", fake_chat)
    return {"project": project, "material": material, "captured": captured}


def _upload(db, material, name, content, note=None):
    return material_service.upload_version(db, material, "liang.xz26", name, content, None, note)


class TestTextDiff:
    def test_done_with_summary_and_prompt_carries_diff_not_fulltext(self, db, pm_env):
        m = pm_env["material"]
        _upload(db, m, "报价.md", b"# Price\nL1: 8%\nL2: 10%\n")
        v2 = _upload(db, m, "报价.md", "# Price\nL1: 8-12%\nL2: 10%\nSample: $45\n".encode())
        diff_service.generate_diff_summary(db, v2.id)
        db.refresh(v2)
        assert v2.diff_status == "done"
        assert "样品包" in v2.diff_summary or "新增" in v2.diff_summary
        # 喂给模型的是 diff（含 +/- 行），不是两份全文
        user_msg = pm_env["captured"]["messages"][0]["content"]
        assert "-L1: 8%" in user_msg and "+L1: 8-12%" in user_msg
        assert pm_env["captured"]["preset"] == "pm_diff"

    def test_identical_content_short_circuits(self, db, pm_env):
        m = pm_env["material"]
        _upload(db, m, "a.txt", b"same content")
        v2 = _upload(db, m, "a.txt", b"same content")
        diff_service.generate_diff_summary(db, v2.id)
        db.refresh(v2)
        assert v2.diff_status == "done"
        assert "完全一致" in v2.diff_summary
        assert "messages" not in pm_env["captured"]  # 没调 AI

    def test_v1_not_applicable(self, db, pm_env):
        v1 = _upload(db, pm_env["material"], "a.txt", b"first")
        diff_service.generate_diff_summary(db, v1.id)
        db.refresh(v1)
        assert v1.diff_status == "not_applicable"

    def test_unsupported_type_not_applicable(self, db, pm_env):
        _upload(db, pm_env["material"], "a.zip", b"PK\x03\x04")
        v2 = _upload(db, pm_env["material"], "b.zip", b"PK\x05\x06")
        diff_service.generate_diff_summary(db, v2.id)
        db.refresh(v2)
        assert v2.diff_status == "not_applicable"
        assert v2.diff_error == "该类型不支持对比"

    def test_ai_failure_marks_failed_and_keeps_version(self, db, pm_env, monkeypatch):
        def boom(*_a, **_kw):
            raise RuntimeError("provider 502")

        monkeypatch.setattr("app.ai.service.chat", boom)
        m = pm_env["material"]
        _upload(db, m, "a.txt", b"old")
        v2 = _upload(db, m, "a.txt", b"new")
        diff_service.generate_diff_summary(db, v2.id)  # 不抛出
        db.refresh(v2)
        assert v2.diff_status == "failed"
        assert "502" in v2.diff_error
        assert v2.file_path  # 版本本身完好


class TestXlsxDiff:
    def test_cell_level_changes(self, db, pm_env):
        import openpyxl
        from app.pm.service import to_abs

        m = pm_env["material"]
        v1 = _upload(db, m, "价格.xlsx", b"placeholder")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "L1价格"
        ws["A1"] = "SKU"
        ws["B1"] = "价格"
        ws["A2"] = "LS-001"
        ws["B2"] = 100
        wb.save(to_abs(v1.file_path))

        v2 = _upload(db, m, "价格.xlsx", b"placeholder2")
        wb2 = openpyxl.Workbook()
        ws2 = wb2.active
        ws2.title = "L1价格"
        ws2["A1"] = "SKU"
        ws2["B1"] = "价格"
        ws2["A2"] = "LS-001"
        ws2["B2"] = 120  # 修改
        ws2["A3"] = "LS-002"
        ws2["B3"] = 45  # 新增行
        wb2.save(to_abs(v2.file_path))

        # 直接测本地 diff（不依赖 AI）
        diff_text, truncated = diff_service.diff_xlsx(to_abs(v1.file_path), to_abs(v2.file_path))
        assert not truncated
        assert "100 → 120" in diff_text
        assert "LS-002" in diff_text

        diff_service.generate_diff_summary(db, v2.id)
        db.refresh(v2)
        assert v2.diff_status == "done"


class TestWatchdog:
    def test_stale_pending_recovered(self, db, pm_env, monkeypatch):
        from datetime import timedelta
        m = pm_env["material"]
        _upload(db, m, "a.txt", b"1")
        v2 = _upload(db, m, "a.txt", b"2")
        assert v2.diff_status == "pending"
        # 把 created_at 拨到阈值之前，模拟服务重启遗留的在途任务
        v2.created_at = bj_now() - timedelta(seconds=diff_service.STALE_DIFF_SECS + 10)
        db.commit()
        v2_id = v2.id  # recover 会 close() 会话使实例 detach——先取普通值
        monkeypatch.setattr(diff_service, "SessionLocal", lambda: db)
        diff_service.recover_stale_diffs()
        # recover 内部会 close() 测试会话——重新查询而非 refresh
        v2 = db.query(PmMaterialVersion).filter(PmMaterialVersion.id == v2_id).one()
        assert v2.diff_status == "failed"
        assert "回收" in v2.diff_error

    def test_fresh_pending_not_touched(self, db, pm_env, monkeypatch):
        m = pm_env["material"]
        _upload(db, m, "a.txt", b"1")
        v2 = _upload(db, m, "a.txt", b"2")
        v2_id = v2.id
        monkeypatch.setattr(diff_service, "SessionLocal", lambda: db)
        diff_service.recover_stale_diffs()
        v2 = db.query(PmMaterialVersion).filter(PmMaterialVersion.id == v2_id).one()
        assert v2.diff_status == "pending"
